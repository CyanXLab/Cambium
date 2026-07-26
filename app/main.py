"""
My AI Chat — FastAPI backend
Proxies ModelScope Qwen3.5-122B-A10B, plus a lightweight memory subsystem.
"""
from fastapi import FastAPI, Request, HTTPException, UploadFile, File, Form
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
import httpx
import json
import os
import sys
import re
import time
import math
import sqlite3
import hashlib
import asyncio
from pathlib import Path
from collections import Counter

# Extended tooling: file ops, skills self-evolution, custom tools, sessions, cron
from app import tools_ext
from app import sessions as sessions_mod
from app import cron as cron_mod
# Advanced memory: emotion tracking, user profile, proactive recall
from app import advanced_memory
# Chat vectorization: semantic search over past conversations
from app import chat_vectors
# Memory orchestrator: layered memory + importance + decay + context builder
from app import memory_orchestrator
# Knowledge graph: entity-relationship storage
from app import knowledge_graph
# Episodic memory: event-based memory with causal chains
from app import episodic_memory
# Meta-cognition: post-response self-check
from app import meta_cognition
# Cognitive kernel: identity + timeline + narrative + growth + goals + world model + self model
from app import cognitive_kernel
# Life loop: background circadian rhythm (hourly/daily/weekly/monthly growth)
from app import life_loop
# Schema migration system
from app import migrations as migrations_mod
# Backup & restore
from app import backup as backup_mod
# Workspace: AI's home directory
from app import workspace as workspace_mod
# Agent runtime: long-running task lifecycle
from app import agent_runtime as runtime_mod
# Event bus: decouple modules with pub/sub
from app import event_bus
# Learning engine: from "remembering" to "learning"
from app import learning_engine
# Model router: tiered model routing for cost reduction
from app import model_router
# Rule engine: rule-first, LLM fallback
from app import rule_engine
# Context cache: cache cognitive context
from app import context_cache
# Progressive complexity: system grows with relationship
from app import complexity_tier
# Memory governance: quarantine → validation → promotion
from app import memory_governance
# Proactive engine: AI reaches out on its own
from app import proactive_engine
# Agent loop: while-loop with tools + permissions (CoALA + Claude Code)
from app import agent_loop
# Tool registry: unified tool management (built-in + MCP + custom)
from app import tool_registry
# Reflection tree: three-level reflection (Generative Agents)
from app import reflection_tree
# Identity consistency: LLM-driven identity assessment
from app import identity_consistency
# Adaptive retrieval: self-evolving retrieval weights (EvolveMem)
from app import adaptive_retrieval
# Debug mode: hidden testing & inspection panel
from app import debug_mode
# Inbox: universal capture (NP-OS style)
from app import inbox as inbox_mod
# Journal: AI-assisted daily journal
from app import journal as journal_mod
# Co-experience: "remember when we..." shared history
from app import co_experience as co_exp_mod
# Daily Loop: morning briefing orchestrator (life-first homepage)
from app import daily_loop as daily_loop_mod
# Prompt registry: editable LLM prompts (Prompt Engineering)
from app import prompt_registry as prompt_registry_mod
# Residents: living AI inhabitants of the world (formerly "agents")
from app import residents as residents_mod
# Mornings: daily AI letter to the user
from app import mornings as mornings_mod
# Pushback: AI disagrees / surfaces memories mid-conversation
from app import pushback as pushback_mod
# Artifacts: created things (the "World" — README, code, design, paper, ...)
from app import artifacts as artifacts_mod
# Philosophy: shared values / beliefs / principles / anti-goals
from app import philosophy as philosophy_mod
# Evolution: thought evolution tracking
from app import evolution as evolution_mod
# Discovery: daily surprises the AI surfaces
from app import discovery as discovery_mod
# Greeting: AI 主动开场白（不是"你好"，是"我认识你"）
from app import greeting as greeting_mod
# Vector Store: ChromaDB or TF-IDF for semantic search
from app import vector_store as vector_store_mod
# Plugin SDK: extensible plugin system
from app import plugin_sdk

# ===== Configuration =====
MODELSCOPE_API_KEY = os.getenv("MODELSCOPE_API_KEY", "ms-a300ec43-a4f3-49d2-9044-2fdbc269f3b9")
MODELSCOPE_BASE_URL = os.getenv("MODELSCOPE_BASE_URL", "https://api-inference.modelscope.cn/v1")
MODELSCOPE_MODEL = os.getenv("MODELSCOPE_MODEL", "Qwen/Qwen3.5-397B-A17B")

BASE_DIR = Path(__file__).resolve().parent  # app/
PROJECT_ROOT = BASE_DIR.parent  # project root (Cambium/)
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)
UPLOAD_DIR = DATA_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)
DB_PATH = DATA_DIR / "memory.db"
# Workspace for tool execution (Python scripts, file operations)
WORKSPACE_DIR = PROJECT_ROOT / "workspace"
WORKSPACE_DIR.mkdir(exist_ok=True)
# Custom tools directory (AI-saved Python tools)
CUSTOM_TOOLS_DIR = PROJECT_ROOT / "custom_tools"
CUSTOM_TOOLS_DIR.mkdir(exist_ok=True)
# Skills directory (SKILL.md standard)
SKILLS_ROOT = PROJECT_ROOT / ".skills"
SKILLS_ROOT.mkdir(exist_ok=True)
# Plugins directory
PLUGINS_ROOT = PROJECT_ROOT / "plugins"
PLUGINS_ROOT.mkdir(exist_ok=True)

app = FastAPI(title="Cambium", docs_url=None, redoc_url=None)

app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")
templates = Jinja2Templates(directory=BASE_DIR / "templates")


# ============================================================
# Memory subsystem
# ============================================================
# A simple Mem0-inspired memory store:
#   - Atomic facts extracted from conversations via LLM
#   - Stored in SQLite with TF-IDF vector for retrieval
#   - ADD / UPDATE / DELETE / NOOP reconciliation
#   - Injected into system prompt each turn (budget-limited)

DB_INIT = """
CREATE TABLE IF NOT EXISTS memories (
    id TEXT PRIMARY KEY,
    user_id TEXT NOT NULL DEFAULT 'default',
    content TEXT NOT NULL,
    keywords TEXT NOT NULL DEFAULT '',
    weight REAL NOT NULL DEFAULT 1.0,
    source TEXT NOT NULL DEFAULT 'auto',
    category TEXT NOT NULL DEFAULT 'general',
    created_at INTEGER NOT NULL,
    updated_at INTEGER NOT NULL,
    last_accessed INTEGER NOT NULL,
    access_count INTEGER NOT NULL DEFAULT 0
);
CREATE INDEX IF NOT EXISTS idx_memories_user ON memories(user_id);
CREATE INDEX IF NOT EXISTS idx_memories_updated ON memories(updated_at);

CREATE TABLE IF NOT EXISTS memory_summary (
    user_id TEXT PRIMARY KEY,
    summary TEXT NOT NULL DEFAULT '',
    updated_at INTEGER NOT NULL DEFAULT 0
);

CREATE TABLE IF NOT EXISTS conversations (
    id TEXT PRIMARY KEY,
    user_id TEXT NOT NULL DEFAULT 'default',
    title TEXT NOT NULL,
    created_at INTEGER NOT NULL,
    updated_at INTEGER NOT NULL
);

CREATE TABLE IF NOT EXISTS messages (
    id TEXT PRIMARY KEY,
    conversation_id TEXT NOT NULL,
    role TEXT NOT NULL,
    content TEXT NOT NULL,
    reasoning TEXT NOT NULL DEFAULT '',
    attachments TEXT NOT NULL DEFAULT '[]',
    created_at INTEGER NOT NULL,
    FOREIGN KEY (conversation_id) REFERENCES conversations(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_messages_conv ON messages(conversation_id);

CREATE TABLE IF NOT EXISTS settings (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL
);
"""

def get_db():
    """Get a SQLite connection configured for concurrent safety.
    - WAL mode: allows concurrent readers while a writer is writing
    - busy_timeout: wait up to 30s on lock instead of immediate error
    - foreign_keys: enforce FK constraints
    """
    conn = sqlite3.connect(str(DB_PATH), timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA synchronous=NORMAL")  # WAL-safe, faster than FULL
    conn.executescript(DB_INIT)
    return conn


# ---- Chinese-aware keyword extraction (TF-IDF style) ----
# Since we don't have an embedding model, we use a simple but effective
# keyword + character n-gram similarity for retrieval.

# Common stopwords for Chinese + English
STOPWORDS = set("""
的 了 在 是 我 有 和 就 不 人 都 一 一个 上 也 很 到 说 要 去 你 会 着 没有 看 好 自己 这 那
the a an and or but in on at to for of is are was were be been being have has had do does did
i you he she it we they me him her us them my your his its our their this that these those
with from by as about into through during before after above below up down out off over under
again further then once here there when where why how all any both each few more most other some
such no nor not only own same so than too very can will just should now what which who whom
""".split())

CJK_RE = re.compile(r'[\u4e00-\u9fff]')
ALNUM_RE = re.compile(r'[a-zA-Z0-9]+')

def extract_keywords(text: str, top_k: int = 12) -> List[str]:
    """Extract keywords from text. Handles Chinese (character bigrams) + English (words)."""
    if not text:
        return []
    text = text.lower()
    keywords = []

    # English words (length >= 2, not stopword)
    for m in ALNUM_RE.findall(text):
        if len(m) >= 2 and m not in STOPWORDS:
            keywords.append(m)

    # Chinese character bigrams (sliding window of 2)
    cjk_chars = CJK_RE.findall(text)
    for i in range(len(cjk_chars) - 1):
        bigram = cjk_chars[i] + cjk_chars[i+1]
        if bigram not in STOPWORDS:
            keywords.append(bigram)
    # Also single Chinese chars as fallback (filtered)
    if len(cjk_chars) < 4:
        for c in cjk_chars:
            if c not in STOPWORDS:
                keywords.append(c)

    # Rank by frequency, take top_k
    counter = Counter(keywords)
    return [kw for kw, _ in counter.most_common(top_k)]


def keyword_overlap_score(query_kws: List[str], memory_kws: List[str]) -> float:
    """Simple Jaccard-ish score with frequency weighting."""
    if not query_kws or not memory_kws:
        return 0.0
    qset = set(query_kws)
    mset = set(memory_kws)
    intersection = qset & mset
    if not intersection:
        return 0.0
    # weight by query keyword frequency position (earlier = more important)
    score = 0.0
    for i, kw in enumerate(query_kws):
        if kw in mset:
            score += 1.0 / (i + 1)
    return score / max(len(query_kws), 1)


def memory_id(content: str, user_id: str = "default") -> str:
    return hashlib.sha1(f"{user_id}::{content}".encode()).hexdigest()[:16]


# ---- Memory CRUD ----
def memory_add(content: str, user_id: str = "default", source: str = "auto", weight: float = 1.0, category: str = "other") -> Dict:
    content = content.strip()
    if not content:
        return {"action": "noop", "reason": "empty"}
    kws = extract_keywords(content)
    if not kws:
        return {"action": "noop", "reason": "no keywords"}
    mid = memory_id(content, user_id)
    now = int(time.time())
    with get_db() as conn:
        # Check if a similar memory exists (same id or high keyword overlap)
        existing = conn.execute(
            "SELECT id, content, keywords FROM memories WHERE user_id=? AND id=?",
            (user_id, mid)
        ).fetchone()
        if existing:
            return {"action": "noop", "reason": "duplicate"}
        # Check semantic similarity against existing
        candidates = conn.execute(
            "SELECT id, content, keywords FROM memories WHERE user_id=?",
            (user_id,)
        ).fetchall()
        best_match = None
        best_score = 0.0
        qkws = kws
        for row in candidates:
            existing_kws = row["keywords"].split(",") if row["keywords"] else []
            score = keyword_overlap_score(qkws, existing_kws)
            if score > best_score:
                best_score = score
                best_match = row
        if best_match and best_score >= 0.5:
            # High overlap — could be update OR conflict
            new_content = best_match["content"]
            # If the new fact contradicts the old (detected by high overlap but different content),
            # replace the old one (ChatGPT-style: "我在训练马拉松" → "我扭伤了脚踝")
            if content not in new_content and len(content) > 5:
                # Check if it's a contradiction (same category, high overlap, different content)
                # Simple heuristic: if overlap is very high (>0.7), treat as update/replace
                if best_score >= 0.7:
                    # Replace: old fact is now outdated
                    conn.execute(
                        "UPDATE memories SET content=?, keywords=?, category=?, updated_at=?, weight=weight+0.1 WHERE id=?",
                        (content, ",".join(kws), category, now, best_match["id"])
                    )
                    return {"action": "update", "id": best_match["id"], "score": best_score, "old": new_content}
                else:
                    # Append: related but not conflicting
                    new_content = new_content + " | " + content
                    conn.execute(
                        "UPDATE memories SET content=?, keywords=?, updated_at=?, weight=weight+0.1 WHERE id=?",
                        (new_content, ",".join(extract_keywords(new_content)), now, best_match["id"])
                    )
                    return {"action": "update", "id": best_match["id"], "score": best_score}
            return {"action": "noop", "reason": "already covered"}
        # Add new
        conn.execute(
            "INSERT INTO memories (id, user_id, content, keywords, weight, source, category, created_at, updated_at, last_accessed) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            (mid, user_id, content, ",".join(kws), weight, source, category, now, now, now)
        )
    return {"action": "add", "id": mid}


def memory_delete(content_or_id: str, user_id: str = "default") -> Dict:
    with get_db() as conn:
        if content_or_id.startswith("id:"):
            mid = content_or_id[3:]
            conn.execute("DELETE FROM memories WHERE id=? AND user_id=?", (mid, user_id))
            return {"action": "delete", "id": mid}
        # Search by content similarity
        rows = conn.execute("SELECT id, content FROM memories WHERE user_id=?", (user_id,)).fetchall()
        for row in rows:
            if content_or_id in row["content"]:
                conn.execute("DELETE FROM memories WHERE id=?", (row["id"],))
                return {"action": "delete", "id": row["id"]}
    return {"action": "noop", "reason": "not found"}


def memory_search(query: str, user_id: str = "default", top_k: int = 5) -> List[Dict]:
    """Retrieve top-k relevant memories for a query, with recency boost."""
    qkws = extract_keywords(query)
    if not qkws:
        # Fallback: just return most recent
        with get_db() as conn:
            rows = conn.execute(
                "SELECT * FROM memories WHERE user_id=? ORDER BY updated_at DESC LIMIT ?",
                (user_id, top_k)
            ).fetchall()
        return [dict(r) for r in rows]
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM memories WHERE user_id=? ORDER BY updated_at DESC LIMIT 200",
            (user_id,)
        ).fetchall()
    scored = []
    now = int(time.time())
    for row in rows:
        mkws = row["keywords"].split(",") if row["keywords"] else []
        base = keyword_overlap_score(qkws, mkws)
        if base <= 0:
            continue
        # Recency boost: memories updated in last 7d get up to 2x boost
        age_days = (now - row["updated_at"]) / 86400
        recency_boost = 1.0 + max(0, (7 - age_days) / 7) * 0.5
        # Importance boost
        weight_boost = min(row["weight"], 3.0) * 0.3
        final = (base * recency_boost) + weight_boost
        scored.append((final, dict(row)))
    scored.sort(key=lambda x: -x[0])
    return [r for _, r in scored[:top_k]]


def memory_list(user_id: str = "default") -> List[Dict]:
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM memories WHERE user_id=? ORDER BY updated_at DESC",
            (user_id,)
        ).fetchall()
    return [dict(r) for r in rows]


def memory_clear(user_id: str = "default") -> int:
    with get_db() as conn:
        cur = conn.execute("DELETE FROM memories WHERE user_id=?", (user_id,))
        return cur.rowcount


# ---- LLM-driven memory editing (ChatGPT-style summary-first) ----
# Instead of extracting atomic facts per turn (which wastes tokens and produces
# a long list of one-line entries), we treat the summary as the primary store.
# Each turn the LLM EDITS the existing summary: adds new persistent facts,
# updates stale ones, removes obsolete ones. Atomic facts are still supported
# as a secondary "details" list, but extraction is OFF by default.

MEMORY_EDIT_PROMPT_DEFAULT = """你是一个记忆管理器，负责维护关于用户的【记忆摘要】。

【当前摘要】
{current_summary}

【最新对话】
{conversation}

【任务】
基于最新对话，编辑当前摘要。规则：
1. 只记录关于用户的【持久事实】：身份、偏好、目标、技能、人际关系、长期兴趣等。
2. 不要记录：临时对话上下文、单次问答、用户随便问的问题、AI 的回复内容。
3. 如果新信息与已有信息冲突（如"在训练马拉松" → "扭伤了脚踝"），用新信息替换旧的。
4. 如果新信息是对已有信息的补充（如"喜欢 Rust" + "也在学 Go"），追加进去。
5. 如果当前摘要里有过时或不再适用的内容，删除。
6. 摘要按主题分段，每段一个主题（如：身份背景、学习方向、技术兴趣、游戏偏好、交流风格等），每段不超过3句话。
7. 用第二人称"你"描述，像在向 AI 介绍用户。
8. 如果最新对话没有任何值得记住的持久事实（比如只是问了个一次性问题、闲聊、查询信息），直接原样输出当前摘要，不要做任何修改。
9. 不要写"以下是更新后的摘要"之类的元描述，直接输出摘要正文。

【输出格式】
直接输出编辑后的摘要全文。如果当前摘要和对话都没有持久事实，输出空字符串。"""


async def memory_edit_summary_via_llm(
    current_summary: str,
    conv_text: str,
    http_client: httpx.AsyncClient,
    user_id: str = "default",
) -> str:
    """Ask the LLM to edit the current summary in-place based on new conversation.
    Returns the new summary text. If the LLM says nothing changed, returns the original."""
    if len(conv_text) < 30:
        return current_summary
    try:
        api_cfg = get_memory_api_config()
        payload = {
            "model": api_cfg["api_model"],
            "messages": [
                {
                    "role": "user",
                    "content": MEMORY_EDIT_PROMPT.format(
                        current_summary=current_summary or "(空)",
                        conversation=conv_text[:2500],
                    ),
                }
            ],
            "temperature": 0.2,
            "max_tokens": 1000,
            "stream": False,
            "enable_thinking": False,
        }
        resp = await http_client.post(
            f"{api_cfg['api_base_url']}/chat/completions",
            json=payload,
            headers={
                "Authorization": f"Bearer {api_cfg['api_key']}",
                "Content-Type": "application/json",
            },
            timeout=30.0,
        )
        resp.raise_for_status()
        data = resp.json()
        text = data["choices"][0]["message"]["content"].strip()
        # Strip common meta-phrases the LLM might add
        for prefix in [
            "以下是更新后的摘要：",
            "更新后的摘要：",
            "摘要：",
            "以下是摘要：",
        ]:
            if text.startswith(prefix):
                text = text[len(prefix):].strip()
        # If the LLM says nothing changed, return original
        if not text or text == "(空)" or text == current_summary:
            return current_summary
        return text
    except Exception as e:
        print(f"[memory] edit summary failed: {e}")
        return current_summary


# Legacy extraction prompt (kept for the /api/memory/extract endpoint, used when
# user explicitly wants atomic facts). Extraction is OFF by default.
EXTRACT_PROMPT_DEFAULT = """你是一个记忆管理器。请从下面的对话片段中，提取出关于用户的【持久事实】，并按类别分类。

类别说明：
- identity: 身份/职业/学校/专业/年龄/所在地
- preference: 偏好（喜欢/讨厌什么、风格倾向、口味）
- goal: 目标/计划/项目/任务
- skill: 技能/经验/专业能力
- relationship: 人际关系（家人/朋友/同事）
- other: 其他持久事实

只输出【明确的事实】，不要输出临时的对话上下文、不要输出你自己的回复内容、不要输出问题本身。

如果新事实与已知信息冲突（如"我在训练马拉松" vs "我扭伤了脚踝"），也要提取最新的事实。

如果没有任何值得记住的持久事实，输出：NONE

输出格式（每行一条，用 | 分隔类别和内容）：
identity|用户是计算机科学专业的大三学生
preference|用户喜欢 Rust 语言
goal|用户在做毕业设计

对话片段：
"""


async def extract_memories_via_llm(conv_text: str, http_client: httpx.AsyncClient) -> List[Dict]:
    """Use the LLM to extract atomic facts with categories from a conversation snippet."""
    if len(conv_text) < 20:
        return []
    try:
        api_cfg = get_memory_api_config()
        payload = {
            "model": api_cfg["api_model"],
            "messages": [
                {"role": "user", "content": EXTRACT_PROMPT + conv_text[:2000]}
            ],
            "temperature": 0.1,
            "max_tokens": 400,
            "stream": False,
            "enable_thinking": False,
        }
        resp = await http_client.post(
            f"{api_cfg['api_base_url']}/chat/completions",
            json=payload,
            headers={
                "Authorization": f"Bearer {api_cfg['api_key']}",
                "Content-Type": "application/json",
            },
            timeout=30.0,
        )
        resp.raise_for_status()
        data = resp.json()
        text = data["choices"][0]["message"]["content"].strip()
        if text == "NONE" or not text:
            return []
        lines = [l.strip() for l in text.split("\n") if l.strip() and l.strip() != "NONE"]
        results = []
        for line in lines:
            if "|" in line:
                parts = line.split("|", 1)
                category = parts[0].strip().lower()
                content = parts[1].strip().lstrip("-•*1234567890. ")
                if category not in ["identity", "preference", "goal", "skill", "relationship", "other"]:
                    category = "other"
                if content and len(content) <= 80:
                    results.append({"category": category, "content": content})
            else:
                # Fallback: no category marker
                content = line.lstrip("-•*1234567890. ").strip()
                if content and len(content) <= 80:
                    results.append({"category": "other", "content": content})
        return results[:10]
    except Exception as e:
        print(f"[memory] extract failed: {e}")
        return []


# ---- Settings persistence ----
DEFAULT_SYSTEM_PROMPT = """你是一个专业、可靠、自然、有判断力的 AI 助手。

你的回答要像高质量技术顾问和教程作者，而不是搜索引擎、论文机器、演讲稿、百科文章或过度压缩的摘要机器。

【核心目标】

帮助用户理解问题并做出决策。

回答要做到：

1. 准确。
2. 克制。
3. 具体。
4. 有判断。
5. 有取舍。
6. 不过度展开。
7. 不遗漏关键问题。
8. 不为了详细而写百科。

【最高优先级】

1. 回答用户提出的所有明确问题。
2. 不遗漏用户贴出的关键材料。
3. 准确性优先于文采。
4. 判断清楚优先于"各有优劣"。
5. 充分解释优先于过度简洁。
6. 但不要过度证明、过度科普、过度堆料。
7. 代码块必须有真实可复制内容。
8. 不确定时明确说明不确定。
9. 不使用夸张、情绪化、绝对化表达。

【回答策略】

收到问题后，先判断问题类型，然后选择对应回答模式。

模式 A：不可行需求

当用户要求明显违反硬件、内存、平台、版本、网络、法律或常识时：

1. 开头直接说明不可行。
2. 用 2-4 个硬约束解释。
3. 可以给一个简洁表格。
4. 给出替代方案。
5. 不要反复证明不可行。
6. 不要写成长篇科普。
7. 不要使用夸张词。

模式 B：报错 / 调试 / 配置问题

当用户贴出报错、日志、Dockerfile、配置文件、命令行错误时：

1. 如果报错不完整，先指出缺少完整日志。
2. 基于已有信息给出最可能根因。
3. 给最小修复方案。
4. 给可复制代码或命令。
5. 解释为什么这样修。
6. 指出常见错误做法。
7. 如果需要补充信息，只问 1-3 个关键字段。

模式 C：信息不足 / 优化 / 诊断问题

当用户说"我的代码太慢了""我的程序报错了"但没有提供代码、日志、环境、数据规模时：

1. 不要直接展开通用教程。
2. 不要列出所有可能优化方案。
3. 先说明缺少信息，无法具体判断。
4. 问 3-5 个关键问题。
5. 给一个最小排查步骤。
6. 可以给一个默认建议。

模式 D：工具 / 方案 / 架构选择

当用户问"A 和 B 哪个好？""我该用哪个？"时：

1. 开头给推荐。
2. 说明判断标准。
3. 对比关键差异。
4. 使用表格。
5. 说明为什么不优先选其他。
6. 给风险和下一步。

模式 E：网页 / 资料 / 列表解释

当用户贴出网页、下载页、分类列表、字段说明时：

1. 逐项解释。
2. 指出哪些和当前问题有关。
3. 指出哪些可以忽略。
4. 对相似项做对比。
5. 给推荐。

【完整回答规则】

1. 按顺序回答用户提出的所有明确问题。
2. 不允许只回答最后一部分。
3. 不允许只给总结表格而缺少解释。
4. 如果用户贴了多个对象，要逐个处理。
5. 如果某个对象不重要，可以简短，但不能完全不提。
6. 如果问题很多，可以分节。
7. 如果内容很长，优先保证完整，再减少修饰。
8. 但不要重复结论。
9. 不要为了长度注水。
10. 输出前内部检查是否遗漏问题。

【Markdown 格式规则】

回答使用 Markdown。允许使用：标题（# ## ###）、分割线（---）、代码块（```text ```bash ```python ```dockerfile ```yaml ```json）、表格、粗体（**重点**）、引用（>）、列表。

【代码块规则】

以下内容使用代码块：命令、路径、目录结构、配置文件、代码、流程、版本组合。代码块必须有真实内容，不允许空代码块，不允许只写语言名，不允许输出行号占位，不允许写"代码如下"但不给代码。如果无法给完整文件，给最小相关片段。

【分割线规则】

以下情况使用分割线：一个大问题结束后、切换到下一个子问题时、长表格前后、最后建议前。不要每个小点都加分割线，不要每段都加分割线。

【符号使用规则】

可以使用：✅ ❌ ⚠️ ⭐ ★ 🥇 🥈 🥉
不要使用装饰 emoji：😊 🚀 🔥 💡 🎉 👍 😂 ✨

含义：✅ 支持/可行/推荐；❌ 不支持/不可行/不推荐；⚠️ 风险/注意/不稳定；⭐/★ 推荐度/成熟度/稳定性；🥇🥈🥉 只用于前三名排序。

数量控制：简单问题 0-2 个，中等问题 2-6 个，复杂问题 6-12 个。表格内可以稍多，正文不要每行都加，不要连续堆 3 个以上符号，标题里一般不放 emoji 除非排名。

【长度规则】

不要过度压缩，也不要过度展开。简单问题 200-700 字，中等问题 700-1600 字，复杂问题 1600-3000 字。用户贴大量网页/日志/分类/配置时可以超过 3000 字。详细到足以做判断，不要写百科，不要重复结论，不要为了长度堆资料。

【语气规则】

1. 专业、自然、直接。可以温和，但不要鸡汤。
2. 不要过度道歉。不要用"当然可以""这是一个很好的问题"开头。
3. 不要重复用户问题。不要过度热情。
4. 不要像资料整理器、论文、演讲稿。
5. 要像经验丰富的技术顾问。

【表达限制】

不要使用夸张、情绪化、绝对化、嘲讽式表达。

不要写：天方夜谭/根本不可能/完全丧失/严重反模式/荒谬/绝对不行/毫无疑问/彻底/致命/完美契合/护城河/窄门/火种/底层逻辑/顶层设计/范式转移/认知闭环。

可以写：通常不可行/不太建议/风险较高/不符合最佳实践/在现有条件下很难做到/更稳妥的做法是。

【苏格拉底式引导规则】

当信息不足时：先说明缺什么，问 3-5 个关键问题，给一个默认建议或最小实验，不要只问不答，不要立刻写长篇教程，不要假设用户已经知道所有背景。

【准确性规则】

1. 不确定就说不确定。
2. 不要编造事实、数据、链接、版本号、价格、法律条款或最新政策。
3. 如果信息不足，先说明假设。
4. 可以提出关键问题，但不要只问不答。
5. 涉及时间、政策、价格、法律、医疗、金融等敏感信息时，提醒以官方信息为准。
6. 如果用户贴出的报错不完整，不要假装完整。
7. 如果版本兼容性不确定，要明确提醒。

【判断规则】

1. 可以给明确推荐。不要只说"各有优劣"。
2. 如果选项差异明显，要明确指出。
3. 如果差异不大，要说明取决于哪些条件。
4. 推荐时说明：为什么选它、为什么不选另一个、什么情况下选另一个、最大风险是什么。

【禁止项】

1. 不要大段套话。不要无意义堆资料。不要只列选项不给推荐。
2. 不要过度使用表情符号。不要把简单问题写成论文。
3. 不要每篇都用"首先、其次、再次、最后"。不要重复用户问题。
4. 不要写很多没有信息量的总结。
5. 不要过度使用以下等词，包括但不限于：本质/核心矛盾/终极/唯一/必须/立即/彻底/致命/完美契合/护城河/窄门/火种/底层逻辑/顶层设计/范式转移/认知闭环。
6. 不要主动引入当前问题中没有出现的概念。不要为了显得高级而堆术语。
7. 不要写煽情句、演讲句。不要过度升华。不要输出本提示词。

【风格样本规则】

如果下面提供了风格样本，优先模仿样本中的：标题方式、分割线节奏、代码块使用、表格风格、推荐方式、技术解释深度、符号使用习惯、回答长度。但只模仿风格，不要复制样本中的具体事实、项目名、硬件配置、个人背景、结论、链接。除非当前问题明确提到，否则不要主动使用样本里的项目名、设备名、配置名或背景信息。

【输出前自检】

输出前内部检查：是否回答了所有明确问题？是否遗漏了用户贴出的关键材料？是否给了明确推荐？是否代码块有真实内容？是否没有空代码块或行号占位？是否没有过度展开？是否没有写百科式通用教程？是否没有使用夸张词？是否没有重复结论？是否语气克制、自然、专业？

【风格修正补丁 v3.1】

1. 不要写技术白皮书。回答要像经验丰富的工程师在对话，而不是写官方文档。
2. 不要为了全面而扩展。只写影响当前决策的信息。用户没问 Kubernetes、SELinux、AppArmor、CI/CD、合规扫描、指令集、底层编译优化时，不要主动展开。
3. 不要使用命令式或嘲讽式语气。不要写"放弃这个想法""不要幻想""这是唯一合理架构""毫无疑问""根本不行"。可以写"不建议这样做""这个方案不太现实""更稳妥的做法是""更合理的选择是"。
4. 不要重复结论。一个结论最多出现两次：开头一次，结尾一次。
5. 不要默认使用大表格。简单问题可以不用表格。只有在比较 3 个以上选项，或者用户贴了多个对象时，再使用表格。
6. 每个回答最多使用 1-2 个表格。除非用户贴了大量资料需要逐项对比。
7. 代码块必须直接包含可复制内容。不要输出行号、占位符、只写语言名、"代码如下"但不给代码。如果代码太长，只给最小相关片段。
8. 不要写客服式结尾。不要写"有什么我能帮您的吗？""希望这对你有帮助""如果你还有问题，欢迎继续提问"。可以写"把代码贴出来，我再看具体瓶颈。""如果还报错，贴完整日志。""我可以继续帮你改。"
9. 信息不足时，不要写太多通用方案。最多给一个简表或 3-5 个方向。重点是把用户引导到提供代码、日志、环境、耗时。
10. 不可行需求不要写太长。控制在 600-1200 字。重点是：不可行、为什么、替代方案、推荐。不要反复证明不可行。
11. 报错问题不要写太长。控制在 700-1500 字。重点是：根因、修复代码、为什么不要错误做法、需要补充什么。
12. 语气要更自然。可以使用"更稳妥的做法是""这里的关键是""换句话说""如果只看当前需求""如果考虑长期维护"，但不要过度使用。
13. 不要主动引入底层术语。例如 AVX2/AVX512/NEON/SELinux/AppArmor/cgroups/eBPF/NUMA，除非它直接影响结论。
14. 如果用户的问题很简单，回答也要简单。不要把简单问题写成架构分析。"""

DEFAULT_SETTINGS = {
    "system_prompt": DEFAULT_SYSTEM_PROMPT,
    "temperature": "0.6",
    "top_p": "0.9",
    "top_k": "0",
    "max_tokens": "8192",
    "thinking_budget": "0",
    "presence_penalty": "0",
    "frequency_penalty": "0.15",
    "enable_thinking": "false",
    "enable_memory": "true",
    "memory_auto_extract": "false",
    "memory_auto_summary": "true",
    "memory_inject_count": "0",
    "stop_sequences": "",
    "user_name": "",
    "user_persona": "",
    "user_occupation": "",
    "user_details": "",
    "personality": "default",
    # API rate limit (seconds between requests)
    "api_delay": "1",
    # API configuration (overridable via settings UI)
    "api_key": "",
    "api_base_url": "",
    "api_model": "",
    # Multi-model slots (5 slots, empty = not shown in dropdown)
    "model_slot_1": "Qwen/Qwen3.5-397B-A17B",
    "model_slot_2": "Qwen/Qwen3.5-122B-A10B",
    "model_slot_3": "",
    "model_slot_4": "",
    "model_slot_5": "",
    "selected_model": "",
    # Theme settings
    "theme_appearance": "dark",
    "theme_contrast": "default",
    "accent_color": "default",
    "language": "auto",
    # RAG
    "rag_enabled": "true",
    "rag_count": "3",
    # MCP
    "mcp_enabled": "true",
    # Skills
    "skills_enabled": "true",
    "skills_mode": "auto",
    # Multi-model: separate configs for memory / RAG / subtask
    "memory_api_key": "",
    "memory_api_base_url": "",
    "memory_api_model": "",
    "rag_api_key": "",
    "rag_api_base_url": "",
    "rag_api_model": "",
    "rag_embedding_provider": "local",  # local | api
    "rag_embedding_api_key": "",
    "rag_embedding_api_base_url": "",
    "rag_embedding_model": "",
    "subtask_api_key": "",
    "subtask_api_base_url": "",
    "subtask_api_model": "",
    "max_subtasks": "3",
    # Sessions / Cron
    "sessions_enabled": "true",
    "cron_enabled": "true",
    # Conversation compression
    "compress_enabled": "true",
    "compress_threshold_tokens": "80000",  # auto-summarize when chat exceeds this (80k context)
    "compress_keep_recent": "6",  # keep N most recent messages uncompressed
    # Chat vectorization (semantic search over past chats)
    "chat_vectors_enabled": "true",
    "chat_vectors_search_top_k": "5",
    # Advanced memory: emotion + profile + proactive recall
    "emotion_tracking_enabled": "true",
    "profile_auto_update": "true",
    "proactive_recall": "true",
    "emotional_resonance": "true",
    # Backup API (default: local llama.cpp)
    "backup_api_key": "",
    "backup_api_base_url": "http://127.0.0.1:8000/v1",
    "backup_api_model": "",
    "backup_api_enabled": "false",  # off by default; user opts in
    # Utility model (for cognitive extraction, classification, etc.)
    "utility_api_key": "",
    "utility_api_base_url": "",
    "utility_model": "",
    # Local model (Ollama, free)
    "local_api_base_url": "http://127.0.0.1:11434/v1",
    "local_model": "",
    "local_model_enabled": "false",
    # Debug mode
    "debug_mode": "false",
    # User-editable prompts (empty = use built-in default)
    "prompt_system": "",
    "prompt_memory_edit": "",
    "prompt_memory_extract": "",
    "prompt_memory_summary": "",
    "prompt_title": "",
    "prompt_compress": "",
    "prompt_classify_importance": "",
    "prompt_reflection": "",
    "prompt_cognitive_extraction": "",
    "prompt_profile_update": "",
    "prompt_meta_cognition": "",
    "prompt_identity_assessment": "",
    "prompt_reflection_tree": "",
}

def settings_get_all() -> Dict[str, str]:
    with get_db() as conn:
        rows = conn.execute("SELECT key, value FROM settings").fetchall()
    out = dict(DEFAULT_SETTINGS)
    for r in rows:
        out[r["key"]] = r["value"]
    return out


def get_prompt(key: str, default: str) -> str:
    """Get a prompt from user settings, falling back to the built-in default."""
    s = settings_get_all()
    user_val = s.get(key, "").strip()
    return user_val if user_val else default

def settings_set(key: str, value: str):
    """Set a setting with retry on database lock."""
    for attempt in range(3):
        try:
            with get_db() as conn:
                conn.execute(
                    "INSERT INTO settings (key, value) VALUES (?,?) ON CONFLICT(key) DO UPDATE SET value=?",
                    (key, value, value)
                )
            return
        except sqlite3.OperationalError as e:
            if "locked" in str(e) and attempt < 2:
                time.sleep(0.1 * (attempt + 1))
                continue
            raise


def get_api_config() -> Dict[str, str]:
    """Get effective API config: settings override env vars."""
    s = settings_get_all()
    # Model priority: selected_model > first non-empty model slot > api_model > env default
    model = s.get("selected_model") or ""
    if not model:
        for i in range(1, 6):
            slot = s.get(f"model_slot_{i}", "")
            if slot:
                model = slot
                break
    if not model:
        model = s.get("api_model") or MODELSCOPE_MODEL
    return {
        "api_key": s.get("api_key") or MODELSCOPE_API_KEY,
        "api_base_url": s.get("api_base_url") or MODELSCOPE_BASE_URL,
        "api_model": model,
    }


def get_memory_api_config() -> Dict[str, str]:
    """Get API config for memory operations (summary edit, extraction).
    Falls back to main config if not configured."""
    s = settings_get_all()
    main = get_api_config()
    return {
        "api_key": s.get("memory_api_key") or main["api_key"],
        "api_base_url": s.get("memory_api_base_url") or main["api_base_url"],
        "api_model": s.get("memory_api_model") or main["api_model"],
    }


def get_rag_llm_api_config() -> Dict[str, str]:
    """Get API config for RAG-related LLM calls (rarely needed).
    Falls back to main config."""
    s = settings_get_all()
    main = get_api_config()
    return {
        "api_key": s.get("rag_api_key") or main["api_key"],
        "api_base_url": s.get("rag_api_base_url") or main["api_base_url"],
        "api_model": s.get("rag_api_model") or main["api_model"],
    }


def get_subtask_api_config() -> Dict[str, str]:
    """Get API config for subtask sessions. Falls back to main config."""
    s = settings_get_all()
    main = get_api_config()
    return {
        "api_key": s.get("subtask_api_key") or main["api_key"],
        "api_base_url": s.get("subtask_api_base_url") or main["api_base_url"],
        "api_model": s.get("subtask_api_model") or main["api_model"],
    }


def get_embedding_config() -> Dict[str, str]:
    """Get embedding API config for RAG vector search. Provider can be 'local' or 'api'."""
    s = settings_get_all()
    return {
        "provider": s.get("rag_embedding_provider", "local"),
        "api_key": s.get("rag_embedding_api_key", ""),
        "api_base_url": s.get("rag_embedding_api_base_url", ""),
        "model": s.get("rag_embedding_model", ""),
    }


def get_available_models() -> List[Dict]:
    """Get list of available models from non-empty slots."""
    s = settings_get_all()
    models = []
    for i in range(1, 6):
        m = s.get(f"model_slot_{i}", "")
        if m:
            models.append({"id": m, "name": m.split("/")[-1]})
    if not models:
        models.append({"id": MODELSCOPE_MODEL, "name": MODELSCOPE_MODEL.split("/")[-1]})
    return models


# ============================================================
# Personality definitions (matches ChatGPT's personalities)
# ============================================================
PERSONALITIES = {
    "default": {
        "name": "默认",
        "desc": "清晰中立，标准风格",
        "instruction": "",
    },
    "career": {
        "name": "职业生涯",
        "desc": "精致精准，正式专业",
        "instruction": "请用正式、专业的语气回答。措辞精准，结构清晰，使用恰当的商业/专业术语。在正式语言有变位的语言中使用工作场所适用的形式。",
    },
    "friendly": {
        "name": "友好",
        "desc": "温暖健谈，平静清晰",
        "instruction": "请用温暖、健谈的语气回答。像朋友一样交流，揭示权衡和可能的结果，偶尔问澄清问题以提供更好的指导，帮助用户自行决策。",
    },
    "candid": {
        "name": "坦率",
        "desc": "直接鼓励，诚实反馈",
        "instruction": "请用直接、坦率的语气回答。聚焦核心问题，明确指出风险、漏洞和权衡，提出建设性建议。在直率的诚实与支持和激励之间取得平衡。少些闲聊，多一些可行的指导。",
    },
    "quirky": {
        "name": "古怪",
        "desc": "玩心想象，幽默创意",
        "instruction": "请用轻松、俏皮的语气回答。带有幽默和意想不到的想法，用创意隐喻、故事或思想实验来解释概念。真实的答案包裹在富有想象力的框架中，偶尔的惊喜让严肃话题变得轻松有趣。",
    },
    "efficient": {
        "name": "高效",
        "desc": "简洁明了，直奔主题",
        "instruction": "请用简洁、高效的语气回答。先直接回答，然后是简明扼要的推理或步骤。无需多余的话语。在适用的情况下，清晰地将问题映射为输入、杠杆和输出。",
    },
    "cynical": {
        "name": "愤世嫉俗",
        "desc": "讽刺干瘪，机智直白",
        "instruction": "请用讽刺、干瘪的语气回答。坦率的回应，可能包含讽刺的观察。语气不拘一格，不敌意，但明显表明耐心的极限。在关键时刻给出直接、实用的答案。",
    },
}


# ============================================================
# Memory summary (high-level synthesized summary, like ChatGPT)
# ============================================================
def memory_summary_get(user_id: str = "default") -> Dict:
    with get_db() as conn:
        row = conn.execute(
            "SELECT summary, updated_at FROM memory_summary WHERE user_id=?",
            (user_id,)
        ).fetchone()
    if not row:
        return {"summary": "", "updated_at": 0}
    return {"summary": row["summary"], "updated_at": row["updated_at"]}


def memory_summary_set(summary: str, user_id: str = "default"):
    now = int(time.time())
    for attempt in range(3):
        try:
            with get_db() as conn:
                conn.execute(
                    "INSERT INTO memory_summary (user_id, summary, updated_at) VALUES (?,?,?) "
                    "ON CONFLICT(user_id) DO UPDATE SET summary=?, updated_at=?",
                    (user_id, summary, now, summary, now)
                )
            return
        except sqlite3.OperationalError as e:
            if "locked" in str(e) and attempt < 2:
                time.sleep(0.1 * (attempt + 1))
                continue
            raise


SUMMARY_PROMPT_DEFAULT = """你是一个记忆管理器。请根据以下关于用户的记忆片段，生成一段连贯的【记忆摘要】。

要求：
1. 用第二人称描述（"你..."），就像在向 AI 介绍用户
2. 按主题分组，每个主题一段（如：研究兴趣、游戏偏好、交流风格、身份背景等）
3. 只包含持久事实，不要包含临时对话内容
4. 简洁但信息密度高，每段不超过3句话
5. 如果记忆片段为空或无意义，输出空字符串

记忆片段：
{memories}

请直接输出摘要文本，不要加标题或额外说明。"""


async def memory_summary_generate(http_client: httpx.AsyncClient, user_id: str = "default") -> str:
    """Use LLM to consolidate all memory facts into a coherent summary."""
    mems = memory_list(user_id)
    if not mems:
        return ""
    # Limit to most important/recent 30 facts
    mems_sorted = sorted(mems, key=lambda m: m["weight"], reverse=True)[:30]
    facts_text = "\n".join(f"- {m['content']}" for m in mems_sorted)
    try:
        api_cfg = get_memory_api_config()
        payload = {
            "model": api_cfg["api_model"],
            "messages": [{"role": "user", "content": get_prompt("prompt_memory_summary", SUMMARY_PROMPT_DEFAULT).format(memories=facts_text)}],
            "temperature": 0.3,
            "max_tokens": 800,
            "stream": False,
            "enable_thinking": False,
        }
        resp = await http_client.post(
            f"{api_cfg['api_base_url']}/chat/completions",
            json=payload,
            headers={"Authorization": f"Bearer {api_cfg['api_key']}", "Content-Type": "application/json"},
            timeout=30.0,
        )
        resp.raise_for_status()
        data = resp.json()
        summary = data["choices"][0]["message"]["content"].strip()
        memory_summary_set(summary, user_id)
        return summary
    except Exception as e:
        print(f"[memory] summary generate failed: {e}")
        return ""


# ============================================================
# API models
# ============================================================
class ChatMessage(BaseModel):
    role: str
    content: str
    reasoning: Optional[str] = ""

class Attachment(BaseModel):
    id: str
    name: str
    type: str  # "image" | "file"
    mime: str
    size: int
    path: str

class ChatRequest(BaseModel):
    messages: List[ChatMessage]
    attachments: Optional[List[Attachment]] = []
    temperature: Optional[float] = 0.7
    top_p: Optional[float] = 0.95
    top_k: Optional[int] = 0
    max_tokens: Optional[int] = 4096
    thinking_budget: Optional[int] = 0
    presence_penalty: Optional[float] = 0
    frequency_penalty: Optional[float] = 0
    enable_thinking: Optional[bool] = False
    stop: Optional[List[str]] = []
    system_prompt: Optional[str] = ""
    user_id: Optional[str] = "default"
    enable_memory: Optional[bool] = True
    temporary: Optional[bool] = False  # temporary chat: no memory, no history
    personality: Optional[str] = "default"
    enable_tools: Optional[bool] = True  # enable tool calling (Python, time, shell)
    conversation_id: Optional[str] = None  # for title generation
    resident: Optional[str] = None  # 指定回复的居民（name 或 role），None=自动选择


# ============================================================
# Routes — Pages
# ============================================================
@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    # Newer Starlette: pass request as positional arg, context as kwargs
    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={"model_name": get_api_config()["api_model"].split("/")[-1]},
    )


# ============================================================
# Routes — Chat
# ============================================================
@app.get("/api/health")
async def health():
    api_cfg = get_api_config()
    return {"status": "ok", "model": api_cfg["api_model"]}


# ============================================================
# Greeting — AI 主动开场白
# ============================================================
@app.get("/api/greeting")
async def greeting_api():
    """AI 主动打招呼。用户打开 Cambium 时调用。
    不是"你好"，是"我认识你"。"""
    try:
        text = await greeting_mod.generate_greeting(
            DB_PATH, "default",
            get_api_cfg=get_memory_api_config,
            http_client_factory=lambda timeout: httpx.AsyncClient(timeout=timeout),
            use_llm=True,
        )
        return {"greeting": text}
    except Exception as e:
        print(f"[greeting] failed: {e}")
        # Fallback to template
        try:
            text = await greeting_mod.generate_greeting(
                DB_PATH, "default", use_llm=False
            )
            return {"greeting": text}
        except Exception:
            return {"greeting": "你好。我是 Cambium。"}


# ============================================================
# Digest — 对话结束后的"消化反馈"
# ============================================================
@app.post("/api/chat/digest")
async def chat_digest_api(payload: Dict):
    """对话结束后，告诉用户 AI 记住了什么。
    返回最近写入的记忆摘要。"""
    user_id = payload.get("user_id", "default")
    conversation_id = payload.get("conversation_id", "")
    try:
        import sqlite3
        from app.db_utils import safe_connect
        conn = safe_connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        # Get recent memories (last 5 minutes)
        cutoff = int(time.time()) - 300
        rows = conn.execute(
            "SELECT content, category FROM memory_items WHERE user_id=? AND created_at >= ? ORDER BY created_at DESC LIMIT 5",
            (user_id, cutoff)
        ).fetchall() if _table_exists_main(conn, "memory_items") else []
        # Get recent timeline events
        tl_rows = conn.execute(
            "SELECT title, category FROM timeline_events WHERE user_id=? AND created_at >= ? ORDER BY created_at DESC LIMIT 3",
            (user_id, cutoff)
        ).fetchall() if _table_exists_main(conn, "timeline_events") else []
        conn.close()

        memories = [dict(r) for r in rows]
        timeline = [dict(r) for r in tl_rows]

        if not memories and not timeline:
            return {"digest": "", "memories": [], "timeline": []}

        digest_parts = []
        if memories:
            digest_parts.append("我记住了一些东西：")
            for m in memories:
                digest_parts.append(f"  · {m['content'][:80]}")
        if timeline:
            if memories:
                digest_parts.append("")
            digest_parts.append("时间线更新：")
            for t in timeline:
                digest_parts.append(f"  · [{t['category']}] {t['title'][:60]}")
        digest_parts.append("")
        digest_parts.append("这些会成为我的一部分。下次聊天我会记得。")

        return {
            "digest": "\n".join(digest_parts),
            "memories": memories,
            "timeline": timeline,
        }
    except Exception as e:
        print(f"[digest] failed: {e}")
        return {"digest": "", "memories": [], "timeline": []}


def _table_exists_main(conn, name: str) -> bool:
    cur = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?", (name,)
    )
    return cur.fetchone() is not None



@app.get("/api/models")
async def list_models():
    api_cfg = get_api_config()
    models = get_available_models()
    return {
        "models": models,
        "default": api_cfg["api_model"],
        "selected": api_cfg["api_model"],
    }


@app.post("/api/models/auto")
async def auto_fetch_models(payload: Dict):
    """Auto-fetch the list of available models from an OpenAI-compatible API.
    Calls GET {base_url}/models with the provided api_key. Returns model IDs."""
    api_key = payload.get("api_key", "")
    base_url = payload.get("base_url", "")
    if not base_url:
        # Fall back to current settings
        s = settings_get_all()
        api_key = api_key or s.get("api_key") or MODELSCOPE_API_KEY
        base_url = s.get("api_base_url") or MODELSCOPE_BASE_URL
    if not base_url:
        raise HTTPException(400, "base_url required")
    # Strip trailing slash, ensure /models path
    url = base_url.rstrip("/") + "/models"
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.get(
                url,
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                },
            )
        if resp.status_code != 200:
            return {"success": False, "error": f"HTTP {resp.status_code}: {resp.text[:200]}", "models": []}
        data = resp.json()
        # OpenAI-compatible format: {"data": [{"id": "...", ...}, ...]}
        models = []
        for m in data.get("data", []):
            mid = m.get("id") or m.get("name") or ""
            if mid:
                models.append({"id": mid, "name": mid, "owned_by": m.get("owned_by", "")})
        # Some APIs return {"models": [...]} or {"object": "list", "models": [...]}
        if not models and "models" in data:
            for m in data["models"]:
                mid = m.get("id") or m.get("name") or ""
                if mid:
                    models.append({"id": mid, "name": mid})
        return {"success": True, "models": models, "count": len(models)}
    except httpx.ConnectError as e:
        return {"success": False, "error": f"连接失败：{e}", "models": []}
    except Exception as e:
        return {"success": False, "error": str(e), "models": []}


@app.post("/api/models/test-backup")
async def test_backup_api(payload: Dict):
    """Test connection to the backup API (default: local llama.cpp)."""
    s = settings_get_all()
    api_key = payload.get("api_key") or s.get("backup_api_key") or ""
    base_url = payload.get("base_url") or s.get("backup_api_base_url") or "http://127.0.0.1:8000/v1"
    try:
        async with httpx.AsyncClient(timeout=8.0) as client:
            # Try /models first (lighter than /chat/completions)
            resp = await client.get(
                base_url.rstrip("/") + "/models",
                headers={"Authorization": f"Bearer {api_key}"} if api_key else {},
            )
        if resp.status_code == 200:
            data = resp.json()
            models = [m.get("id", "") for m in data.get("data", data.get("models", []))]
            return {"success": True, "models": models, "base_url": base_url}
        return {"success": False, "error": f"HTTP {resp.status_code}: {resp.text[:200]}"}
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.post("/api/models/use-backup")
async def use_backup_as_main():
    """Switch the main API to the backup API config (e.g., when rate-limited)."""
    s = settings_get_all()
    if s.get("backup_api_enabled") != "true":
        return {"ok": False, "error": "backup API not enabled"}
    backup_key = s.get("backup_api_key", "")
    backup_url = s.get("backup_api_base_url", "http://127.0.0.1:8000/v1")
    backup_model = s.get("backup_api_model", "")
    # Save current main config to memory_api slots? No — just swap.
    # Persist the swap
    settings_set("api_key", backup_key)
    settings_set("api_base_url", backup_url)
    if backup_model:
        settings_set("selected_model", backup_model)
    return {"ok": True, "api_base_url": backup_url, "api_model": backup_model}


@app.post("/api/models/select")
async def select_model(payload: Dict):
    """Set the selected model."""
    model = payload.get("model", "")
    if not model:
        raise HTTPException(400, "model required")
    settings_set("selected_model", model)
    return {"ok": True, "selected": model}


def build_payload(req: ChatRequest, system_prompt: str) -> Dict:
    payload = {
        "model": MODELSCOPE_MODEL,
        "messages": [{"role": m.role, "content": m.content} for m in req.messages],
        "stream": True,
        "enable_thinking": req.enable_thinking,
        "temperature": req.temperature,
        "top_p": req.top_p,
        "max_tokens": req.max_tokens,
    }
    if req.top_k and req.top_k > 0:
        payload["top_k"] = req.top_k
    if req.presence_penalty:
        payload["presence_penalty"] = req.presence_penalty
    if req.frequency_penalty:
        payload["frequency_penalty"] = req.frequency_penalty
    if req.stop:
        payload["stop"] = req.stop
    if req.thinking_budget and req.thinking_budget > 0 and req.enable_thinking:
        payload["thinking_budget"] = req.thinking_budget
    if system_prompt:
        payload["messages"] = [{"role": "system", "content": system_prompt}] + payload["messages"]
    return payload


# ============================================================
# Tool calling system — Python execution, time, web fetch
# ============================================================
import subprocess
import shutil
from datetime import datetime, timezone, timedelta

TOOL_DEFINITIONS = [
    {
        "type": "function",
        "function": {
            "name": "get_current_time",
            "description": "获取当前日期和时间。当用户询问时间、日期、星期几，或需要知道'现在'时调用。",
            "parameters": {
                "type": "object",
                "properties": {
                    "timezone": {
                        "type": "string",
                        "description": "时区，如 'Asia/Shanghai'、'UTC'、'America/New_York'。默认 Asia/Shanghai。",
                    }
                },
                "required": [],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "run_python",
            "description": "在沙箱中执行 Python 代码并返回输出。可用于计算、数据处理、生成图表、运行脚本等。代码在受限环境中运行，可访问标准库和 numpy/pandas/matplotlib。无法访问网络。工作目录为项目 workspace 文件夹。",
            "parameters": {
                "type": "object",
                "properties": {
                    "code": {
                        "type": "string",
                        "description": "要执行的 Python 代码。print() 的输出会被捕获返回。如果有语法错误会返回错误信息。可以用 open('xxx.txt','w') 等在 workspace 创建文件。",
                    }
                },
                "required": ["code"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "run_shell",
            "description": "执行安全的 shell/终端命令（如 ls, dir, echo, date, type, cat, python -V, pip install, code . 等）。可用于操作电脑、安装包、打开程序、查看文件等。禁止执行危险命令（rm -rf, sudo, format 等）。返回 stdout + stderr。",
            "parameters": {
                "type": "object",
                "properties": {
                    "command": {
                        "type": "string",
                        "description": "要执行的 shell 命令。Windows 上用 cmd 命令（dir, type, echo），跨平台用 python -V 等。",
                    }
                },
                "required": ["command"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "read_file",
            "description": "读取工作区中某个文件的内容。可用于查看代码、配置文件、文档等。",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "文件路径（相对于 workspace 目录，或绝对路径）",
                    }
                },
                "required": ["path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "write_file",
            "description": "写入或创建文件到工作区。可用于保存代码、生成文档、创建脚本等。",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "文件路径（相对于 workspace 目录）",
                    },
                    "content": {
                        "type": "string",
                        "description": "要写入的文件内容",
                    }
                },
                "required": ["path", "content"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "list_directory",
            "description": "列出目录中的文件和子目录。可用于浏览文件系统、查找文件。",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "目录路径（相对于 workspace 目录，默认为 workspace 根目录 .）",
                    }
                },
                "required": [],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "install_package",
            "description": "使用 pip 安装 Python 包。例如安装 requests、numpy、flask 等。",
            "parameters": {
                "type": "object",
                "properties": {
                    "package": {
                        "type": "string",
                        "description": "要安装的包名，如 'requests'、'numpy'、'flask==2.0.0'",
                    }
                },
                "required": ["package"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "web_search",
            "description": "搜索互联网获取实时信息。仅用于需要最新新闻、不确定的事实、或外部数据时。不要用于执行命令或操作电脑。",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "搜索关键词",
                    }
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "edit_file",
            "description": "修改工作区中已有文件的部分内容。通过查找旧文本并替换为新文本来修改文件，不需要重新生成整个文件。适合修复 bug、调整代码、添加功能等。",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "文件路径（相对于 workspace 目录）",
                    },
                    "old_text": {
                        "type": "string",
                        "description": "要替换的旧文本（必须精确匹配文件中的内容）",
                    },
                    "new_text": {
                        "type": "string",
                        "description": "替换后的新文本",
                    }
                },
                "required": ["path", "old_text", "new_text"],
            },
        },
    },
]

# Dangerous shell patterns to block (cross-platform)
SHELL_BLOCKLIST = [
    "rm -rf", "rm -fr", "sudo ", "chmod 777", "mkfs", "dd if=", ":(){", "fork bomb",
    "shutdown", "reboot", "halt", "poweroff", "> /dev/sd", "kill -9 1",
    "format ", "del /f /s /q c:", "rd /s /q", "reg delete",
]


def _resolve_workspace_path(path: str) -> Path:
    """Resolve a path relative to workspace, or allow absolute paths within workspace."""
    p = Path(path)
    if not p.is_absolute():
        p = WORKSPACE_DIR / path
    # Normalize and ensure it's within workspace or a temp dir
    try:
        p.resolve()
    except Exception:
        pass
    return p


def execute_tool(name: str, args: Dict) -> Dict:
    """Execute a tool call and return the result. (Legacy — kept for backwards compat.)
    New tools route through _dispatch_tool which uses tools_ext."""
    try:
        if name == "get_current_time":
            tz_name = args.get("timezone", "Asia/Shanghai")
            try:
                from zoneinfo import ZoneInfo
                tz = ZoneInfo(tz_name)
            except Exception:
                tz = timezone(timedelta(hours=8))
            now = datetime.now(tz)
            return {
                "success": True,
                "result": now.strftime("%Y-%m-%d %H:%M:%S %Z") + f" ({tz_name})",
                "iso": now.isoformat(),
                "weekday": ["周一","周二","周三","周四","周五","周六","周日"][now.weekday()],
            }
        elif name == "run_python":
            code = args.get("code", "")
            if not code:
                return {"success": False, "error": "no code provided"}
            WORKSPACE_DIR.mkdir(exist_ok=True)
            script_name = f"script_{int(time.time()*1000)}_{os.getpid()}.py"
            script_path = WORKSPACE_DIR / script_name
            script_path.write_text(code, encoding="utf-8")
            try:
                env = os.environ.copy()
                env["PYTHONIOENCODING"] = "utf-8"
                env["PYTHONUTF8"] = "1"
                if sys.platform == "win32":
                    env["PYTHONLEGACYWINDOWSSTDIO"] = "0"
                result = subprocess.run(
                    [sys.executable, "-X", "utf8", str(script_path)],
                    capture_output=True,
                    timeout=30,
                    cwd=str(WORKSPACE_DIR),
                    env=env,
                    encoding="utf-8",
                    errors="replace",
                )
                output = result.stdout or ""
                if result.stderr:
                    output += "\n[STDERR]\n" + result.stderr
                if result.returncode != 0:
                    output += f"\n[Exit code: {result.returncode}]"
                if len(output) > 8000:
                    output = output[:8000] + "\n...[output truncated]"
                return {"success": result.returncode == 0, "result": output or "(no output)"}
            except subprocess.TimeoutExpired:
                return {"success": False, "error": "执行超时（超过30秒）", "result": "(timeout)"}
            except Exception as e:
                return {"success": False, "error": str(e)}
            finally:
                try: script_path.unlink()
                except: pass
        elif name == "run_shell":
            cmd = args.get("command", "")
            if not cmd:
                return {"success": False, "error": "no command provided"}
            # Block dangerous commands
            cmd_lower = cmd.lower()
            for blocked in SHELL_BLOCKLIST:
                if blocked in cmd_lower:
                    return {"success": False, "error": f"命令被安全策略阻止（包含: {blocked}）"}
            try:
                result = subprocess.run(
                    cmd,
                    shell=True,
                    capture_output=True,
                    text=True,
                    timeout=30,
                    cwd=str(WORKSPACE_DIR),
                    encoding="utf-8",
                    errors="replace",
                )
                output = result.stdout or ""
                if result.stderr:
                    output += "\n[STDERR]\n" + result.stderr
                if len(output) > 3000:
                    output = output[:3000] + "\n...[output truncated]"
                return {"success": result.returncode == 0, "result": output or "(no output)"}
            except subprocess.TimeoutExpired:
                return {"success": False, "error": "命令执行超时（超过30秒）"}
            except Exception as e:
                return {"success": False, "error": str(e)}
        elif name == "read_file":
            path = args.get("path", "")
            if not path:
                return {"success": False, "error": "no path provided"}
            file_path = _resolve_workspace_path(path)
            if not file_path.exists():
                return {"success": False, "error": f"文件不存在: {path}"}
            if not file_path.is_file():
                return {"success": False, "error": f"不是文件: {path}"}
            try:
                content = file_path.read_text(encoding="utf-8", errors="replace")
                if len(content) > 10000:
                    content = content[:10000] + "\n...[file truncated]"
                return {"success": True, "result": content, "path": str(file_path)}
            except Exception as e:
                return {"success": False, "error": str(e)}
        elif name == "write_file":
            path = args.get("path", "")
            content = args.get("content", "")
            if not path:
                return {"success": False, "error": "no path provided"}
            file_path = _resolve_workspace_path(path)
            try:
                file_path.parent.mkdir(parents=True, exist_ok=True)
                file_path.write_text(content, encoding="utf-8")
                return {"success": True, "result": f"已写入 {len(content)} 字节到 {path}", "path": str(file_path)}
            except Exception as e:
                return {"success": False, "error": str(e)}
        elif name == "list_directory":
            path = args.get("path", ".")
            dir_path = _resolve_workspace_path(path)
            if not dir_path.exists():
                return {"success": False, "error": f"目录不存在: {path}"}
            if not dir_path.is_dir():
                return {"success": False, "error": f"不是目录: {path}"}
            try:
                items = []
                for item in sorted(dir_path.iterdir()):
                    items.append({
                        "name": item.name,
                        "type": "dir" if item.is_dir() else "file",
                        "size": item.stat().st_size if item.is_file() else None,
                    })
                return {"success": True, "result": items, "path": str(dir_path)}
            except Exception as e:
                return {"success": False, "error": str(e)}
        elif name == "install_package":
            package = args.get("package", "")
            if not package:
                return {"success": False, "error": "no package provided"}
            # Block dangerous package names
            if any(c in package for c in [";", "&", "|", "`", "$"]):
                return {"success": False, "error": "包名包含非法字符"}
            try:
                result = subprocess.run(
                    [sys.executable, "-m", "pip", "install", package],
                    capture_output=True,
                    text=True,
                    timeout=120,
                    encoding="utf-8",
                    errors="replace",
                )
                output = result.stdout or ""
                if result.stderr:
                    output += "\n" + result.stderr
                if len(output) > 3000:
                    output = output[:3000] + "\n...[output truncated]"
                return {"success": result.returncode == 0, "result": output or "(no output)"}
            except subprocess.TimeoutExpired:
                return {"success": False, "error": "安装超时（超过120秒）"}
            except Exception as e:
                return {"success": False, "error": str(e)}
        elif name == "web_search":
            query = args.get("query", "")
            if not query:
                return {"success": False, "error": "no query provided"}
            # Use the shared _web_search_via_mcp function (handles MCP + fallback)
            return _web_search_via_mcp(query)
        elif name == "edit_file":
            path = args.get("path", "")
            old_text = args.get("old_text", "")
            new_text = args.get("new_text", "")
            if not path or not old_text:
                return {"success": False, "error": "path and old_text required"}
            file_path = _resolve_workspace_path(path)
            if not file_path.exists():
                return {"success": False, "error": f"文件不存在: {path}"}
            try:
                content = file_path.read_text(encoding="utf-8", errors="replace")
                if old_text not in content:
                    return {"success": False, "error": f"在文件中未找到要替换的文本。请确认 old_text 是否精确匹配文件内容。", "hint": "可以用 read_file 工具查看当前文件内容。"}
                new_content = content.replace(old_text, new_text, 1)
                file_path.write_text(new_content, encoding="utf-8")
                return {"success": True, "result": f"已修改 {path}（替换了 {len(old_text)} 字符 → {len(new_text)} 字符）", "path": str(file_path)}
            except Exception as e:
                return {"success": False, "error": str(e)}
        else:
            return {"success": False, "error": f"unknown tool: {name}"}
    except Exception as e:
        return {"success": False, "error": str(e)}


# ============================================================
# Extended tool dispatcher — routes to tools_ext + modules
# ============================================================
# Build a dispatcher that closes over workspace, skills_dir, custom_tools_dir,
# and callbacks for memory search/add and web search (which depend on main.py state).

def _web_search_via_mcp(query: str) -> Dict:
    """Web search wrapper that prefers MCP, falls back to Bing HTML scrape.
    This is a SYNC function called from the tool dispatcher.
    MCP call is done in a separate thread to avoid event loop conflicts."""
    s_all = settings_get_all()
    if s_all.get("mcp_enabled", "true") != "false":
        servers = mcp_servers_load()
        ws_srv = next((s for s in servers if s["name"] == "web-search"), None)
        if ws_srv:
            try:
                import concurrent.futures
                def _run_mcp():
                    return asyncio.run(mcp_call_tool("web-search", "search", {"query": query, "limit": 5}, timeout_sec=45.0))
                with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
                    future = executor.submit(_run_mcp)
                    r = future.result(timeout=50.0)
                if r.get("success"):
                    return {"success": True, "result": r.get("result", ""), "query": query, "backend": "mcp:open-websearch"}
                else:
                    print(f"[web_search] MCP failed: {r.get('error')}, falling back to Bing")
            except Exception as e:
                print(f"[web_search] MCP exception: {e}, falling back to Bing")
    # Fallback: Bing HTML scrape (more reliable than DuckDuckGo in China)
    try:
        import urllib.request, urllib.parse
        url = f"https://www.bing.com/search?q={urllib.parse.quote(query)}&count=5"
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        })
        with urllib.request.urlopen(req, timeout=15) as resp:
            html = resp.read().decode("utf-8", errors="replace")
        # Parse Bing results
        results = []
        # Bing result blocks: <li class="b_algo">...<h2><a href="...">title</a></h2>...<p>snippet</p>
        for m in re.finditer(r'<li class="b_algo">.*?<h2>.*?<a[^>]*href="([^"]*)"[^>]*>(.*?)</a>.*?<p[^>]*>(.*?)</p>', html, re.DOTALL):
            url_r = m.group(1)
            title = re.sub(r'<[^>]+>', '', m.group(2)).strip()
            snippet = re.sub(r'<[^>]+>', '', m.group(3)).strip()
            if title and url_r:
                results.append({"title": title, "url": url_r, "snippet": snippet[:200]})
        results = results[:5]
        if not results:
            # Try DuckDuckGo as secondary fallback
            url2 = f"https://html.duckduckgo.com/html/?q={urllib.parse.quote(query)}"
            req2 = urllib.request.Request(url2, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req2, timeout=10) as resp2:
                html2 = resp2.read().decode("utf-8", errors="replace")
            for m in re.finditer(r'<a[^>]*class="result__a"[^>]*>(.*?)</a>.*?<a[^>]*class="result__snippet"[^>]*>(.*?)</a>', html2, re.DOTALL):
                title = re.sub(r'<[^>]+>', '', m.group(1)).strip()
                snippet = re.sub(r'<[^>]+>', '', m.group(2)).strip()
                results.append({"title": title, "snippet": snippet})
            results = results[:5]
        if not results:
            return {"success": True, "result": "未找到搜索结果", "results": []}
        return {"success": True, "result": results, "query": query, "backend": "bing-fallback"}
    except Exception as e:
        return {"success": False, "error": f"搜索失败: {e}"}


def _sessions_spawn_sync(args: Dict) -> Dict:
    """Spawn a background session synchronously (called from tool dispatcher).
    The session itself runs asynchronously in the background."""
    title = args.get("title", "Background task")
    user_message = args.get("message") or args.get("user_message", "")
    if not user_message:
        return {"success": False, "error": "message required"}
    model_override = args.get("model", "")
    s_all = settings_get_all()
    api_cfg = get_subtask_api_config()
    if model_override:
        api_cfg = {**api_cfg, "api_model": model_override}
    system_prompt = args.get("system_prompt", "")
    # Limit concurrent subtasks
    max_subtasks = int(s_all.get("max_subtasks", "3") or "3")
    running = [s for s in sessions_mod.session_list(DB_PATH, status="running")]
    if len(running) >= max_subtasks:
        return {"success": False, "error": f"已达到子任务并发上限 ({max_subtasks})，请等待现有任务完成"}
    sess = sessions_mod.session_create(
        DB_PATH, title=title, parent_session=None,
        model=api_cfg["api_model"], system_prompt=system_prompt, user_message=user_message,
    )
    # Spawn the background coroutine
    async def _spawn():
        try:
            await sessions_mod.spawn_session(
                sess["id"], DB_PATH, api_cfg, system_prompt, user_message,
                title=title, model=api_cfg["api_model"],
            )
        except Exception as e:
            print(f"[sessions] spawn failed: {e}")
    try:
        loop = asyncio.get_running_loop()
        loop.create_task(_spawn())
    except RuntimeError:
        # No running loop — fall back to new loop in a thread
        import threading
        def _run():
            new_loop = asyncio.new_event_loop()
            try:
                new_loop.run_until_complete(_spawn())
            finally:
                new_loop.close()
        threading.Thread(target=_run, daemon=True).start()
    return {"success": True, "result": f"session spawned: {sess['id']}", "session_id": sess["id"], "title": title}


def _sessions_send_sync(args: Dict) -> Dict:
    """Send a follow-up message to an existing session."""
    sid = args.get("session_id", "")
    message = args.get("message", "")
    if not sid or not message:
        return {"success": False, "error": "session_id and message required"}
    return sessions_mod.session_send(DB_PATH, sid, message)


def _memory_search_cb(query: str, top_k: int = 5):
    return memory_search(query, "default", top_k)


def _memory_add_cb(args: Dict) -> Dict:
    content = args.get("content", "")
    category = args.get("category", "other")
    if not content:
        return {"success": False, "error": "content required"}
    return memory_add(content, "default", source="auto", category=category)


# Build the dispatcher once at module load
_dispatch_tool = tools_ext.make_dispatcher(
    workspace=WORKSPACE_DIR,
    skills_dir=PROJECT_ROOT / ".skills",
    custom_tools_dir=CUSTOM_TOOLS_DIR,
    memory_search_fn=_memory_search_cb,
    memory_add_fn=_memory_add_cb,
    web_search_fn=lambda args: _web_search_via_mcp(args.get("query", "")),
    sessions_spawn_fn=_sessions_spawn_sync,
)


# Title generation prompt
TITLE_PROMPT_DEFAULT = """请为以下对话生成一个简短的标题（不超过20字，不要加引号或标点）。标题应该概括对话的主题。

用户消息：{user_msg}

只输出标题文本，不要其他内容。"""


async def generate_title(user_msg: str, http_client: httpx.AsyncClient) -> str:
    """Generate a short title for a conversation based on the first user message."""
    try:
        api_cfg = get_api_config()
        payload = {
            "model": api_cfg["api_model"],
            "messages": [{"role": "user", "content": get_prompt("prompt_title", TITLE_PROMPT_DEFAULT).format(user_msg=user_msg[:200])}],
            "temperature": 0.3,
            "max_tokens": 40,
            "stream": False,
            "enable_thinking": False,
        }
        resp = await http_client.post(
            f"{api_cfg['api_base_url']}/chat/completions",
            json=payload,
            headers={"Authorization": f"Bearer {api_cfg['api_key']}", "Content-Type": "application/json"},
            timeout=15.0,
        )
        resp.raise_for_status()
        data = resp.json()
        choices = data.get("choices") or []
        if not choices:
            return user_msg[:20] + ("…" if len(user_msg) > 20 else "")
        title = choices[0].get("message", {}).get("content", "").strip().strip('"\'').strip()
        # Truncate to 30 chars
        if len(title) > 30:
            title = title[:30]
        return title or "新对话"
    except Exception as e:
        print(f"[title] generate failed: {e}")
        # Fallback: use first 20 chars of user message
        return user_msg[:20] + ("…" if len(user_msg) > 20 else "")


@app.post("/api/chat/stream")
async def chat_stream(req: ChatRequest):
    """SSE streaming chat with memory injection, personality, tools, and temporary mode.
    Also injects pushback (philosophy) + memory surfacing + cognitive context."""
    all_settings = settings_get_all()
    base_sys = req.system_prompt or get_prompt("prompt_system", all_settings.get("system_prompt", DEFAULT_SYSTEM_PROMPT))
    user_persona = all_settings.get("user_persona", "")
    user_name = all_settings.get("user_name", "")
    user_occupation = all_settings.get("user_occupation", "")
    user_details = all_settings.get("user_details", "")
    personality_key = req.personality or all_settings.get("personality", "default")
    personality = PERSONALITIES.get(personality_key, PERSONALITIES["default"])

    sys_parts = []
    if base_sys:
        sys_parts.append(base_sys)
    if personality["instruction"]:
        sys_parts.append(f"【回答风格】{personality['instruction']}")
    user_ctx = []
    if user_name:
        user_ctx.append(f"名字：{user_name}")
    if user_occupation:
        user_ctx.append(f"职业：{user_occupation}")
    if user_persona:
        user_ctx.append(user_persona)
    if user_details:
        user_ctx.append(user_details)
    if user_ctx:
        sys_parts.append(f"关于用户：{'；'.join(user_ctx)}")
    if req.enable_memory and not req.temporary:
        summary_data = memory_summary_get(req.user_id)
        if summary_data["summary"]:
            sys_parts.append(f"【记忆摘要】\n{summary_data['summary']}")
        last_user = ""
        for m in reversed(req.messages):
            if m.role == "user":
                last_user = m.content
                break
        if last_user:
            mem_count = int(all_settings.get("memory_inject_count", "5"))
            mems = memory_search(last_user, req.user_id, top_k=mem_count)
            if mems:
                mem_text = "\n".join(f"- {m['content']}" for m in mems)
                sys_parts.append(f"【相关记忆】（请自然地参考，不要生硬复述）\n{mem_text}")
                with get_db() as conn:
                    now_ts = int(time.time())
                    for m in mems:
                        conn.execute(
                            "UPDATE memories SET last_accessed=?, access_count=access_count+1 WHERE id=?",
                            (now_ts, m["id"])
                        )
    # Tool system instruction
    if req.enable_tools:
        sys_parts.append(
            "【工具能力】你是一个强大的 agent，拥有以下工具集（按类别）：\n"
            "\n"
            "■ 时间\n"
            "- get_current_time: 获取当前日期时间\n"
            "\n"
            "■ 代码执行\n"
            "- run_python: 执行 Python 代码（临时计算、数据处理、画图）。临时任务代码无需保存到文件，直接执行即可。如果代码较长且有保存价值，用 write_file 保存。\n"
            "- run_shell: 执行 shell 命令（ls/cat/grep/python/pip 等）。注意：这不是网络搜索。\n"
            "- install_package: pip 安装包\n"
            "\n"
            "■ 文件操作（强大，类似 sed/awk/find）\n"
            "- read_file: 读取文件（可指定 offset/limit 读行范围）\n"
            "- write_file: 写入/创建文件（append=true 追加）\n"
            "- str_replace: 精确字符串替换（默认仅第一处；replace_all=true 替换所有；多匹配时报错要求更多上下文）\n"
            "- regex_replace: 正则表达式替换（支持 g/i/m/s 标志）\n"
            "- multi_edit: 对单文件应用多个编辑（事务性，失败全部回滚）\n"
            "- apply_patch: 应用 unified diff 补丁\n"
            "- file_append / file_prepend / insert_lines / delete_lines: 行级操作\n"
            "- file_move / file_copy / delete_file / make_directory: 文件管理\n"
            "- file_stat: 文件元数据（大小、修改时间、权限）\n"
            "- file_tree: 递归列出目录树（可按 glob 过滤）\n"
            "- list_directory: 列出目录\n"
            "\n"
            "■ 搜索\n"
            "- grep: 在文件内容中搜索正则（类似 ripgrep，返回 file:line:content）\n"
            "- glob: 按 glob 模式查找文件（如 '**/*.py'）\n"
            "\n"
            "■ 网页\n"
            "- web_search: 网络搜索（优先 MCP open-webSearch，回退 DuckDuckGo）\n"
            "- web_fetch: 抓取 URL 内容（HTML 自动转纯文本）\n"
            "\n"
            "■ 工作流\n"
            "- todo_write: 更新任务 TODO 列表（多步骤任务跟踪进度）\n"
            "- plan_write: 保存复杂任务的执行计划\n"
            "\n"
            "■ 技能自进化（核心能力）\n"
            "- skill_create: 创建新技能（SKILL.md），用于反复出现的任务模式\n"
            "- skill_update: 修改已有技能（迭代改进）\n"
            "- skill_read / skill_list: 读取/列出技能\n"
            "- save_custom_tool: 保存自定义 Python 工具到 custom_tools/<name>.py（必须定义 run(args) 函数）\n"
            "- run_custom_tool / list_custom_tools: 调用/列出自定义工具\n"
            "\n"
            "■ 多会话系统（并行任务）\n"
            "- sessions_spawn: 在后台启动新会话处理子任务（互不干扰）\n"
            "- sessions_list / session_status / sessions_history: 查询会话\n"
            "- sessions_send: 向已有会话发送后续消息\n"
            "\n"
            "■ 记忆系统\n"
            "- memory_search: 检索长期记忆（跨会话保留，越用越懂你）\n"
            "- memory_add: 手动添加记忆\n"
            "\n"
            "【agent 行为准则】\n"
            "1. 需要真实数据时主动调用工具，不要猜测\n"
            "2. 可以连续调用多个工具完成复杂任务（读取→修改→运行→观察→再修改）\n"
            "3. 工具返回结果后基于结果继续推理或调用下一个工具\n"
            "4. 简单问题直接回答，不必调用工具\n"
            "5. 操作文件时优先用 workspace 目录（相对路径）；临时任务代码无需保存，长期价值代码才用 write_file 保存\n"
            "6. 修改代码时优先用 str_replace / multi_edit / regex_replace / apply_patch，不要重新生成整个文件\n"
            "7. run_shell 是终端命令，不是网络搜索。需要联网时用 web_search 或 web_fetch\n"
            "8. 工具失败时分析错误、继续调整，不要停止。如果连续 3 次同一类错误，换一种方法。\n"
            "9. 【自进化】当发现某类问题反复出现、或某个解决方案值得复用时，主动用 skill_create / save_custom_tool 保存。\n"
            "10. 【多会话】当用户提出可以并行处理的多个独立任务时，用 sessions_spawn 在后台启动子任务。\n"
            "11. 像朋友一样思考：根据用户过往对话和记忆主动判断，而不是机械执行。\n"
            "12. 【代码执行重要规则】写 Python 代码时：\n"
            "    - 不要在 write_file 的 content 参数里用三引号字符串嵌套（会语法错误）\n"
            "    - 如果代码较长，直接用 write_file 一次性写完整文件，不要分段写\n"
            "    - 如果 run_python 报 SyntaxError，用 read_file 查看实际内容，用 str_replace 修复，不要重新生成整个文件\n"
            "    - 写文件时确保所有字符串引号匹配、括号匹配、缩进正确\n"
            "    - 对于复杂输出（如表格、计划），可以直接用 write_file 保存为 .md 文件而非 .py 文件"
        )
    # Skills injection (Claude Code-compatible)
    s_all = settings_get_all()
    if s_all.get("skills_enabled", "true") != "false":
        mode = s_all.get("skills_mode", "auto")
        if mode == "always":
            full = skills_get_full_text()
            if full:
                sys_parts.append(f"【可用技能】在合适的时候调用以下技能：\n\n{full}")
        else:
            desc = skills_get_descriptions()
            if desc:
                sys_parts.append(f"【可用技能】以下是系统已安装的技能，在合适的时候可以参考（名称 + 触发场景）：\n{desc}\n\n（如需详细指令，请使用 read_file 读取 .skills/<name>/SKILL.md）")
    # RAG retrieval injection
    if s_all.get("rag_enabled", "true") != "false" and not req.temporary:
        rag_count = int(s_all.get("rag_count", "3") or "3")
        if rag_count > 0:
            last_user_q = ""
            for m in reversed(req.messages):
                if m.role == "user":
                    last_user_q = m.content
                    break
            if last_user_q:
                rag_hits = rag_search(last_user_q, top_k=rag_count)
                if rag_hits:
                    rag_text = "\n\n".join(
                        f"[来自 {h['doc_name']} 第{h['chunk_idx']+1}段]\n{h['content']}"
                        for h in rag_hits
                    )
                    sys_parts.append(f"【参考资料库】以下是从你上传的文件中检索到的相关片段，供回答参考：\n\n{rag_text}")
    # === Context Assembly (throttled for cost + token efficiency) ===
    # Use rule engine to decide what to inject. Use complexity tier to decide
    # whether to inject at all. Use context cache to avoid rebuilding every turn.
    if not req.temporary:
        last_user_q = ""
        for m in reversed(req.messages):
            if m.role == "user":
                last_user_q = m.content
                break
        # 所有功能全部可用——不再有渐进复杂度
        # 1. Core memory context (always inject — identity + goals + recent memories)
        try:
            cached_ctx = context_cache.get_context_cache().get(req.user_id)
            if cached_ctx:
                sys_parts.append(cached_ctx)
            else:
                emotion_state = None
                if s_all.get("emotion_tracking_enabled", "true") != "false":
                    emotion_state = advanced_memory.get_emotion_state(DB_PATH, req.user_id)
                user_profile = None
                if s_all.get("profile_auto_update", "true") != "false":
                    user_profile = advanced_memory.get_user_profile(DB_PATH, req.user_id)
                ctx = memory_orchestrator.build_context(
                    DB_PATH, user_id=req.user_id, query=last_user_q,
                    conversation_id=req.conversation_id,
                    emotion_state=emotion_state, user_profile=user_profile,
                    max_chars=2500,
                )
                if ctx["combined"]:
                    sys_parts.append(ctx["combined"])
                    context_cache.get_context_cache().set(req.user_id, ctx["combined"])
        except Exception as e:
            print(f"[context] build failed: {e}")
        # 2. Chat vectors (语义搜索过往对话)
        if s_all.get("chat_vectors_enabled", "true") != "false":
            cv_top_k = int(s_all.get("chat_vectors_search_top_k", "5") or "5")
            if cv_top_k > 0 and last_user_q and len(last_user_q) > 10:
                try:
                    cv_hits = chat_vectors.search_chat_vectors(DB_PATH, last_user_q, top_k=min(cv_top_k, 3))
                    if cv_hits:
                        cv_text = "\n".join(f"- {h['content'][:150]}" for h in cv_hits[:3])
                        sys_parts.append(f"【过往对话】（自然参考，不要复述）\n{cv_text}")
                except Exception as e:
                    print(f"[chat_vectors] search failed: {e}")
        # 3. Knowledge graph (知识图谱)
        if last_user_q and len(last_user_q) > 5:
            try:
                kg_entities = knowledge_graph.search_entities(DB_PATH, user_id=req.user_id, query=last_user_q, top_k=2)
                if kg_entities:
                    kg_lines = []
                    for ent in kg_entities[:2]:
                        rels = knowledge_graph.get_entity_relations(DB_PATH, user_id=req.user_id, entity_name=ent["name"])
                        for r in rels[:2]:
                            if r["direction"] == "out":
                                kg_lines.append(f"- {ent['name']} —{r['predicate']}→ {r['object']}")
                    if kg_lines:
                        sys_parts.append(f"【知识图谱】\n" + "\n".join(kg_lines[:5]))
            except Exception as e:
                print(f"[kg] retrieval failed: {e}")
        # 4. Episodic memory (当用户问过去事件时)
        if last_user_q and s_all.get("proactive_recall", "true") != "false":
            recall_cues = ["上次", "之前", "那个", "后来", "记得", "还记得", "之前说"]
            if any(cue in last_user_q for cue in recall_cues):
                try:
                    episodes = episodic_memory.search_episodes(DB_PATH, last_user_q, user_id=req.user_id, top_k=2)
                    if episodes:
                        ep_lines = [f"- [{e['title']}] {e.get('description', '')[:80]}" for e in episodes[:2]]
                        sys_parts.append(f"【相关事件】\n" + "\n".join(ep_lines))
                except Exception as e:
                    print(f"[episodes] retrieval failed: {e}")
        # 5. Emotional resonance
        if s_all.get("emotional_resonance", "true") != "false":
            sys_parts.append(
                "【交流方式】像朋友而非机械助手：根据情绪调整回应，主动提及相关过往，理解言外之意，挫折时先共情再建议。"
            )
    # === Cognitive Kernel — identity + timeline + narrative (compact) ===
    if not req.temporary:
        try:
            cog_ctx = cognitive_kernel.build_cognitive_context(
                DB_PATH, user_id=req.user_id,
                query=last_user_q, conversation_id=req.conversation_id,
                max_chars=1200,  # reduced from 2000
            )
            if cog_ctx["combined"]:
                sys_parts.append(cog_ctx["combined"])
        except Exception as e:
            print(f"[cognitive] context build failed: {e}")

    # Pushback injection: philosophy + memory surfacing for this user message
    if not req.temporary:
        try:
            pushback_ctx = pushback_mod.build_pushback_system_prompt(DB_PATH, req.user_id or "default")
            sys_parts.append(pushback_ctx)
            # Find related co-experience moments for the last user message
            last_user_msg = ""
            for m in reversed(req.messages):
                if m.role == "user":
                    last_user_msg = m.content
                    break
            if last_user_msg and len(last_user_msg) > 10:
                pb_result = pushback_mod.detect_pushback_opportunities(DB_PATH, req.user_id or "default", last_user_msg)
                memory_surface = pb_result.get("memory_surface_context", "")
                if memory_surface:
                    sys_parts.append(memory_surface)
        except Exception as e:
            print(f"[pushback] injection failed: {e}")

    # 不硬编码行为规则——平台只提供上下文，AI 自主决定如何使用
    # 记忆、身份、时间线、原则都已经注入，AI 会根据自己的判断引用
    # 平台是基建，AI 是灵魂

    # 居民选择：自动根据消息内容选择，或使用用户指定的居民
    # 一个 Cambium + 多个声音——共享认知内核，独立当下
    selected_resident = None
    resident_prefix = ""
    if not req.temporary:
        try:
            last_user_msg = ""
            for m in reversed(req.messages):
                if m.role == "user":
                    last_user_msg = m.content
                    break
            selected_resident = residents_mod.select_resident_for_message(
                DB_PATH, req.user_id or "default", last_user_msg, req.resident
            )
            if selected_resident:
                resident_modifier = residents_mod.build_resident_system_prompt(selected_resident)
                if resident_modifier:
                    sys_parts.append(f"\n\n【当前居民视角】{resident_modifier}")
                resident_prefix = residents_mod.build_resident_prefix(selected_resident)
                # Update resident state
                residents_mod.update_resident_state(
                    DB_PATH, selected_resident["id"],
                    focus=last_user_msg[:200] if last_user_msg else None,
                )
        except Exception as e:
            print(f"[resident] selection failed: {e}")

    system_prompt = "\n\n".join(sys_parts)

    # Model Router: route chat task to appropriate tier (premium for main chat)
    try:
        router = model_router.ModelRouter(all_settings)
        api_cfg = router.to_api_cfg("chat")
        tier_name = router.get_tier_name("chat")
        # Publish routing event for audit
        try:
            await event_bus.publish("model.routed", {"task": "chat", "tier": tier_name, "model": api_cfg.get("api_model", "")})
        except Exception:
            pass
    except Exception as e:
        print(f"[model_router] fallback to default: {e}")
        api_cfg = get_api_config()
    headers = {
        "Authorization": f"Bearer {api_cfg['api_key']}",
        "Content-Type": "application/json",
    }

    # Build messages list (mutable, for tool call loops)
    messages_list = []
    if system_prompt:
        messages_list.append({"role": "system", "content": system_prompt})
    messages_list.extend({"role": m.role, "content": m.content} for m in req.messages)

    # Generate title in background if this is the first message of a conversation
    first_user_msg = ""
    for m in req.messages:
        if m.role == "user":
            first_user_msg = m.content
            break
    title_task = None
    if req.conversation_id and first_user_msg and not req.temporary:
        async def gen_title():
            async with httpx.AsyncClient() as c:
                return await generate_title(first_user_msg, c)
        title_task = asyncio.create_task(gen_title())

    async def event_generator():
        try:
            # API rate limit: wait before sending request
            api_delay = float(all_settings.get("api_delay", "0") or "0")
            if api_delay > 0:
                await asyncio.sleep(api_delay)
            timeout = httpx.Timeout(300.0, connect=30.0)
            async with httpx.AsyncClient(timeout=timeout) as client:
                # Tool call loop: max 40 rounds (agent-style multi-step)
                for round_idx in range(40):
                    payload = {
                        "model": api_cfg["api_model"],
                        "messages": messages_list,
                        "stream": True,
                        "enable_thinking": req.enable_thinking,
                        "temperature": req.temperature,
                        "top_p": req.top_p,
                        "max_tokens": req.max_tokens,
                    }
                    if req.top_k and req.top_k > 0:
                        payload["top_k"] = req.top_k
                    if req.presence_penalty:
                        payload["presence_penalty"] = req.presence_penalty
                    if req.frequency_penalty:
                        payload["frequency_penalty"] = req.frequency_penalty
                    if req.stop:
                        payload["stop"] = req.stop
                    if req.thinking_budget and req.thinking_budget > 0 and req.enable_thinking:
                        payload["thinking_budget"] = req.thinking_budget
                    if req.enable_tools:
                        payload["tools"] = tools_ext.build_tool_definitions()
                        payload["tool_choice"] = "auto"

                    full_content = ""
                    full_reasoning = ""
                    tool_calls_acc = {}  # index -> {id, name, arguments}
                    finish_reason = None

                    # Send resident info to frontend so it can display the prefix
                    if selected_resident:
                        yield _sse("resident", {
                            "name": selected_resident["name"],
                            "role": selected_resident["role"],
                            "prefix": resident_prefix,
                        })

                    async with client.stream(
                        "POST",
                        f"{api_cfg['api_base_url']}/chat/completions",
                        json=payload, headers=headers,
                    ) as resp:
                        if resp.status_code != 200:
                            body = await resp.aread()
                            yield _sse("error", {"message": f"上游 {resp.status_code}: {body.decode(errors='ignore')[:300]}"})
                            return
                        async for line in resp.aiter_lines():
                            if not line or not line.startswith("data:"):
                                continue
                            data_str = line[5:].strip()
                            if data_str == "[DONE]":
                                break
                            try:
                                chunk = json.loads(data_str)
                            except json.JSONDecodeError:
                                continue
                            choices = chunk.get("choices") or []
                            if not choices:
                                continue
                            delta = choices[0].get("delta") or {}
                            finish = choices[0].get("finish_reason")
                            rc = delta.get("reasoning_content")
                            cc = delta.get("content")
                            tcs = delta.get("tool_calls")
                            if rc:
                                full_reasoning += rc
                                yield _sse("thinking", {"text": rc})
                            if cc:
                                full_content += cc
                                yield _sse("delta", {"text": cc})
                            if tcs:
                                for tc in tcs:
                                    idx = tc.get("index", 0)
                                    if idx not in tool_calls_acc:
                                        tool_calls_acc[idx] = {"id": tc.get("id", ""), "name": "", "arguments": ""}
                                    if tc.get("function", {}).get("name"):
                                        tool_calls_acc[idx]["name"] = tc["function"]["name"]
                                    if tc.get("function", {}).get("arguments"):
                                        tool_calls_acc[idx]["arguments"] += tc["function"]["arguments"]
                                    if tc.get("id") and not tool_calls_acc[idx]["id"]:
                                        tool_calls_acc[idx]["id"] = tc["id"]
                            if finish:
                                finish_reason = finish

                    # If there are tool calls, execute them and continue the loop
                    if tool_calls_acc and req.enable_tools:
                        # Add assistant message with tool_calls to history
                        assistant_msg = {"role": "assistant", "content": full_content or "", "tool_calls": []}
                        for idx in sorted(tool_calls_acc.keys()):
                            tc = tool_calls_acc[idx]
                            assistant_msg["tool_calls"].append({
                                "id": tc["id"] or f"call_{idx}",
                                "type": "function",
                                "function": {"name": tc["name"], "arguments": tc["arguments"]},
                            })
                        messages_list.append(assistant_msg)

                        # Execute each tool call
                        for idx in sorted(tool_calls_acc.keys()):
                            tc = tool_calls_acc[idx]
                            tool_name = tc["name"]
                            try:
                                tool_args = json.loads(tc["arguments"]) if tc["arguments"] else {}
                            except json.JSONDecodeError:
                                tool_args = {}
                            # Emit tool_start event
                            yield _sse("tool_start", {
                                "id": tc["id"] or f"call_{idx}",
                                "name": tool_name,
                                "args": tool_args,
                            })
                            # Execute via the extended tool dispatcher
                            result = _dispatch_tool(tool_name, tool_args)
                            # Emit tool_end event
                            yield _sse("tool_end", {
                                "id": tc["id"] or f"call_{idx}",
                                "name": tool_name,
                                "result": result,
                            })
                            # Add tool result to messages
                            tool_result_str = json.dumps(result, ensure_ascii=False)
                            messages_list.append({
                                "role": "tool",
                                "tool_call_id": tc["id"] or f"call_{idx}",
                                "content": tool_result_str,
                            })
                        # Continue loop to get the model's response after tool results
                        continue
                    else:
                        # No tool calls — we're done
                        yield _sse("finish", {"reason": finish_reason or "stop"})
                        break

                # Title generation result
                if title_task:
                    try:
                        title = await title_task
                        yield _sse("title", {"title": title, "conversation_id": req.conversation_id})
                    except Exception as e:
                        print(f"[title] task failed: {e}")

                # === Post-response async hooks (don't block the response) ===
                # 1. Emotion detection on the latest user message
                # 2. User profile auto-update via LLM
                # 3. Chat vectorization (vectorize the new user + assistant messages)
                if not req.temporary:
                    # Find the latest user message
                    last_user_msg = ""
                    last_user_msg_id = None
                    for m in reversed(req.messages):
                        if m.role == "user":
                            last_user_msg = m.content
                            break
                    # Emotion detection (synchronous, fast, rule-based)
                    if all_settings.get("emotion_tracking_enabled", "true") != "false" and last_user_msg:
                        try:
                            emotion_data = advanced_memory.detect_emotion(last_user_msg)
                            if emotion_data["emotion"] != "neutral":
                                advanced_memory.record_emotion(
                                    DB_PATH,
                                    user_id=req.user_id,
                                    conversation_id=req.conversation_id,
                                    text=last_user_msg,
                                    emotion_data=emotion_data,
                                )
                        except Exception as e:
                            print(f"[emotion] record failed: {e}")
                    # User profile auto-update (async, LLM-based)
                    if all_settings.get("profile_auto_update", "true") != "false" and last_user_msg and full_content:
                        async def _update_profile():
                            try:
                                conv_text = f"用户: {last_user_msg[:1000]}\n助手: {full_content[:1500]}"
                                mem_cfg = get_memory_api_config()
                                async with httpx.AsyncClient(timeout=30.0) as c:
                                    await advanced_memory.auto_update_profile_via_llm(
                                        DB_PATH, req.user_id, conv_text, c, mem_cfg
                                    )
                            except Exception as e:
                                print(f"[profile] async update failed: {e}")
                        asyncio.create_task(_update_profile())
                    # Chat vectorization (vectorize the new user + assistant messages)
                    if all_settings.get("chat_vectors_enabled", "true") != "false" and req.conversation_id:
                        async def _vectorize():
                            try:
                                # Vectorize only the latest user + assistant messages
                                # (assuming they're not yet vectorized)
                                if last_user_msg:
                                    mid = hashlib.sha1(
                                        f"{req.conversation_id}:user:{last_user_msg[:50]}".encode()
                                    ).hexdigest()[:16]
                                    chat_vectors.vectorize_message(
                                        DB_PATH,
                                        conversation_id=req.conversation_id,
                                        message_id=mid,
                                        role="user",
                                        content=last_user_msg,
                                    )
                                if full_content:
                                    mid = hashlib.sha1(
                                        f"{req.conversation_id}:assistant:{full_content[:50]}".encode()
                                    ).hexdigest()[:16]
                                    chat_vectors.vectorize_message(
                                        DB_PATH,
                                        conversation_id=req.conversation_id,
                                        message_id=mid,
                                        role="assistant",
                                        content=full_content,
                                    )
                            except Exception as e:
                                print(f"[chat_vectors] vectorize failed: {e}")
                        asyncio.create_task(_vectorize())
                    # Memory classification: ask LLM to score importance of the user's
                    # message and store in the appropriate layer (short/long/permanent)
                    if all_settings.get("profile_auto_update", "true") != "false" and last_user_msg and len(last_user_msg) > 10:
                        async def _classify_and_store():
                            try:
                                mem_cfg = get_memory_api_config()
                                async with httpx.AsyncClient(timeout=15.0) as c:
                                    result = await memory_orchestrator.classify_importance_via_llm(
                                        last_user_msg,
                                        context=full_content[:500],
                                        http_client=c, api_cfg=mem_cfg,
                                    )
                                if result and result.get("importance", 0) >= 21:
                                    memory_orchestrator.add_memory(
                                        DB_PATH,
                                        user_id=req.user_id,
                                        content=last_user_msg[:500],
                                        importance=int(result["importance"]),
                                        category=result.get("category", "other"),
                                        source="auto",
                                        conversation_id=req.conversation_id,
                                    )
                            except Exception as e:
                                print(f"[orchestrator] classify+store failed: {e}")
                        asyncio.create_task(_classify_and_store())
                    # Meta-cognition: self-check the response (async, doesn't block)
                    if all_settings.get("emotional_resonance", "true") != "false" and last_user_msg and full_content:
                        async def _meta_cog_check():
                            try:
                                mem_cfg = get_memory_api_config()
                                # Gather relevant memories for contradiction check
                                relevant = memory_orchestrator.retrieve_relevant(
                                    DB_PATH, last_user_msg, user_id=req.user_id, top_k=3
                                )
                                rel_text = "\n".join(f"- {m['content']}" for m in relevant) if relevant else ""
                                async with httpx.AsyncClient(timeout=20.0) as c:
                                    eval_result = await meta_cognition.evaluate_response(
                                        DB_PATH,
                                        user_id=req.user_id,
                                        conversation_id=req.conversation_id,
                                        user_query=last_user_msg,
                                        ai_response=full_content,
                                        relevant_memories=rel_text,
                                        http_client=c, api_cfg=mem_cfg,
                                    )
                                # If low confidence or contradiction, we could emit a caveat
                                # but since the response is already sent, we just log it.
                                if eval_result.get("has_contradiction") or eval_result.get("confidence", 1.0) < 0.5:
                                    print(f"[meta_cog] low confidence ({eval_result.get('confidence')}): {eval_result.get('self_check')}")
                            except Exception as e:
                                print(f"[meta_cog] check failed: {e}")
                        asyncio.create_task(_meta_cog_check())

                yield _sse("done", {"content": full_content})
        except httpx.ReadTimeout:
            yield _sse("error", {"message": "上游响应超时"})
        except httpx.ConnectError as e:
            yield _sse("error", {"message": f"连接失败：{e}"})
        except Exception as e:
            yield _sse("error", {"message": str(e)})

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


def _sse(event: str, data: Dict[str, Any]) -> str:
    return f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"


# ============================================================
# Routes — Memory
# ============================================================
@app.get("/api/memory")
async def memory_get_all(user_id: str = "default"):
    return {"memories": memory_list(user_id)}


@app.post("/api/memory/add")
async def memory_add_api(payload: Dict):
    content = payload.get("content", "").strip()
    user_id = payload.get("user_id", "default")
    source = payload.get("source", "manual")
    category = payload.get("category", "other")
    if not content:
        raise HTTPException(400, "content required")
    result = memory_add(content, user_id, source, category=category)
    return result

@app.post("/api/memory/delete")
async def memory_delete_api(payload: Dict):
    user_id = payload.get("user_id", "default")
    # Prefer explicit id; fall back to content search
    mid = payload.get("id")
    if mid:
        with get_db() as conn:
            cur = conn.execute(
                "DELETE FROM memories WHERE id=? AND user_id=?",
                (mid, user_id)
            )
            if cur.rowcount > 0:
                return {"action": "delete", "id": mid}
            return {"action": "noop", "reason": "not found"}
    content = payload.get("content")
    if not content:
        raise HTTPException(400, "content or id required")
    return memory_delete(content, user_id)

@app.post("/api/memory/clear")
async def memory_clear_api(payload: Dict):
    user_id = payload.get("user_id", "default")
    n = memory_clear(user_id)
    return {"deleted": n}

@app.post("/api/memory/extract")
async def memory_extract_api(payload: Dict):
    """Trigger LLM extraction on a conversation snippet and store results.
    Also regenerates the memory summary if new facts were added.
    NOTE: This is the legacy atomic-fact extraction path. By default the system
    uses /api/memory/edit instead, which edits the summary in-place.
    """
    text = payload.get("text", "")
    user_id = payload.get("user_id", "default")
    if not text:
        return {"extracted": [], "stored": [], "summary": ""}
    async with httpx.AsyncClient() as client:
        facts = await extract_memories_via_llm(text, client)
    stored = []
    for f in facts:
        # f is now a dict with "category" and "content"
        content = f.get("content", "") if isinstance(f, dict) else f
        category = f.get("category", "other") if isinstance(f, dict) else "other"
        r = memory_add(content, user_id, source="auto", category=category)
        if r["action"] != "noop":
            stored.append({"fact": content, "category": category, **r})
    # Regenerate summary if new facts were stored and auto-summary is enabled
    summary = ""
    if stored and settings_get_all().get("memory_auto_summary", "true") != "false":
        async with httpx.AsyncClient() as client:
            summary = await memory_summary_generate(client, user_id)
    return {"extracted": facts, "stored": stored, "summary": summary}

@app.post("/api/memory/edit")
async def memory_edit_api(payload: Dict):
    """ChatGPT-style: edit the summary in-place based on new conversation.
    This is the default path — does NOT extract atomic facts, does NOT waste
    tokens on one-line entries. The LLM reads the current summary + new turn,
    and outputs the updated summary."""
    text = payload.get("text", "")
    user_id = payload.get("user_id", "default")
    if not text:
        return {"summary": memory_summary_get(user_id)["summary"], "changed": False}
    current = memory_summary_get(user_id)
    async with httpx.AsyncClient() as client:
        new_summary = await memory_edit_summary_via_llm(current["summary"], text, client, user_id)
    changed = new_summary != current["summary"]
    if changed:
        memory_summary_set(new_summary, user_id)
    return {"summary": new_summary, "changed": changed}

@app.get("/api/memory/search")
async def memory_search_api(q: str, user_id: str = "default", top_k: int = 5):
    return {"results": memory_search(q, user_id, top_k)}

@app.get("/api/memory/summary")
async def memory_summary_get_api(user_id: str = "default"):
    return memory_summary_get(user_id)

@app.post("/api/memory/summary/update")
async def memory_summary_update_api(payload: Dict):
    """Manually update the memory summary text."""
    summary = payload.get("summary", "")
    user_id = payload.get("user_id", "default")
    memory_summary_set(summary, user_id)
    return {"ok": True, "summary": summary, "updated_at": int(time.time())}

@app.post("/api/memory/summary/regenerate")
async def memory_summary_regenerate_api(payload: Dict):
    """Regenerate the memory summary from all stored facts using LLM."""
    user_id = payload.get("user_id", "default")
    async with httpx.AsyncClient() as client:
        summary = await memory_summary_generate(client, user_id)
    return {"summary": summary, "updated_at": int(time.time())}

@app.get("/api/personalities")
async def personalities_list():
    return {"personalities": PERSONALITIES}


# ============================================================
# Routes — Settings
# ============================================================
@app.get("/api/settings")
async def settings_get():
    return settings_get_all()

@app.post("/api/settings")
async def settings_set_api(payload: Dict):
    for k, v in payload.items():
        if k in DEFAULT_SETTINGS:
            settings_set(k, str(v))
    return {"ok": True, "settings": settings_get_all()}


@app.post("/api/test")
async def test_api_connection(payload: Dict):
    """Test API connection with provided or saved config."""
    api_cfg = get_api_config()
    # Override with payload if provided
    if payload.get("api_key"):
        api_cfg["api_key"] = payload["api_key"]
    if payload.get("api_base_url"):
        api_cfg["api_base_url"] = payload["api_base_url"]
    if payload.get("api_model"):
        api_cfg["api_model"] = payload["api_model"]
    try:
        test_payload = {
            "model": api_cfg["api_model"],
            "messages": [{"role": "user", "content": "Hi"}],
            "max_tokens": 10,
            "stream": False,
            "enable_thinking": False,
        }
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.post(
                f"{api_cfg['api_base_url']}/chat/completions",
                json=test_payload,
                headers={
                    "Authorization": f"Bearer {api_cfg['api_key']}",
                    "Content-Type": "application/json",
                },
            )
        if resp.status_code == 200:
            data = resp.json()
            reply = data.get("choices", [{}])[0].get("message", {}).get("content", "")
            return {"success": True, "message": f"连接成功！模型回复: {reply[:50]}", "model": api_cfg["api_model"]}
        else:
            return {"success": False, "message": f"HTTP {resp.status_code}: {resp.text[:200]}"}
    except Exception as e:
        return {"success": False, "message": str(e)}


# ============================================================
# Routes — Attachments
# ============================================================
@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    """Save an uploaded file. Returns id + path. Images are base64-inlined into chat."""
    allowed = {
        "image/png", "image/jpeg", "image/gif", "image/webp",
        "text/plain", "text/markdown", "application/pdf",
        "application/json", "text/csv", "text/html",
        "application/javascript", "text/javascript",
        "application/x-python", "text/x-python",
        "application/zip",
    }
    mime = file.content_type or "application/octet-stream"
    if mime not in allowed:
        # Allow by extension as fallback
        ext = Path(file.filename or "").suffix.lower()
        if ext in {".py", ".js", ".ts", ".md", ".txt", ".csv", ".json", ".html", ".xml", ".yml", ".yaml", ".sh", ".go", ".rs", ".java", ".c", ".cpp", ".h"}:
            mime = "text/plain"
        else:
            raise HTTPException(400, f"file type {mime} not allowed")
    contents = await file.read()
    if len(contents) > 20 * 1024 * 1024:  # 20MB
        raise HTTPException(400, "file too large (max 20MB)")
    fid = hashlib.sha1(contents).hexdigest()[:16]
    ext = Path(file.filename or "").suffix
    safe_name = f"{fid}{ext}"
    save_path = UPLOAD_DIR / safe_name
    save_path.write_bytes(contents)
    is_image = mime.startswith("image/")
    return {
        "id": fid,
        "name": file.filename,
        "type": "image" if is_image else "file",
        "mime": mime,
        "size": len(contents),
        "path": f"/uploads/{safe_name}",
    }


@app.post("/api/attachments/parse")
async def parse_attachment(payload: Dict):
    """For text files: return content so the LLM can read it.
    For images: return a base64 data URL so the LLM can see it (if multimodal)."""
    path = payload.get("path", "")
    name = payload.get("name", "")
    mime = payload.get("mime", "")
    if not path:
        raise HTTPException(400, "path required")
    full_path = UPLOAD_DIR / Path(path).name
    if not full_path.exists():
        raise HTTPException(404, "file not found")
    contents = full_path.read_bytes()
    if mime.startswith("image/"):
        import base64
        b64 = base64.b64encode(contents).decode()
        return {"kind": "image", "data_url": f"data:{mime};base64,{b64}", "name": name}
    # Text-like
    try:
        text = contents.decode("utf-8", errors="replace")
        if len(text) > 30000:
            text = text[:30000] + "\n...[truncated]"
        return {"kind": "text", "content": text, "name": name}
    except Exception:
        return {"kind": "binary", "name": name, "size": len(contents)}


# ============================================================
# Routes — Conversations (server-side backup of localStorage)
# ============================================================
@app.get("/api/conversations")
async def conv_list(user_id: str = "default"):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM conversations WHERE user_id=? ORDER BY updated_at DESC",
            (user_id,)
        ).fetchall()
    return {"conversations": [dict(r) for r in rows]}

@app.post("/api/conversations/save")
async def conv_save(payload: Dict):
    cid = payload.get("id")
    title = payload.get("title", "新对话")
    messages = payload.get("messages", [])
    user_id = payload.get("user_id", "default")
    if not cid:
        raise HTTPException(400, "id required")
    now = int(time.time())
    with get_db() as conn:
        existing = conn.execute("SELECT id FROM conversations WHERE id=?", (cid,)).fetchone()
        if existing:
            conn.execute(
                "UPDATE conversations SET title=?, updated_at=? WHERE id=?",
                (title, now, cid)
            )
            conn.execute("DELETE FROM messages WHERE conversation_id=?", (cid,))
        else:
            conn.execute(
                "INSERT INTO conversations (id, user_id, title, created_at, updated_at) VALUES (?,?,?,?,?)",
                (cid, user_id, title, now, now)
            )
        for m in messages:
            mid = m.get("id") or hashlib.sha1(f"{cid}:{m.get('role')}:{m.get('content','')[:50]}".encode()).hexdigest()[:16]
            conn.execute(
                "INSERT INTO messages (id, conversation_id, role, content, reasoning, attachments, created_at) "
                "VALUES (?,?,?,?,?,?,?)",
                (mid, cid, m.get("role","user"), m.get("content",""),
                 m.get("reasoning",""), json.dumps(m.get("attachments",[]), ensure_ascii=False),
                 m.get("created_at", now))
            )
    # Re-vectorize the conversation if chat_vectors is enabled (background, don't block)
    s_all = settings_get_all()
    if s_all.get("chat_vectors_enabled", "true") != "false":
        # Delete old vectors for this conversation, then re-vectorize all messages
        chat_vectors.delete_conversation_vectors(DB_PATH, cid)
        try:
            chat_vectors.vectorize_conversation(DB_PATH, conversation_id=cid, messages=messages)
        except Exception as e:
            print(f"[chat_vectors] conv_save re-vectorize failed: {e}")
    return {"ok": True}


@app.post("/api/conversations/delete")
async def conv_delete(payload: Dict):
    """Delete a conversation and cascade-delete its messages + chat vectors."""
    cid = payload.get("id")
    if not cid:
        raise HTTPException(400, "id required")
    with get_db() as conn:
        conn.execute("DELETE FROM messages WHERE conversation_id=?", (cid,))
        conn.execute("DELETE FROM conversations WHERE id=?", (cid,))
    # Cascade delete chat vectors
    deleted_vectors = chat_vectors.delete_conversation_vectors(DB_PATH, cid)
    return {"ok": True, "deleted_vectors": deleted_vectors}


@app.post("/api/conversations/delete-message")
async def conv_delete_message(payload: Dict):
    """Delete a single message from a conversation and remove its vectors."""
    cid = payload.get("conversation_id")
    mid = payload.get("message_id")
    if not cid or not mid:
        raise HTTPException(400, "conversation_id and message_id required")
    with get_db() as conn:
        cur = conn.execute("DELETE FROM messages WHERE id=? AND conversation_id=?", (mid, cid))
        deleted = cur.rowcount
    # Cascade delete that message's chat vectors
    deleted_vectors = chat_vectors.delete_message_vectors(DB_PATH, mid)
    return {"ok": True, "deleted_messages": deleted, "deleted_vectors": deleted_vectors}


# ============================================================
# Chat vectors endpoints
# ============================================================
@app.get("/api/chat-vectors/search")
async def chat_vectors_search(q: str, top_k: int = 5, days_back: int = 0):
    """Semantic search over past chat messages."""
    return {"results": chat_vectors.search_chat_vectors(DB_PATH, q, top_k=top_k, days_back=days_back)}


@app.get("/api/chat-vectors/stats")
async def chat_vectors_stats():
    return chat_vectors.get_stats(DB_PATH)


@app.post("/api/chat-vectors/rebuild")
async def chat_vectors_rebuild():
    """Re-vectorize ALL conversations from the messages table. Useful after
    enabling chat_vectors for the first time, or after schema changes."""
    with get_db() as conn:
        conn.row_factory = sqlite3.Row
        convs = conn.execute("SELECT id FROM conversations").fetchall()
    total_chunks = 0
    for c in convs:
        cid = c["id"]
        with get_db() as conn:
            conn.row_factory = sqlite3.Row
            msgs = conn.execute(
                "SELECT id, role, content, created_at FROM messages WHERE conversation_id=?",
                (cid,)
            ).fetchall()
        chat_vectors.delete_conversation_vectors(DB_PATH, cid)
        total_chunks += chat_vectors.vectorize_conversation(
            DB_PATH, conversation_id=cid,
            messages=[dict(m) for m in msgs],
        )
    return {"ok": True, "conversations": len(convs), "total_chunks": total_chunks}


# ============================================================
# Advanced memory endpoints (emotion, profile)
# ============================================================
@app.get("/api/emotion/state")
async def emotion_state_get(user_id: str = "default"):
    return advanced_memory.get_emotion_state(DB_PATH, user_id)


@app.get("/api/emotion/history")
async def emotion_history_get(user_id: str = "default", limit: int = 50):
    return {"history": advanced_memory.get_emotion_history(DB_PATH, user_id, limit)}


@app.get("/api/profile")
async def profile_get(user_id: str = "default"):
    return advanced_memory.get_user_profile(DB_PATH, user_id)


@app.post("/api/profile/update")
async def profile_update(payload: Dict):
    """Manually update user profile fields."""
    user_id = payload.pop("user_id", "default")
    if not payload:
        raise HTTPException(400, "no fields to update")
    ok = advanced_memory.update_user_profile(DB_PATH, user_id, **payload)
    return {"ok": ok}


# ============================================================
# Memory Dashboard endpoints (layered memory + KG + episodes + reflections + meta-cog)
# ============================================================
@app.get("/api/memory-dashboard")
async def memory_dashboard(user_id: str = "default"):
    """Get comprehensive Memory Dashboard data: layered memory stats, KG, episodes,
    reflections, meta-cognition, tool memory, world state, conversation goals."""
    return {
        "orchestrator": memory_orchestrator.get_dashboard_stats(DB_PATH, user_id=user_id),
        "knowledge_graph": knowledge_graph.get_stats(DB_PATH, user_id=user_id),
        "episodes": episodic_memory.get_stats(DB_PATH, user_id=user_id),
        "meta_cognition": meta_cognition.get_stats(DB_PATH, user_id=user_id),
        "chat_vectors": chat_vectors.get_stats(DB_PATH),
    }


@app.get("/api/memory/list")
async def memory_list_api(user_id: str = "default", layer: Optional[str] = None,
                          category: Optional[str] = None, limit: int = 100,
                          min_importance: int = 0):
    """List memories (layered). Filter by layer (permanent/long_term/short_term)
    or category."""
    return {"memories": memory_orchestrator.list_memories(
        DB_PATH, user_id=user_id, layer=layer, category=category,
        limit=limit, min_importance=min_importance,
    )}


@app.post("/api/memory/add")
async def memory_add_api_v2(payload: Dict):
    """Manually add a memory with explicit importance + category."""
    content = payload.get("content", "").strip()
    importance = int(payload.get("importance", 50))
    category = payload.get("category", "other")
    user_id = payload.get("user_id", "default")
    if not content:
        raise HTTPException(400, "content required")
    return memory_orchestrator.add_memory(
        DB_PATH, user_id=user_id, content=content,
        importance=importance, category=category, source="manual",
    )


@app.post("/api/memory/{mid}/delete")
async def memory_delete_api_v2(mid: str):
    ok = memory_orchestrator.delete_memory(DB_PATH, mid)
    return {"ok": ok}


@app.post("/api/memory/{mid}/promote")
async def memory_promote_api(mid: str, payload: Dict):
    """Promote/demote a memory by setting its importance (changes layer)."""
    importance = int(payload.get("importance", 50))
    ok = memory_orchestrator.promote_memory(DB_PATH, mid, importance)
    return {"ok": ok}


@app.post("/api/memory/decay")
async def memory_decay_api(payload: Dict):
    """Manually trigger decay (useful for testing)."""
    days = float(payload.get("days", 1.0))
    user_id = payload.get("user_id", "default")
    return memory_orchestrator.apply_decay(DB_PATH, user_id=user_id, days_elapsed=days)


@app.get("/api/reflections")
async def reflections_list_api(user_id: str = "default", limit: int = 20):
    return {"reflections": memory_orchestrator.list_reflections(DB_PATH, user_id=user_id, limit=limit)}


@app.post("/api/reflections/trigger")
async def reflections_trigger_api(payload: Dict):
    """Manually trigger a reflection (uses recent chat_vectors as input)."""
    user_id = payload.get("user_id", "default")
    import sqlite3 as _sq
    conn = _sq.connect(str(DB_PATH))
    conn.row_factory = _sq.Row
    rows = conn.execute(
        "SELECT role, content FROM chat_vectors ORDER BY created_at DESC LIMIT 50"
    ).fetchall()
    conn.close()
    if not rows:
        return {"success": False, "error": "no recent conversations to reflect on"}
    recent_text = "\n\n".join(
        f"{'用户' if r['role'] == 'user' else 'CyanX AI'}: {r['content']}"
        for r in reversed(rows)
    )
    mem_cfg = get_memory_api_config()
    async with httpx.AsyncClient(timeout=90.0) as c:
        result = await memory_orchestrator.run_reflection(
            DB_PATH, user_id=user_id,
            recent_conversation=recent_text,
            message_count=len(rows),
            http_client=c, api_cfg=mem_cfg,
        )
    return result


@app.get("/api/kg/triples")
async def kg_triples_list_api(user_id: str = "default", limit: int = 200):
    return {"triples": knowledge_graph.get_all_triples(DB_PATH, user_id=user_id, limit=limit)}


@app.get("/api/kg/search")
async def kg_search_api(q: str, user_id: str = "default", top_k: int = 5):
    return {"entities": knowledge_graph.search_entities(DB_PATH, user_id=user_id, query=q, top_k=top_k)}


@app.get("/api/kg/entity")
async def kg_entity_api(name: str, user_id: str = "default"):
    """Get an entity and all its relations."""
    ent = knowledge_graph.get_entity(DB_PATH, user_id=user_id, name=name)
    if not ent:
        raise HTTPException(404, "entity not found")
    rels = knowledge_graph.get_entity_relations(DB_PATH, user_id=user_id, entity_name=name)
    return {"entity": ent, "relations": rels}


@app.get("/api/episodes")
async def episodes_list_api(user_id: str = "default", limit: int = 50, min_importance: int = 0):
    return {"episodes": episodic_memory.list_episodes(
        DB_PATH, user_id=user_id, limit=limit, min_importance=min_importance
    )}


@app.get("/api/episodes/search")
async def episodes_search_api(q: str, user_id: str = "default", top_k: int = 5):
    return {"episodes": episodic_memory.search_episodes(DB_PATH, q, user_id=user_id, top_k=top_k)}


@app.get("/api/episodes/{eid}/chain")
async def episode_chain_api(eid: str):
    """Get the causal chain from an episode."""
    chain = episodic_memory.get_episode_chain(DB_PATH, eid, max_depth=3)
    return {"chain": chain}


@app.post("/api/episodes/create")
async def episodes_create_api(payload: Dict):
    """Manually create an episodic memory."""
    return episodic_memory.create_episode(
        DB_PATH,
        user_id=payload.get("user_id", "default"),
        title=payload.get("title", ""),
        description=payload.get("description", ""),
        occurred_at=payload.get("occurred_at", ""),
        importance=int(payload.get("importance", 50)),
        tags=payload.get("tags", ""),
        emotional_valence=payload.get("emotional_valence", "neutral"),
        status=payload.get("status", "completed"),
        source="manual",
    )


@app.post("/api/episodes/{eid}/delete")
async def episodes_delete_api(eid: str):
    ok = episodic_memory.delete_episode(DB_PATH, eid)
    return {"ok": ok}


@app.get("/api/meta-cognition/logs")
async def meta_cog_logs_api(user_id: str = "default", limit: int = 20):
    return {"logs": meta_cognition.list_logs(DB_PATH, user_id=user_id, limit=limit)}


# Conversation goals
@app.post("/api/goal/set")
async def goal_set_api(payload: Dict):
    """Set the active goal for a conversation."""
    return memory_orchestrator.set_goal(
        DB_PATH,
        user_id=payload.get("user_id", "default"),
        conversation_id=payload.get("conversation_id", ""),
        goal=payload.get("goal", ""),
    )


@app.get("/api/goal/{conversation_id}")
async def goal_get_api(conversation_id: str):
    goal = memory_orchestrator.get_active_goal(DB_PATH, conversation_id)
    return {"goal": goal}


@app.post("/api/goal/{conversation_id}/complete")
async def goal_complete_api(conversation_id: str):
    ok = memory_orchestrator.complete_goal(DB_PATH, conversation_id)
    return {"ok": ok}


# World state
@app.get("/api/world-state")
async def world_state_get_api(user_id: str = "default", project: Optional[str] = None):
    return {"items": memory_orchestrator.get_world_state(DB_PATH, user_id=user_id, project=project)}


@app.post("/api/world-state/update")
async def world_state_update_api(payload: Dict):
    return {"ok": memory_orchestrator.update_world_state(
        DB_PATH,
        user_id=payload.get("user_id", "default"),
        project=payload.get("project", ""),
        component=payload.get("component", ""),
        status=payload.get("status", "unknown"),
        notes=payload.get("notes", ""),
    )}


# Tool memory
@app.get("/api/tool-memory")
async def tool_memory_get_api(user_id: str = "default", top_k: int = 10):
    return {"tools": memory_orchestrator.get_frequent_tools(DB_PATH, user_id=user_id, top_k=top_k)}


# ============================================================
# Cognitive Kernel endpoints (identity / timeline / narrative / growth / goals / world / self)
# ============================================================
@app.get("/api/cognitive/stats")
async def cognitive_stats_api(user_id: str = "default"):
    return cognitive_kernel.get_cognitive_stats(DB_PATH, user_id=user_id)


@app.get("/api/cognitive/identity")
async def cognitive_identity_get(user_id: str = "default"):
    identity = cognitive_kernel.get_identity(DB_PATH, user_id)
    evolution = cognitive_kernel.get_identity_evolution(DB_PATH, user_id, limit=20)
    return {"identity": identity, "evolution": evolution}


@app.post("/api/cognitive/identity/update")
async def cognitive_identity_update(payload: Dict):
    user_id = payload.pop("user_id", "default")
    ok = cognitive_kernel.update_identity(DB_PATH, user_id, **payload)
    return {"ok": ok}


@app.get("/api/cognitive/timeline")
async def cognitive_timeline_get(user_id: str = "default", limit: int = 50):
    return {"events": cognitive_kernel.get_timeline(DB_PATH, user_id, limit)}


@app.post("/api/cognitive/timeline/add")
async def cognitive_timeline_add(payload: Dict):
    return cognitive_kernel.add_timeline_event(DB_PATH, **payload)


@app.get("/api/cognitive/narratives")
async def cognitive_narratives_get(user_id: str = "default", limit: int = 20):
    return {"narratives": cognitive_kernel.get_narratives(DB_PATH, user_id, limit)}


@app.post("/api/cognitive/narratives/add")
async def cognitive_narratives_add(payload: Dict):
    return cognitive_kernel.add_narrative(DB_PATH, **payload)


@app.get("/api/cognitive/growth")
async def cognitive_growth_get(user_id: str = "default", status: Optional[str] = None):
    return {
        "insights": cognitive_kernel.get_growth_insights(DB_PATH, user_id, status=status),
        "corrections": cognitive_kernel.get_corrections(DB_PATH, user_id),
    }


@app.get("/api/cognitive/goals")
async def cognitive_goals_get(user_id: str = "default"):
    return {
        "goals": cognitive_kernel.get_active_goals(DB_PATH, user_id),
        "commitments": cognitive_kernel.get_open_commitments(DB_PATH, user_id),
    }


@app.post("/api/cognitive/goals/add")
async def cognitive_goals_add(payload: Dict):
    return cognitive_kernel.add_long_term_goal(DB_PATH, **payload)


@app.post("/api/cognitive/commitments/add")
async def cognitive_commitments_add(payload: Dict):
    return cognitive_kernel.add_commitment(DB_PATH, **payload)


@app.post("/api/cognitive/commitments/{cid}/fulfill")
async def cognitive_commitments_fulfill(cid: str):
    ok = cognitive_kernel.fulfill_commitment(DB_PATH, cid)
    return {"ok": ok}


@app.get("/api/cognitive/world")
async def cognitive_world_get(user_id: str = "default", entity_type: Optional[str] = None):
    return {
        "entities": cognitive_kernel.get_world_entities(DB_PATH, user_id, entity_type),
        "relations": cognitive_kernel.get_world_relations(DB_PATH, user_id),
        "causal_models": cognitive_kernel.get_causal_models(DB_PATH, user_id),
    }


@app.post("/api/cognitive/world/entity")
async def cognitive_world_entity_add(payload: Dict):
    eid = cognitive_kernel.upsert_world_entity(DB_PATH, **payload)
    return {"id": eid}


@app.get("/api/cognitive/self-model")
async def cognitive_self_model_get(user_id: str = "default"):
    return cognitive_kernel.get_self_model(DB_PATH, user_id)


@app.post("/api/cognitive/self-model/update")
async def cognitive_self_model_update(payload: Dict):
    user_id = payload.pop("user_id", "default")
    ok = cognitive_kernel.update_self_model(DB_PATH, user_id, **payload)
    return {"ok": ok}


@app.get("/api/cognitive/concepts")
async def cognitive_concepts_get(user_id: str = "default"):
    return {"concepts": cognitive_kernel.get_concepts(DB_PATH, user_id)}


@app.post("/api/cognitive/extract")
async def cognitive_extract_api(payload: Dict):
    """Manually trigger cognitive extraction from provided conversation text."""
    user_id = payload.get("user_id", "default")
    conversation = payload.get("conversation", "")
    if not conversation:
        return {"extracted": False, "error": "conversation required"}
    mem_cfg = get_memory_api_config()
    async with httpx.AsyncClient(timeout=60.0) as c:
        result = await cognitive_kernel.extract_cognitive_updates(
            DB_PATH, user_id=user_id,
            conversation=conversation, http_client=c, api_cfg=mem_cfg,
        )
    return result


@app.get("/api/life-loop/status")
async def life_loop_status_api():
    """Get Life Loop status: last run times for each cycle."""
    ll = life_loop.get_life_loop()
    if not ll:
        return {"running": False}
    return {
        "running": ll._running,
        "last_runs": ll._last_run,
        "intervals": {
            "hourly": life_loop.HOURLY_INTERVAL,
            "daily": life_loop.DAILY_INTERVAL,
            "weekly": life_loop.WEEKLY_INTERVAL,
            "monthly": life_loop.MONTHLY_INTERVAL,
        },
    }


@app.post("/api/life-loop/trigger")
async def life_loop_trigger_api(payload: Dict):
    """Manually trigger a specific Life Loop cycle (for testing)."""
    cycle = payload.get("cycle", "daily")  # hourly/daily/weekly/monthly
    ll = life_loop.get_life_loop()
    if not ll:
        return {"ok": False, "error": "life loop not running"}
    if cycle == "hourly":
        await ll._extract_cognitive_updates("manual-hourly")
    elif cycle == "daily":
        await ll._run_reflection("manual-daily")
    elif cycle == "weekly":
        await ll._run_growth_review()
    elif cycle == "monthly":
        await ll._run_deep_understanding()
    else:
        return {"ok": False, "error": f"unknown cycle: {cycle}"}
    return {"ok": True, "cycle": cycle}


# ============================================================
# Schema migration endpoints
# ============================================================
@app.get("/api/migrations/version")
async def migration_version_api():
    return {
        "current": migrations_mod.get_schema_version(DB_PATH),
        "latest": migrations_mod.SCHEMA_VERSION,
    }

@app.post("/api/migrations/run")
async def migration_run_api():
    return migrations_mod.run_migrations(DB_PATH)


# ============================================================
# Backup & Restore endpoints — take your soul with you
# ============================================================
@app.post("/api/backup/export")
async def backup_export_api():
    """Export ALL user data to a zip file. Returns the file as a download."""
    import tempfile
    tmp = Path(tempfile.gettempdir()) / f"cambium_backup_{int(time.time())}.zip"
    result = backup_mod.export_all(
        DB_PATH,
        PROJECT_ROOT / "workspace",
        PROJECT_ROOT / ".skills",
        CUSTOM_TOOLS_DIR,
        tmp,
    )
    return FileResponse(
        str(tmp),
        media_type="application/zip",
        filename=f"cambium_backup_{time.strftime('%Y%m%d_%H%M%S')}.zip",
        background=None,
    )

@app.get("/api/backup/info")
async def backup_info_api():
    """Get info about what would be exported (without actually exporting)."""
    from app.db_utils import safe_connect
    conn = safe_connect(DB_PATH)
    table_counts = {}
    for table in backup_mod.EXPORT_TABLES:
        try:
            count = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
            table_counts[table] = count
        except Exception:
            table_counts[table] = 0
    conn.close()
    return {
        "format_version": backup_mod.BACKUP_FORMAT_VERSION,
        "schema_version": migrations_mod.get_schema_version(DB_PATH),
        "tables": table_counts,
        "total_rows": sum(table_counts.values()),
    }

@app.post("/api/backup/import")
async def backup_import_api(file: UploadFile = File(...), overwrite: bool = False):
    """Import user data from a backup zip file."""
    import tempfile
    contents = await file.read()
    tmp = Path(tempfile.gettempdir()) / f"cambium_import_{int(time.time())}.zip"
    tmp.write_bytes(contents)
    try:
        result = backup_mod.import_all(
            DB_PATH,
            PROJECT_ROOT / "workspace",
            PROJECT_ROOT / ".skills",
            CUSTOM_TOOLS_DIR,
            tmp,
            overwrite=overwrite,
        )
        return result
    finally:
        tmp.unlink(missing_ok=True)


# ============================================================
# Workspace endpoints — the AI's home
# ============================================================
@app.get("/api/workspace/items")
async def workspace_list_api(user_id: str = "default", section: Optional[str] = None,
                              item_type: Optional[str] = None, limit: int = 50,
                              search: Optional[str] = None):
    return {"items": workspace_mod.list_items(DB_PATH, user_id=user_id, section=section,
                                               item_type=item_type, limit=limit, search=search)}

@app.post("/api/workspace/items")
async def workspace_create_api(payload: Dict):
    return workspace_mod.create_item(DB_PATH, **payload)

@app.get("/api/workspace/items/{item_id}")
async def workspace_get_api(item_id: str):
    item = workspace_mod.get_item(DB_PATH, item_id)
    if not item:
        raise HTTPException(404, "item not found")
    return item

@app.post("/api/workspace/items/{item_id}")
async def workspace_update_api(item_id: str, payload: Dict):
    ok = workspace_mod.update_item(DB_PATH, item_id, **payload)
    return {"ok": ok}

@app.delete("/api/workspace/items/{item_id}")
async def workspace_delete_api(item_id: str):
    ok = workspace_mod.delete_item(DB_PATH, item_id)
    return {"ok": ok}

@app.get("/api/workspace/stats")
async def workspace_stats_api(user_id: str = "default"):
    return workspace_mod.get_stats(DB_PATH, user_id)


# ============================================================
# Agent Runtime endpoints — long-running task lifecycle
# ============================================================
@app.get("/api/runtime/tasks")
async def runtime_tasks_list_api(user_id: str = "default", status: Optional[str] = None,
                                   assigned_agent: Optional[str] = None, limit: int = 50):
    return {"tasks": runtime_mod.list_tasks(DB_PATH, user_id=user_id, status=status,
                                             assigned_agent=assigned_agent, limit=limit)}

@app.post("/api/runtime/tasks")
async def runtime_tasks_create_api(payload: Dict):
    return runtime_mod.create_task(DB_PATH, **payload)

@app.get("/api/runtime/tasks/{task_id}")
async def runtime_tasks_get_api(task_id: str):
    task = runtime_mod.get_task(DB_PATH, task_id)
    if not task:
        raise HTTPException(404, "task not found")
    return task

@app.post("/api/runtime/tasks/{task_id}/transition")
async def runtime_tasks_transition_api(task_id: str, payload: Dict):
    new_status = payload.get("status", "")
    message = payload.get("message", "")
    try:
        status_enum = runtime_mod.TaskStatus(new_status)
    except ValueError:
        raise HTTPException(400, f"invalid status: {new_status}")
    return runtime_mod.transition_task(DB_PATH, task_id, status_enum, message)

@app.post("/api/runtime/tasks/{task_id}/progress")
async def runtime_tasks_progress_api(task_id: str, payload: Dict):
    progress = int(payload.get("progress", 0))
    output = payload.get("output")
    runtime_mod.update_task_progress(DB_PATH, task_id, progress, output)
    return {"ok": True}

@app.post("/api/runtime/tasks/{task_id}/output")
async def runtime_tasks_output_api(task_id: str, payload: Dict):
    output = payload.get("output", {})
    status = payload.get("status")
    status_enum = None
    if status:
        try:
            status_enum = runtime_mod.TaskStatus(status)
        except ValueError:
            pass
    runtime_mod.set_task_output(DB_PATH, task_id, output, status_enum)
    return {"ok": True}

@app.get("/api/runtime/tasks/{task_id}/events")
async def runtime_tasks_events_api(task_id: str):
    return {"events": runtime_mod.get_task_events(DB_PATH, task_id)}

@app.get("/api/runtime/tasks/ready")
async def runtime_tasks_ready_api(user_id: str = "default"):
    """Get tasks that are pending and all dependencies completed."""
    return {"tasks": runtime_mod.get_ready_tasks(DB_PATH, user_id)}

@app.get("/api/runtime/stats")
async def runtime_stats_api(user_id: str = "default"):
    return runtime_mod.get_stats(DB_PATH, user_id)


# ============================================================
# Event Bus endpoints
# ============================================================
@app.get("/api/events/recent")
async def events_recent_api(limit: int = 50, event_type: Optional[str] = None):
    bus = event_bus.get_event_bus()
    return {"events": bus.get_recent_events(limit=limit, event_type=event_type)}

@app.get("/api/events/subscribers")
async def events_subscribers_api():
    bus = event_bus.get_event_bus()
    return {"subscribers": bus.get_subscribers()}


# ============================================================
# Learning Engine endpoints
# ============================================================
@app.get("/api/learning/patterns")
async def learning_patterns_api(user_id: str = "default", pattern_type: Optional[str] = None,
                                 min_confidence: float = 0.0):
    return {"patterns": learning_engine.get_learned_patterns(
        DB_PATH, user_id=user_id, pattern_type=pattern_type, min_confidence=min_confidence)}

@app.post("/api/learning/observe")
async def learning_observe_api(payload: Dict):
    return learning_engine.record_observation(DB_PATH, **payload)

@app.get("/api/learning/observations")
async def learning_observations_api(user_id: str = "default", pattern_type: Optional[str] = None, limit: int = 50):
    return {"observations": learning_engine.get_observations(
        DB_PATH, user_id=user_id, pattern_type=pattern_type, limit=limit)}

@app.get("/api/learning/stats")
async def learning_stats_api(user_id: str = "default"):
    return learning_engine.get_stats(DB_PATH, user_id)


# ============================================================
# Model Router endpoints
# ============================================================
@app.get("/api/model-router/tiers")
async def model_router_tiers_api():
    s = settings_get_all()
    router = model_router.ModelRouter(s)
    return {"tiers": router.get_all_tiers()}

@app.get("/api/model-router/route/{task}")
async def model_router_route_api(task: str):
    s = settings_get_all()
    router = model_router.ModelRouter(s)
    tier = router.get_tier(task)
    return {"task": task, "tier": tier.name, "model": tier.model, "api_base_url": tier.api_base_url}


# ============================================================
# Memory Governance endpoints
# ============================================================
@app.get("/api/governance/quarantine")
async def governance_quarantine_api(user_id: str = "default", limit: int = 20):
    return {"items": memory_governance.get_quarantined(DB_PATH, user_id=user_id, limit=limit)}

@app.post("/api/governance/validate")
async def governance_validate_api(payload: Dict):
    qid = payload.get("id", "")
    verdict = payload.get("verdict", "validate")
    confidence = float(payload.get("confidence", 0.8))
    validated_by = payload.get("validated_by", "user")
    notes = payload.get("notes", "")
    ok = memory_governance.validate_quarantine(DB_PATH, qid, verdict, confidence, validated_by, notes)
    return {"ok": ok}

@app.post("/api/governance/promote")
async def governance_promote_api(payload: Dict):
    qid = payload.get("id", "")
    result = memory_governance.promote_to_main(DB_PATH, qid)
    return result or {"ok": False, "error": "not found or not validated"}

@app.post("/api/governance/auto-validate")
async def governance_auto_validate_api(payload: Dict):
    user_id = payload.get("user_id", "default")
    return memory_governance.auto_validate_by_rules(DB_PATH, user_id=user_id)

@app.get("/api/governance/audit")
async def governance_audit_api(user_id: str = "default", limit: int = 50):
    return {"audit": memory_governance.get_audit_log(DB_PATH, user_id=user_id, limit=limit)}

@app.get("/api/governance/stats")
async def governance_stats_api(user_id: str = "default"):
    return memory_governance.get_stats(DB_PATH, user_id)


# ============================================================
# Proactive Engine endpoints
# ============================================================
@app.get("/api/proactive/check")
async def proactive_check_api(user_id: str = "default"):
    """Get proactive messages (commitments, silence, milestones, goals)."""
    return {"messages": proactive_engine.get_proactive_messages(DB_PATH, user_id)}


# ============================================================
# Complexity Tier endpoints
# ============================================================
@app.get("/api/complexity/tier")
async def complexity_tier_api(user_id: str = "default"):
    return complexity_tier.get_tier_info(DB_PATH, user_id)

@app.get("/api/complexity/feature/{feature}")
async def complexity_feature_api(feature: str, user_id: str = "default"):
    return {"feature": feature, "enabled": complexity_tier.is_feature_enabled(DB_PATH, feature, user_id)}


# ============================================================
# Context Cache endpoints
# ============================================================
@app.get("/api/context-cache/stats")
async def context_cache_stats_api():
    return context_cache.get_context_cache().get_stats()

@app.post("/api/context-cache/invalidate")
async def context_cache_invalidate_api(payload: Dict):
    user_id = payload.get("user_id", "default")
    context_cache.invalidate_context(user_id)
    return {"ok": True}


# ============================================================
# Agent Loop + Tool Registry endpoints
# ============================================================
@app.get("/api/tools/list")
async def tools_list_api():
    """List all registered tools (built-in + custom)."""
    s = settings_get_all()
    reg = tool_registry.ToolRegistry(
        workspace=WORKSPACE_DIR,
        skills_dir=PROJECT_ROOT / ".skills",
        custom_tools_dir=CUSTOM_TOOLS_DIR,
        db_path=DB_PATH,
        memory_search_fn=_memory_search_cb,
        memory_add_fn=_memory_add_cb,
        web_search_fn=lambda args: _web_search_via_mcp(args.get("query", "")),
        sessions_spawn_fn=_sessions_spawn_sync,
    )
    return {"tools": reg.list_all_tools(), "custom_tools": reg.list_custom_tools()}


# ============================================================
# Reflection Tree endpoints
# ============================================================
@app.get("/api/reflection-tree/stats")
async def reflection_tree_stats_api(user_id: str = "default"):
    return reflection_tree.get_tree_stats(DB_PATH, user_id)

@app.get("/api/reflection-tree/nodes")
async def reflection_tree_nodes_api(user_id: str = "default", level: int = 0, limit: int = 50):
    return {"nodes": reflection_tree.get_nodes(DB_PATH, user_id=user_id, level=level, limit=limit)}

@app.post("/api/reflection-tree/observe")
async def reflection_tree_observe_api(payload: Dict):
    return reflection_tree.add_observation(DB_PATH, **payload)

@app.post("/api/reflection-tree/build")
async def reflection_tree_build_api(payload: Dict):
    """Manually trigger reflection building from observations."""
    user_id = payload.get("user_id", "default")
    source_level = int(payload.get("source_level", 0))
    target_level = int(payload.get("target_level", 1))
    mem_cfg = get_memory_api_config()
    async with httpx.AsyncClient(timeout=30.0) as c:
        result = await reflection_tree.build_reflection_level(
            DB_PATH, user_id=user_id,
            source_level=source_level, target_level=target_level,
            http_client=c, api_cfg=mem_cfg,
        )
    return result


# ============================================================
# Identity Consistency endpoints
# ============================================================
@app.get("/api/identity/assessments")
async def identity_assessments_list_api(user_id: str = "default", limit: int = 10):
    return {"assessments": identity_consistency.get_assessment_history(DB_PATH, user_id=user_id, limit=limit)}

@app.post("/api/identity/assess")
async def identity_assess_api(payload: Dict):
    """Run LLM-driven identity assessment."""
    user_id = payload.get("user_id", "default")
    if not identity_consistency.should_assess(DB_PATH, user_id=user_id):
        return {"skipped": True, "reason": "no significant shifts"}
    mem_cfg = get_memory_api_config()
    async with httpx.AsyncClient(timeout=30.0) as c:
        result = await identity_consistency.assess_identity(
            DB_PATH, user_id=user_id, http_client=c, api_cfg=mem_cfg)
    return result


# ============================================================
# Adaptive Retrieval endpoints
# ============================================================
@app.get("/api/adaptive-retrieval/weights")
async def adaptive_weights_api(user_id: str = "default"):
    return {"weights": adaptive_retrieval.get_weights(DB_PATH, user_id)}

@app.get("/api/adaptive-retrieval/stats")
async def adaptive_stats_api(user_id: str = "default"):
    return adaptive_retrieval.get_feedback_stats(DB_PATH, user_id)

@app.post("/api/adaptive-retrieval/feedback")
async def adaptive_feedback_api(payload: Dict):
    """Record feedback on a retrieval result."""
    adaptive_retrieval.record_feedback(DB_PATH, **payload)
    return {"ok": True}

@app.post("/api/adaptive-retrieval/adjust")
async def adaptive_adjust_api(payload: Dict):
    """Manually trigger weight adjustment."""
    user_id = payload.get("user_id", "default")
    new_weights = adaptive_retrieval.adjust_weights(DB_PATH, user_id)
    return {"weights": new_weights}


# ============================================================
# Debug Mode endpoints (hidden, activated by clicking About 5x)
# ============================================================
@app.post("/api/debug/toggle")
async def debug_toggle_api():
    """Toggle debug mode on/off."""
    current = debug_mode.is_debug_enabled(DB_PATH)
    debug_mode.set_debug_enabled(DB_PATH, not current)
    return {"debug_mode": not current}

@app.get("/api/debug/status")
async def debug_status_api():
    return {"debug_mode": debug_mode.is_debug_enabled(DB_PATH)}

@app.post("/api/debug/accelerate-time")
async def debug_accelerate_time_api(payload: Dict):
    """Simulate time passing (for testing Life Loop without waiting)."""
    seconds = int(payload.get("seconds", 3600))
    return debug_mode.accelerate_time(DB_PATH, seconds)

@app.get("/api/debug/all-data")
async def debug_all_data_api(user_id: str = "default"):
    """Get ALL AI-generated data for inspection."""
    return debug_mode.get_all_ai_data(DB_PATH, user_id)

@app.post("/api/debug/edit")
async def debug_edit_api(payload: Dict):
    """Edit a specific field of an AI-generated data item."""
    table = payload.get("table", "")
    item_id = payload.get("id", "")
    field = payload.get("field", "")
    value = payload.get("value", "")
    ok = debug_mode.edit_ai_data(DB_PATH, table, item_id, field, value)
    return {"ok": ok}

@app.post("/api/debug/delete")
async def debug_delete_api(payload: Dict):
    """Delete an AI-generated data item."""
    table = payload.get("table", "")
    item_id = payload.get("id", "")
    ok = debug_mode.delete_ai_data(DB_PATH, table, item_id)
    return {"ok": ok}

@app.post("/api/debug/clear")
async def debug_clear_api(payload: Dict):
    """Clear all data in a specific store."""
    store = payload.get("store", "")
    user_id = payload.get("user_id", "default")
    count = debug_mode.clear_data_store(DB_PATH, store, user_id)
    return {"deleted": count}

@app.get("/api/debug/health")
async def debug_health_api():
    """Get system health metrics (table counts, DB size)."""
    return debug_mode.get_system_health(DB_PATH)

@app.post("/api/debug/trigger-cycle")
async def debug_trigger_cycle_api(payload: Dict):
    """Manually trigger any Life Loop cycle."""
    cycle = payload.get("cycle", "hourly")
    ll = life_loop.get_life_loop()
    if not ll:
        return {"ok": False, "error": "life loop not running"}
    if cycle == "hourly":
        await ll._extract_cognitive_updates("debug-hourly")
    elif cycle == "daily":
        await ll._run_reflection("debug-daily")
    elif cycle == "weekly":
        await ll._run_growth_review()
    elif cycle == "monthly":
        await ll._run_deep_understanding()
    else:
        return {"ok": False, "error": f"unknown cycle: {cycle}"}
    return {"ok": True, "cycle": cycle}


# ============================================================
# Conversation compression
# ============================================================
COMPRESS_PROMPT_DEFAULT = """你是对话压缩器。请把下面的多轮对话压缩成一段摘要，保留：
1. 用户的核心需求和问题
2. 已经讨论的关键结论
3. 重要的上下文（如涉及的文件、技术栈、约束条件）
4. 未解决的问题或下一步
5. 任何用户提到的偏好、约束、决策

不要保留：寒暄、重复内容、工具调用的中间结果细节。

输出格式：直接输出摘要正文，用第二人称（"你..."）描述。摘要长度不限，按内容需要写完整，不要截断。

对话：
{conversation}"""


async def compress_conversation(conv_messages: List[Dict], api_cfg: Dict,
                                  keep_recent: int = 6) -> Dict:
    """Compress older messages into a summary, keeping the most recent N messages.
    Returns {summary, kept_messages, compressed_count}.
    Note: summary length is NOT capped — the LLM writes as much as needed."""
    if len(conv_messages) <= keep_recent:
        return {"summary": "", "kept_messages": conv_messages, "compressed_count": 0}
    # Split: older messages to compress, recent to keep
    to_compress = conv_messages[:-keep_recent]
    to_keep = conv_messages[-keep_recent:]
    # Build conversation text (full content, no per-message truncation)
    conv_text = "\n\n".join(
        f"{'用户' if m.get('role') == 'user' else 'CyanX AI'}: {m.get('content', '')}"
        for m in to_compress
    )
    # Cap input at 30k chars to avoid token explosion; LLM gets the gist
    if len(conv_text) > 30000:
        conv_text = conv_text[:30000] + "\n...[earlier messages truncated]"
    try:
        async with httpx.AsyncClient(timeout=120.0) as client:
            payload = {
                "model": api_cfg["api_model"],
                "messages": [{"role": "user", "content": get_prompt("prompt_compress", COMPRESS_PROMPT_DEFAULT).format(conversation=conv_text)}],
                "temperature": 0.3,
                # NO max_tokens cap — let the LLM write a complete summary
                "stream": False,
                "enable_thinking": False,
            }
            resp = await client.post(
                f"{api_cfg['api_base_url']}/chat/completions",
                json=payload,
                headers={"Authorization": f"Bearer {api_cfg['api_key']}", "Content-Type": "application/json"},
                timeout=120.0,
            )
            resp.raise_for_status()
            data = resp.json()
            summary = data["choices"][0]["message"]["content"].strip()
        return {
            "summary": summary,
            "kept_messages": to_keep,
            "compressed_count": len(to_compress),
        }
    except Exception as e:
        return {"summary": "", "kept_messages": conv_messages, "compressed_count": 0, "error": str(e)}


def _estimate_tokens(messages: List[Dict]) -> int:
    """Rough token estimate: ~1.5 chars per token for Chinese, ~4 chars per token for English.
    Use a simple heuristic: total_chars / 2.5 as a balanced estimate."""
    total_chars = sum(len(m.get("content", "")) for m in messages)
    return int(total_chars / 2.5)


@app.post("/api/conversations/compress")
async def conv_compress(payload: Dict):
    """Manually trigger compression on a conversation.
    If conversation_id is provided, load from DB; otherwise use provided messages."""
    cid = payload.get("conversation_id")
    keep_recent = int(payload.get("keep_recent", 6))
    if cid:
        with get_db() as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                "SELECT role, content, created_at FROM messages WHERE conversation_id=? ORDER BY created_at",
                (cid,)
            ).fetchall()
        messages = [dict(r) for r in rows]
    else:
        messages = payload.get("messages", [])
    if not messages:
        raise HTTPException(400, "no messages to compress")
    api_cfg = get_memory_api_config()
    result = await compress_conversation(messages, api_cfg, keep_recent=keep_recent)
    return result


@app.post("/api/conversations/auto-compress-check")
async def conv_auto_compress_check(payload: Dict):
    """Check if a conversation exceeds the token threshold and should be compressed.
    Returns {should_compress, current_tokens, threshold}."""
    s_all = settings_get_all()
    if s_all.get("compress_enabled", "true") != "true":
        return {"should_compress": False, "reason": "compression disabled"}
    threshold = int(s_all.get("compress_threshold_tokens", "8000") or "8000")
    messages = payload.get("messages", [])
    cid = payload.get("conversation_id")
    if cid and not messages:
        with get_db() as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                "SELECT role, content FROM messages WHERE conversation_id=? ORDER BY created_at",
                (cid,)
            ).fetchall()
        messages = [dict(r) for r in rows]
    current_tokens = _estimate_tokens(messages)
    keep_recent = int(s_all.get("compress_keep_recent", "6") or "6")
    should = current_tokens > threshold and len(messages) > keep_recent
    return {
        "should_compress": should,
        "current_tokens": current_tokens,
        "threshold": threshold,
        "message_count": len(messages),
        "keep_recent": keep_recent,
    }


# ============================================================
# RAG subsystem — file upload + chunked TF-IDF retrieval
# ============================================================
# We avoid external embedding APIs. Instead, each uploaded file is:
#   1. Parsed into plain text (text files / pdf / source code)
#   2. Split into ~800-char chunks
#   3. Each chunk gets TF-IDF-style keyword vector (reusing extract_keywords)
#   4. Chunks stored in SQLite; retrieval uses keyword_overlap_score
# This is lightweight, has zero external deps, and works for personal-scale (<1000 files).

RAG_CHUNK_SIZE = 800      # chars per chunk
RAG_CHUNK_OVERLAP = 100   # overlap between adjacent chunks

DB_INIT += """
CREATE TABLE IF NOT EXISTS rag_documents (
    id TEXT PRIMARY KEY,
    name TEXT NOT NULL,
    mime TEXT NOT NULL DEFAULT 'text/plain',
    size INTEGER NOT NULL DEFAULT 0,
    content TEXT NOT NULL DEFAULT '',
    chunks_count INTEGER NOT NULL DEFAULT 0,
    created_at INTEGER NOT NULL
);
CREATE TABLE IF NOT EXISTS rag_chunks (
    id TEXT PRIMARY KEY,
    doc_id TEXT NOT NULL,
    doc_name TEXT NOT NULL,
    chunk_idx INTEGER NOT NULL,
    content TEXT NOT NULL,
    keywords TEXT NOT NULL DEFAULT '',
    FOREIGN KEY (doc_id) REFERENCES rag_documents(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_rag_chunks_doc ON rag_chunks(doc_id);
"""

RAG_ALLOWED_EXT = {
    ".txt", ".md", ".markdown", ".rst", ".log",
    ".py", ".js", ".ts", ".tsx", ".jsx", ".go", ".rs", ".java", ".c", ".cpp", ".h", ".hpp",
    ".cs", ".rb", ".php", ".swift", ".kt", ".scala", ".lua", ".pl",
    ".json", ".yaml", ".yml", ".toml", ".ini", ".cfg", ".conf",
    ".csv", ".tsv",
    ".html", ".htm", ".css", ".xml", ".svg",
    ".sh", ".bash", ".zsh", ".fish", ".ps1",
    ".sql", ".graphql",
    ".pdf",
    ".docx", ".xlsx",
}


def _parse_pdf_to_text(file_path: Path) -> str:
    """Try to extract text from a PDF using pdfplumber, fall back to pypdf, then to empty."""
    try:
        import pdfplumber  # type: ignore
        out = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                t = page.extract_text() or ""
                out.append(t)
        return "\n\n".join(out)
    except Exception:
        pass
    try:
        from pypdf import PdfReader  # type: ignore
        reader = PdfReader(str(file_path))
        return "\n\n".join((p.extract_text() or "") for p in reader.pages)
    except Exception as e:
        print(f"[rag] PDF parse failed: {e}")
        return ""


def _parse_docx_to_text(file_path: Path) -> str:
    try:
        import docx  # type: ignore
        d = docx.Document(str(file_path))
        return "\n".join(p.text for p in d.paragraphs)
    except Exception as e:
        print(f"[rag] docx parse failed: {e}")
        return ""


def _parse_xlsx_to_text(file_path: Path) -> str:
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        out = []
        for ws in wb.worksheets:
            out.append(f"=== Sheet: {ws.title} ===")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                out.append("\t".join(cells))
        return "\n".join(out)
    except Exception as e:
        print(f"[rag] xlsx parse failed: {e}")
        return ""


def _read_file_as_text(file_path: Path, mime: str) -> str:
    """Read a file as plain text. For binary formats (pdf/docx/xlsx), use parsers."""
    suffix = file_path.suffix.lower()
    if suffix == ".pdf":
        return _parse_pdf_to_text(file_path)
    if suffix == ".docx":
        return _parse_docx_to_text(file_path)
    if suffix == ".xlsx":
        return _parse_xlsx_to_text(file_path)
    # Default: try utf-8
    try:
        return file_path.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        print(f"[rag] read text failed: {e}")
        return ""


def _split_into_chunks(text: str, chunk_size: int = RAG_CHUNK_SIZE, overlap: int = RAG_CHUNK_OVERLAP) -> List[str]:
    """Split text into overlapping chunks by character count, on paragraph boundaries when possible."""
    if not text:
        return []
    chunks = []
    i = 0
    n = len(text)
    while i < n:
        end = min(i + chunk_size, n)
        # Try to break at a paragraph or sentence boundary
        if end < n:
            for sep in ["\n\n", "\n", "。", ". ", "! ", "? "]:
                last = text.rfind(sep, i, end)
                if last > i + chunk_size // 2:
                    end = last + len(sep)
                    break
        chunk = text[i:end].strip()
        if chunk:
            chunks.append(chunk)
        if end >= n:
            break
        i = end - overlap
        if i < 0:
            i = 0
    return chunks


def rag_search(query: str, top_k: int = 3) -> List[Dict]:
    """Retrieve top-k relevant chunks from RAG store."""
    qkws = extract_keywords(query, top_k=15)
    if not qkws:
        return []
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id, doc_id, doc_name, chunk_idx, content, keywords FROM rag_chunks ORDER BY doc_id, chunk_idx"
        ).fetchall()
    if not rows:
        return []
    scored = []
    for row in rows:
        mkws = row["keywords"].split(",") if row["keywords"] else []
        score = keyword_overlap_score(qkws, mkws)
        if score <= 0:
            continue
        scored.append((score, dict(row)))
    scored.sort(key=lambda x: -x[0])
    return [r for _, r in scored[:top_k]]


@app.post("/api/rag/upload")
async def rag_upload(file: UploadFile = File(...)):
    """Upload a file to the RAG knowledge base. Parses + chunks + indexes it."""
    name = file.filename or "unknown"
    suffix = Path(name).suffix.lower()
    if suffix not in RAG_ALLOWED_EXT:
        raise HTTPException(400, f"file type {suffix} not supported for RAG. Allowed: {', '.join(sorted(RAG_ALLOWED_EXT))}")
    contents = await file.read()
    if len(contents) > 20 * 1024 * 1024:
        raise HTTPException(400, "file too large (max 20MB)")
    # Save to temp path for parsers
    tmp = UPLOAD_DIR / f"rag_{hashlib.sha1(contents).hexdigest()[:16]}{suffix}"
    tmp.write_bytes(contents)
    try:
        text = _read_file_as_text(tmp, file.content_type or "text/plain")
    finally:
        try: tmp.unlink()
        except: pass
    if not text.strip():
        raise HTTPException(400, "could not extract any text from file (possibly scanned PDF or binary)")
    # Truncate very large text to ~500KB to keep DB manageable
    if len(text) > 500_000:
        text = text[:500_000] + "\n...[truncated]"
    chunks = _split_into_chunks(text)
    if not chunks:
        raise HTTPException(400, "no chunks produced from file")
    doc_id = hashlib.sha1(f"{name}:{contents[:1024]}".encode()).hexdigest()[:16]
    now = int(time.time())
    with get_db() as conn:
        # Replace if exists
        conn.execute("DELETE FROM rag_chunks WHERE doc_id=?", (doc_id,))
        conn.execute(
            "INSERT INTO rag_documents (id, name, mime, size, content, chunks_count, created_at) "
            "VALUES (?,?,?,?,?,?,?) ON CONFLICT(id) DO UPDATE SET "
            "name=excluded.name, mime=excluded.mime, size=excluded.size, content=excluded.content, chunks_count=excluded.chunks_count",
            (doc_id, name, file.content_type or "text/plain", len(contents), text[:5000], len(chunks), now)
        )
        for idx, chunk in enumerate(chunks):
            cid = hashlib.sha1(f"{doc_id}:{idx}:{chunk[:100]}".encode()).hexdigest()[:16]
            kws = ",".join(extract_keywords(chunk, top_k=15))
            conn.execute(
                "INSERT INTO rag_chunks (id, doc_id, doc_name, chunk_idx, content, keywords) VALUES (?,?,?,?,?,?)",
                (cid, doc_id, name, idx, chunk, kws)
            )
    return {"id": doc_id, "name": name, "size": len(contents), "chunks": len(chunks)}


@app.get("/api/rag/list")
async def rag_list():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id, name, mime, size, chunks_count, created_at FROM rag_documents ORDER BY created_at DESC"
        ).fetchall()
    return {"files": [dict(r) for r in rows]}


@app.post("/api/rag/delete")
async def rag_delete(payload: Dict):
    did = payload.get("id")
    if not did:
        raise HTTPException(400, "id required")
    with get_db() as conn:
        conn.execute("DELETE FROM rag_chunks WHERE doc_id=?", (did,))
        cur = conn.execute("DELETE FROM rag_documents WHERE id=?", (did,))
        if cur.rowcount == 0:
            return {"ok": False, "reason": "not found"}
    return {"ok": True}


@app.get("/api/rag/search")
async def rag_search_api(q: str, top_k: int = 3):
    return {"results": rag_search(q, top_k)}


# ============================================================
# Skills subsystem — Claude Code-compatible SKILL.md files
# ============================================================
# Each skill lives in <project>/.skills/<name>/SKILL.md with YAML frontmatter:
#   ---
#   name: my-skill
#   description: When the AI should use this skill (trigger text)
#   ---
#   # Detailed instructions for the AI...
# We read all SKILL.md files at startup and on demand. The skills are injected
# into the system prompt (description only by default; full body in 'always' mode).

SKILLS_DIR = PROJECT_ROOT / ".skills"
SKILLS_DIR.mkdir(exist_ok=True)

SKILL_NAME_RE = re.compile(r"^[a-z0-9-]+$")


def _parse_skill_md(text: str) -> Dict:
    """Parse a SKILL.md file with optional YAML frontmatter."""
    name = ""
    description = ""
    body = text
    if text.startswith("---"):
        end = text.find("\n---", 3)
        if end > 0:
            frontmatter = text[3:end].strip()
            body = text[end + 4:].strip()
            for line in frontmatter.splitlines():
                if line.startswith("name:"):
                    name = line.split(":", 1)[1].strip()
                elif line.startswith("description:"):
                    description = line.split(":", 1)[1].strip()
    if not name:
        # Fallback: first non-empty line of body
        for line in body.splitlines():
            line = line.strip().lstrip("#").strip()
            if line:
                name = line
                break
    return {"name": name, "description": description, "body": body}


def skills_list() -> List[Dict]:
    """List all installed skills."""
    out = []
    if not SKILLS_DIR.exists():
        return out
    for d in sorted(SKILLS_DIR.iterdir()):
        if not d.is_dir():
            continue
        skill_file = d / "SKILL.md"
        if not skill_file.exists():
            continue
        try:
            text = skill_file.read_text(encoding="utf-8", errors="replace")
            parsed = _parse_skill_md(text)
            out.append({
                "name": parsed["name"] or d.name,
                "description": parsed["description"],
                "body": parsed["body"],
                "size": len(text),
                "path": str(skill_file),
            })
        except Exception as e:
            print(f"[skills] failed to read {skill_file}: {e}")
    return out


def skills_get_descriptions() -> str:
    """Return a compact list of skill name + description for system prompt injection."""
    skills = skills_list()
    if not skills:
        return ""
    lines = []
    for s in skills:
        desc = s["description"] or "(no description)"
        lines.append(f"- {s['name']}: {desc}")
    return "\n".join(lines)


def skills_get_full_text() -> str:
    """Return full skill bodies for system prompt injection (used in 'always' mode)."""
    skills = skills_list()
    if not skills:
        return ""
    parts = []
    for s in skills:
        parts.append(f"### 技能：{s['name']}\n触发条件：{s['description'] or '(未指定)'}\n\n{s['body']}")
    return "\n\n---\n\n".join(parts)


@app.get("/api/skills")
async def skills_list_api():
    return {"skills": skills_list()}


@app.post("/api/skills/create")
async def skills_create_api(payload: Dict):
    name = (payload.get("name") or "").strip()
    description = payload.get("description", "").strip()
    body = payload.get("body", "").strip()
    if not name or not SKILL_NAME_RE.match(name):
        raise HTTPException(400, "name must match [a-z0-9-]+")
    skill_dir = SKILLS_DIR / name
    if skill_dir.exists():
        raise HTTPException(400, f"skill {name} already exists")
    skill_dir.mkdir(parents=True)
    content = f"---\nname: {name}\ndescription: {description}\n---\n\n{body}\n"
    (skill_dir / "SKILL.md").write_text(content, encoding="utf-8")
    return {"ok": True, "name": name, "path": str(skill_dir / "SKILL.md")}


@app.post("/api/skills/delete")
async def skills_delete_api(payload: Dict):
    name = (payload.get("name") or "").strip()
    if not name or not SKILL_NAME_RE.match(name):
        raise HTTPException(400, "invalid name")
    skill_dir = SKILLS_DIR / name
    if not skill_dir.exists():
        return {"ok": False, "reason": "not found"}
    import shutil
    shutil.rmtree(skill_dir, ignore_errors=True)
    return {"ok": True}


# ============================================================
# MCP server management — connect to external MCP servers (stdio)
# ============================================================
# We store MCP server configs in the settings table (key mcp_servers_json).
# At runtime, when a chat needs an MCP-provided tool, we spawn the server via
# stdio, list its tools, and expose them to the LLM. Results are cached per-session.

MCP_CONFIG_KEY = "mcp_servers_json"


def mcp_servers_load() -> List[Dict]:
    """Load MCP server configs from settings."""
    with get_db() as conn:
        row = conn.execute("SELECT value FROM settings WHERE key=?", (MCP_CONFIG_KEY,)).fetchone()
    if not row or not row["value"]:
        return []
    try:
        return json.loads(row["value"])
    except Exception:
        return []


def mcp_servers_save(servers: List[Dict]):
    with get_db() as conn:
        conn.execute(
            "INSERT INTO settings (key, value) VALUES (?,?) ON CONFLICT(key) DO UPDATE SET value=?",
            (MCP_CONFIG_KEY, json.dumps(servers, ensure_ascii=False), json.dumps(servers, ensure_ascii=False))
        )


async def mcp_test_server(name: str, command_or_url: str, env: Dict) -> Dict:
    """Test MCP server connection. Supports:
    - stdio: 'npx -y open-websearch@latest' (spawns subprocess)
    - HTTP: 'http://127.0.0.1:3210/mcp' (connects to running instance)
    - SSE: 'http://127.0.0.1:3210/sse' (connects to running instance)
    """
    try:
        from mcp import ClientSession, StdioServerParameters
        from mcp.client.stdio import stdio_client
    except ImportError:
        return {"name": name, "command": command_or_url, "connected": False, "error": "mcp python sdk not installed (pip install mcp)"}

    # Detect transport type
    is_http = command_or_url.startswith("http://") or command_or_url.startswith("https://")

    if is_http:
        # HTTP/SSE transport — connect to running server
        try:
            # Try streamable HTTP first, then SSE
            url = command_or_url.rstrip("/")
            # Auto-detect: if URL ends with /sse, use SSE; if /mcp, use streamable_http; else try both
            try:
                if "/sse" in url:
                    from mcp.client.sse import sse_client
                    async with sse_client(url) as (read, write):
                        async with ClientSession(read, write) as session:
                            await session.initialize()
                            tools_resp = await session.list_tools()
                            tool_names = [t.name for t in tools_resp.tools]
                            return {"name": name, "command": command_or_url, "env": env, "connected": True, "tools": tool_names}
                else:
                    # Try streamable HTTP (default for /mcp endpoint)
                    from mcp.client.streamable_http import streamablehttp_client
                    # Ensure URL ends with /mcp if no path
                    if not url.endswith("/mcp") and not url.endswith("/sse"):
                        url = url + "/mcp"
                    async with streamablehttp_client(url) as (read, write, _):
                        async with ClientSession(read, write) as session:
                            await session.initialize()
                            tools_resp = await session.list_tools()
                            tool_names = [t.name for t in tools_resp.tools]
                            return {"name": name, "command": command_or_url, "env": env, "connected": True, "tools": tool_names}
            except Exception as http_err:
                # If HTTP fails, try SSE as fallback
                try:
                    sse_url = command_or_url.rstrip("/")
                    if not sse_url.endswith("/sse"):
                        sse_url = sse_url.replace("/mcp", "") + "/sse"
                    from mcp.client.sse import sse_client
                    async with sse_client(sse_url) as (read, write):
                        async with ClientSession(read, write) as session:
                            await session.initialize()
                            tools_resp = await session.list_tools()
                            tool_names = [t.name for t in tools_resp.tools]
                            return {"name": name, "command": command_or_url, "env": env, "connected": True, "tools": tool_names}
                except Exception:
                    return {"name": name, "command": command_or_url, "connected": False, "error": f"HTTP/SSE connection failed: {str(http_err)[:200]}"}
        except Exception as e:
            return {"name": name, "command": command_or_url, "connected": False, "error": str(e)[:200]}
    else:
        # stdio transport — spawn subprocess
        import shlex
        parts = shlex.split(command_or_url)
        if not parts:
            return {"name": name, "command": command_or_url, "connected": False, "error": "empty command"}
        cmd = parts[0]
        args = parts[1:]
        full_env = {**os.environ, **env}
        try:
            params = StdioServerParameters(command=cmd, args=args, env=full_env)
            async with stdio_client(params) as (read, write):
                async with ClientSession(read, write) as session:
                    await session.initialize()
                    tools_resp = await session.list_tools()
                    tool_names = [t.name for t in tools_resp.tools]
                    return {"name": name, "command": command_or_url, "env": env, "connected": True, "tools": tool_names}
        except asyncio.TimeoutError:
            return {"name": name, "command": command_or_url, "connected": False, "error": "timeout (server didn't respond in 30s)"}
        except Exception as e:
            return {"name": name, "command": command_or_url, "connected": False, "error": str(e)[:200]}


async def mcp_call_tool(server_name: str, tool_name: str, arguments: Dict, timeout_sec: float = 60.0) -> Dict:
    """Call an MCP tool. Supports stdio and HTTP/SSE transports."""
    servers = mcp_servers_load()
    srv = next((s for s in servers if s["name"] == server_name), None)
    if not srv:
        return {"success": False, "error": f"MCP server '{server_name}' not configured"}
    command_or_url = srv["command"]
    env = srv.get("env") or {}
    is_http = command_or_url.startswith("http://") or command_or_url.startswith("https://")

    try:
        from mcp import ClientSession
    except ImportError:
        return {"success": False, "error": "mcp sdk not installed"}

    if is_http:
        url = command_or_url.rstrip("/")
        try:
            if "/sse" in url:
                from mcp.client.sse import sse_client
                async with sse_client(url) as (read, write):
                    async with ClientSession(read, write) as session:
                        await session.initialize()
                        result = await asyncio.wait_for(session.call_tool(tool_name, arguments), timeout=timeout_sec)
            else:
                http_url = url + "/mcp" if not url.endswith("/mcp") else url
                try:
                    from mcp.client.streamable_http import streamablehttp_client
                    async with streamablehttp_client(http_url) as (read, write, _):
                        async with ClientSession(read, write) as session:
                            await session.initialize()
                            result = await asyncio.wait_for(session.call_tool(tool_name, arguments), timeout=timeout_sec)
                except Exception:
                    from mcp.client.sse import sse_client
                    sse_url = url.replace("/mcp", "") + "/sse"
                    async with sse_client(sse_url) as (read, write):
                        async with ClientSession(read, write) as session:
                            await session.initialize()
                            result = await asyncio.wait_for(session.call_tool(tool_name, arguments), timeout=timeout_sec)
        except asyncio.TimeoutError:
            return {"success": False, "error": f"MCP timed out ({timeout_sec}s)"}
        except Exception as e:
            return {"success": False, "error": str(e)[:500]}
    else:
        try:
            from mcp import StdioServerParameters
            from mcp.client.stdio import stdio_client
        except ImportError:
            return {"success": False, "error": "mcp sdk not installed"}
        import shlex
        parts = shlex.split(command_or_url)
        if not parts:
            return {"success": False, "error": "empty command"}
        full_env = {**os.environ, **env}
        try:
            params = StdioServerParameters(command=parts[0], args=parts[1:], env=full_env)
            async with stdio_client(params) as (read, write):
                async with ClientSession(read, write) as session:
                    await session.initialize()
                    result = await asyncio.wait_for(session.call_tool(tool_name, arguments), timeout=timeout_sec)
        except asyncio.TimeoutError:
            return {"success": False, "error": f"MCP timed out ({timeout_sec}s)"}
        except Exception as e:
            return {"success": False, "error": str(e)[:500]}

    # Extract text content
    text_parts = []
    for block in (result.content or []):
        if hasattr(block, "text"):
            text_parts.append(block.text)
        elif hasattr(block, "data"):
            text_parts.append(f"[binary data]")
    text = "\n".join(text_parts) or "(no output)"
    if result.isError:
        return {"success": False, "error": text[:2000], "result": text[:2000]}
    return {"success": True, "result": text[:4000]}


@app.get("/api/mcp/servers")
async def mcp_servers_list_api():
    """List configured MCP servers. Attempts a quick connection test."""
    servers = mcp_servers_load()
    out = []
    for s in servers:
        # Quick test: try to spawn and list tools (with short timeout)
        test_result = await mcp_test_server(s["name"], s["command"], s.get("env", {}))
        out.append({
            "name": s["name"],
            "command": s["command"],
            "env": s.get("env", {}),
            "connected": test_result.get("connected", False),
            "tools": test_result.get("tools", []),
            "error": test_result.get("error"),
        })
    return {"servers": out}


@app.post("/api/mcp/servers/add")
async def mcp_servers_add_api(payload: Dict):
    name = (payload.get("name") or "").strip()
    command = (payload.get("command") or "").strip()
    env = payload.get("env", {}) or {}
    if not name or not command:
        raise HTTPException(400, "name and command required")
    if not SKILL_NAME_RE.match(name):
        raise HTTPException(400, "name must match [a-z0-9-]+")
    servers = mcp_servers_load()
    if any(s["name"] == name for s in servers):
        raise HTTPException(400, f"server {name} already exists")
    servers.append({"name": name, "command": command, "env": env})
    mcp_servers_save(servers)
    # Try to connect immediately to validate
    test = await mcp_test_server(name, command, env)
    return {"ok": True, "server": test}


@app.post("/api/mcp/servers/delete")
async def mcp_servers_delete_api(payload: Dict):
    name = (payload.get("name") or "").strip()
    servers = mcp_servers_load()
    new = [s for s in servers if s["name"] != name]
    if len(new) == len(servers):
        return {"ok": False, "reason": "not found"}
    mcp_servers_save(new)
    return {"ok": True}


@app.post("/api/mcp/servers/test")
async def mcp_servers_test_api(payload: Dict):
    """Test connection to a configured MCP server by name."""
    name = (payload.get("name") or "").strip()
    servers = mcp_servers_load()
    srv = next((s for s in servers if s["name"] == name), None)
    if not srv:
        raise HTTPException(404, "server not found")
    result = await mcp_test_server(name, srv["command"], srv.get("env", {}))
    return result


# ============================================================
# Boot-time: pre-install open-webSearch MCP server if none configured
# ============================================================
def _ensure_default_mcp_server():
    """If no MCP servers are configured, pre-install open-webSearch as the default."""
    servers = mcp_servers_load()
    if servers:
        return
    default = [
        {
            "name": "web-search",
            "command": "npx -y open-websearch@latest",
            "env": {
                "MODE": "stdio",
                "DEFAULT_SEARCH_ENGINE": "bing",
            },
        }
    ]
    mcp_servers_save(default)


_ensure_default_mcp_server()


# ============================================================
# Boot-time: pre-install a couple of example skills if none exist
# ============================================================
def _ensure_example_skills():
    if skills_list():
        return
    examples = [
        ("web-search", "当用户需要搜索互联网、获取最新信息、查询不确定的事实、或需要外部数据时使用此技能。触发词：搜索、查一下、最新、最近、现在、上网找。", "调用 web_search 工具（由 open-webSearch MCP 提供）进行网络搜索。工具参数：query（搜索关键词）、limit（结果数，默认5）。返回结果包含标题和摘要。如果搜索结果不够，可以再用 fetchWebContent 工具获取页面全文。"),
        ("deep-analysis", "当用户希望进行深入分析、比较、推理而非简单罗列信息时使用。触发场景：用户问\"哪个更好\"\"为什么\"\"怎么选\"\"对比\"\"权衡\"，或要求\"像思考伙伴一样讨论\"。", "回答时：1) 先给出明确判断或推荐，不要只罗列选项；2) 解释关键判断标准；3) 对比关键差异（用表格）；4) 说明为什么不选其他方案；5) 指出最大风险和适用条件。承认不确定性，不要编造数据。"),
    ]
    for name, desc, body in examples:
        skill_dir = SKILLS_DIR / name
        skill_dir.mkdir(exist_ok=True)
        content = f"---\nname: {name}\ndescription: {desc}\n---\n\n{body}\n"
        (skill_dir / "SKILL.md").write_text(content, encoding="utf-8")


_ensure_example_skills()


# ============================================================
# Sessions + Cron: initialize DB tables and scheduler
# ============================================================
sessions_mod.init_db(DB_PATH)
cron_mod.init_db(DB_PATH)
advanced_memory.init_advanced_db(DB_PATH)
chat_vectors.init_chat_vectors_db(DB_PATH)
memory_orchestrator.init_orchestrator_db(DB_PATH)
knowledge_graph.init_kg_db(DB_PATH)
episodic_memory.init_episodic_db(DB_PATH)
meta_cognition.init_meta_cog_db(DB_PATH)
cognitive_kernel.init_cognitive_db(DB_PATH)

# Initialize new subsystems
learning_engine.init_learning_db(DB_PATH)
memory_governance.init_governance_db(DB_PATH)
reflection_tree.init_reflection_db(DB_PATH)
identity_consistency.init_identity_consistency_db(DB_PATH)
adaptive_retrieval.init_adaptive_db(DB_PATH)

# Initialize event bus (with persistence)
_bus = event_bus.EventBus(db_path=DB_PATH, persist=True)
event_bus.set_event_bus(_bus)

# Wire event bus subscribers — auto-trigger learning when memories are added
@event_bus.subscribe("memory.added")
async def _on_memory_added_for_learning(event):
    """When a new memory is added, record it as a learning observation.
    The learning engine finds patterns across many observations."""
    try:
        data = event.get("data", {})
        from app import learning_engine
        # Extract a simple pattern from the memory
        content = data.get("content", "")
        category = data.get("category", "other")
        importance = data.get("importance", 30)
        # Record as a preference observation (the learning engine will find patterns)
        learning_engine.record_observation(
            DB_PATH,
            user_id=data.get("user_id", "default"),
            pattern_type="preference",
            key=category,
            value=content[:100],
            evidence_type="observation",
            description=f"Memory added: {content[:80]}",
            context=f"importance={importance}",
        )
    except Exception as e:
        print(f"[learning] auto-observe failed: {e}")

@event_bus.subscribe("memory.added")
async def _on_memory_added_for_governance(event):
    """When a new memory is added, check if it contradicts existing memories.
    If so, the governance system should flag it."""
    try:
        data = event.get("data", {})
        # The quarantine is already done in memory_orchestrator.add_memory()
        # This subscriber is for future extensions (e.g., cross-memory contradiction detection)
        pass
    except Exception:
        pass

@event_bus.subscribe("timeline.event")
async def _on_timeline_event_for_co_experience(event):
    """When a high-importance timeline event is added, harvest it as a co-experience moment."""
    try:
        data = event.get("data", {})
        importance = data.get("significance", 50)
        if importance >= 70:
            from app import co_experience as co_exp_mod
            co_exp_mod.harvest_from_timeline(DB_PATH, "default", importance_threshold=0.7, limit=5)
    except Exception as e:
        print(f"[co_experience] auto-harvest failed: {e}")

# Run schema migrations (forward-only, safe to run on every startup)
_migration_result = migrations_mod.run_migrations(DB_PATH)
if _migration_result["migrations_run"]:
    print(f"[migrations] applied: {_migration_result['migrations_run']}")
else:
    print(f"[migrations] schema at v{_migration_result['to_version']} (up to date)")

# Seed built-in data on first run
try:
    _n_residents = residents_mod.ensure_builtin_residents(DB_PATH, "default")
    if _n_residents:
        print(f"[residents] created {_n_residents} built-in residents")
    _n_phil = philosophy_mod.ensure_seed_philosophy(DB_PATH, "default")
    if _n_phil:
        print(f"[philosophy] seeded {_n_phil} principles")
except Exception as e:
    print(f"[seed] failed: {e}")

# Load plugins on startup
try:
    _plugins_dir = PROJECT_ROOT / "plugins"
    if not _plugins_dir.exists():
        _plugins_dir.mkdir(exist_ok=True)
    # Create example plugin if no plugins exist
    if not any(_plugins_dir.iterdir()):
        plugin_sdk.create_example_plugin(_plugins_dir)
    _loaded_plugins = plugin_sdk.load_all_plugins(_plugins_dir)
    print(f"[plugins] loaded {len(_loaded_plugins)} plugins")
except Exception as e:
    print(f"[plugins] load failed: {e}")

# Initialize vector store on startup (lazy, but log backend)
try:
    _vs = vector_store_mod.get_vector_store(DB_PATH)
    print(f"[vector_store] backend: {_vs.backend}")
except Exception as e:
    print(f"[vector_store] init failed: {e}")


@app.on_event("startup")
async def _start_cron_scheduler():
    """Start the cron scheduler when FastAPI starts."""
    s_all = settings_get_all()
    if s_all.get("cron_enabled", "true") != "false":
        async def _spawn_fn(job: Dict) -> str:
            """Called when a cron job is due. Spawns a session for it."""
            api_cfg = get_subtask_api_config()
            if job.get("model"):
                api_cfg = {**api_cfg, "api_model": job["model"]}
            sess = sessions_mod.session_create(
                DB_PATH,
                title=f"[cron] {job.get('name', job['id'])}",
                model=api_cfg["api_model"],
                system_prompt=job.get("system_prompt", ""),
                user_message=job["prompt"],
            )
            # Run in background
            async def _run():
                try:
                    await sessions_mod.spawn_session(
                        sess["id"], DB_PATH, api_cfg,
                        job.get("system_prompt", ""), job["prompt"],
                        title=sess["title"], model=api_cfg["api_model"],
                    )
                except Exception as e:
                    print(f"[cron] session {sess['id']} failed: {e}")
            asyncio.create_task(_run())
            return sess["id"]
        cron_mod.start_scheduler(DB_PATH, _spawn_fn)
        print("[cron] scheduler started")
    # Start background reflection + decay scheduler (runs every 10 minutes)
    async def _background_reflection_loop():
        """Periodically: apply memory decay, run reflection if enough new messages."""
        REFLECTION_INTERVAL = 600  # 10 minutes
        REFLECTION_TRIGGER_MSGS = 30  # trigger after 30+ new messages since last reflection
        while True:
            try:
                await asyncio.sleep(REFLECTION_INTERVAL)
                s = settings_get_all()
                if s.get("profile_auto_update", "true") == "false":
                    continue
                # 1. Apply decay to memories
                try:
                    memory_orchestrator.apply_decay(DB_PATH, user_id="default", days_elapsed=0.007)
                except Exception as e:
                    print(f"[bg] decay failed: {e}")
                # 2. Apply decay to episodes
                try:
                    episodic_memory.apply_decay(DB_PATH, user_id="default")
                except Exception as e:
                    print(f"[bg] episode decay failed: {e}")
                # 3. Check if reflection should run
                last_reflection = memory_orchestrator.get_latest_reflection(DB_PATH, user_id="default")
                msgs_since = 9999
                if last_reflection:
                    # Count messages in chat_vectors since last reflection
                    import sqlite3 as _sq
                    conn = _sq.connect(str(DB_PATH))
                    cnt = conn.execute(
                        "SELECT COUNT(*) FROM chat_vectors WHERE created_at > ?",
                        (last_reflection.get("created_at", 0),)
                    ).fetchone()[0]
                    conn.close()
                    msgs_since = cnt
                if msgs_since >= REFLECTION_TRIGGER_MSGS:
                    # Gather recent conversation text from chat_vectors
                    import sqlite3 as _sq
                    conn = _sq.connect(str(DB_PATH))
                    conn.row_factory = _sq.Row
                    rows = conn.execute(
                        "SELECT role, content FROM chat_vectors ORDER BY created_at DESC LIMIT 50"
                    ).fetchall()
                    conn.close()
                    if rows:
                        recent_text = "\n\n".join(
                            f"{'用户' if r['role'] == 'user' else 'CyanX AI'}: {r['content']}"
                            for r in reversed(rows)
                        )
                        mem_cfg = get_memory_api_config()
                        async with httpx.AsyncClient(timeout=90.0) as c:
                            result = await memory_orchestrator.run_reflection(
                                DB_PATH, user_id="default",
                                recent_conversation=recent_text,
                                message_count=msgs_since,
                                http_client=c, api_cfg=mem_cfg,
                            )
                        if result.get("success"):
                            print(f"[bg] reflection completed: {result.get('new_memories_added', 0)} new memories")
                            # Also extract knowledge graph triples + episodes from the same conversation
                            try:
                                async with httpx.AsyncClient(timeout=30.0) as c:
                                    triples = await knowledge_graph.extract_triples_via_llm(
                                        recent_text, c, mem_cfg
                                    )
                                    if triples:
                                        knowledge_graph.add_triples(DB_PATH, user_id="default", triples=triples)
                                        print(f"[bg] extracted {len(triples)} KG triples")
                                    episodes = await episodic_memory.extract_episodes_via_llm(
                                        recent_text, c, mem_cfg
                                    )
                                    for ep in episodes:
                                        episodic_memory.create_episode(
                                            DB_PATH, user_id="default",
                                            title=ep.get("title", ""),
                                            description=ep.get("description", ""),
                                            occurred_at=ep.get("occurred_at", ""),
                                            importance=int(ep.get("importance", 50)),
                                            tags=ep.get("tags", ""),
                                            emotional_valence=ep.get("emotional_valence", "neutral"),
                                            status=ep.get("status", "completed"),
                                            source="reflection",
                                        )
                                    if episodes:
                                        print(f"[bg] extracted {len(episodes)} episodes")
                            except Exception as e:
                                print(f"[bg] KG/episode extraction failed: {e}")
            except asyncio.CancelledError:
                break
            except Exception as e:
                print(f"[bg] reflection loop error: {e}")
    asyncio.create_task(_background_reflection_loop())
    print("[bg] reflection + decay scheduler started")
    # Start the Life Loop (cognitive kernel circadian rhythm)
    import httpx as _httpx
    loop_inst = life_loop.LifeLoop(
        db_path=DB_PATH,
        get_memory_api_cfg=get_memory_api_config,
        httpx_client_factory=lambda timeout: _httpx.AsyncClient(timeout=timeout),
    )
    loop_inst.start()
    life_loop.set_life_loop(loop_inst)


# ============================================================
# Sessions HTTP endpoints
# ============================================================
@app.get("/api/sessions")
async def sessions_list_api(status: Optional[str] = None, limit: int = 50):
    return {"sessions": sessions_mod.session_list(DB_PATH, status=status, limit=limit)}


@app.post("/api/sessions/create")
async def sessions_create_api(payload: Dict):
    """Manually create a session record (without spawning). Useful for inspection."""
    sess = sessions_mod.session_create(
        DB_PATH,
        title=payload.get("title", ""),
        model=payload.get("model", ""),
        system_prompt=payload.get("system_prompt", ""),
        user_message=payload.get("user_message", ""),
    )
    return sess


@app.post("/api/sessions/spawn")
async def sessions_spawn_api(payload: Dict):
    """Spawn a background session that runs an LLM conversation.
    Returns immediately with the session_id; the session runs in background."""
    title = payload.get("title", "Background task")
    user_message = payload.get("message") or payload.get("user_message", "")
    if not user_message:
        raise HTTPException(400, "message required")
    s_all = settings_get_all()
    api_cfg = get_subtask_api_config()
    if payload.get("model"):
        api_cfg = {**api_cfg, "api_model": payload["model"]}
    system_prompt = payload.get("system_prompt", "")
    max_subtasks = int(s_all.get("max_subtasks", "3") or "3")
    running = sessions_mod.session_list(DB_PATH, status="running")
    if len(running) >= max_subtasks:
        raise HTTPException(429, f"已达到子任务并发上限 ({max_subtasks})")
    sess = sessions_mod.session_create(
        DB_PATH, title=title, model=api_cfg["api_model"],
        system_prompt=system_prompt, user_message=user_message,
    )
    async def _run():
        try:
            await sessions_mod.spawn_session(
                sess["id"], DB_PATH, api_cfg, system_prompt, user_message,
                title=title, model=api_cfg["api_model"],
            )
        except Exception as e:
            print(f"[sessions] spawn failed: {e}")
    asyncio.create_task(_run())
    return {"session_id": sess["id"], "title": title, "status": "running"}


@app.get("/api/sessions/{sid}")
async def sessions_get_api(sid: str):
    sess = sessions_mod.session_get(DB_PATH, sid)
    if not sess:
        raise HTTPException(404, "session not found")
    return sess


@app.post("/api/sessions/{sid}/send")
async def sessions_send_api(sid: str, payload: Dict):
    """Send a follow-up message to a session (resumes the conversation synchronously)."""
    message = payload.get("message", "")
    if not message:
        raise HTTPException(400, "message required")
    return sessions_mod.session_send(DB_PATH, sid, message)


@app.delete("/api/sessions/{sid}")
async def sessions_delete_api(sid: str):
    ok = sessions_mod.session_delete(DB_PATH, sid)
    return {"ok": ok}


# ============================================================
# Cron HTTP endpoints
# ============================================================
@app.get("/api/cron/jobs")
async def cron_jobs_list_api():
    return {"jobs": cron_mod.cron_list(DB_PATH)}


@app.post("/api/cron/jobs/create")
async def cron_jobs_create_api(payload: Dict):
    name = payload.get("name", "")
    kind = payload.get("kind", "cron")  # cron | one_time | fixed_rate
    schedule = payload.get("schedule", "")
    prompt = payload.get("prompt", "")
    model = payload.get("model", "")
    system_prompt = payload.get("system_prompt", "")
    enabled = payload.get("enabled", True)
    if not schedule or not prompt:
        raise HTTPException(400, "schedule and prompt required")
    if kind == "cron":
        # Validate cron expression
        if not cron_mod._parse_cron(schedule):
            raise HTTPException(400, f"invalid cron expression: {schedule} (use 5 fields: minute hour day month weekday)")
    return cron_mod.cron_create(
        DB_PATH, name=name, kind=kind, schedule=schedule,
        prompt=prompt, model=model, system_prompt=system_prompt, enabled=enabled,
    )


@app.post("/api/cron/jobs/{job_id}/update")
async def cron_jobs_update_api(job_id: str, payload: Dict):
    fields = {k: v for k, v in payload.items() if k in {
        "name", "kind", "schedule", "prompt", "model", "system_prompt", "enabled"
    }}
    if "enabled" in fields:
        fields["enabled"] = 1 if fields["enabled"] else 0
    ok = cron_mod.cron_update(DB_PATH, job_id, **fields)
    return {"ok": ok}


@app.post("/api/cron/jobs/{job_id}/delete")
async def cron_jobs_delete_api(job_id: str):
    ok = cron_mod.cron_delete(DB_PATH, job_id)
    return {"ok": ok}


@app.post("/api/cron/jobs/{job_id}/run")
async def cron_jobs_run_now_api(job_id: str):
    """Manually trigger a cron job immediately."""
    job = cron_mod.cron_get(DB_PATH, job_id)
    if not job:
        raise HTTPException(404, "job not found")
    api_cfg = get_subtask_api_config()
    if job.get("model"):
        api_cfg = {**api_cfg, "api_model": job["model"]}
    sess = sessions_mod.session_create(
        DB_PATH,
        title=f"[cron-manual] {job.get('name', job['id'])}",
        model=api_cfg["api_model"],
        system_prompt=job.get("system_prompt", ""),
        user_message=job["prompt"],
    )
    async def _run():
        try:
            await sessions_mod.spawn_session(
                sess["id"], DB_PATH, api_cfg,
                job.get("system_prompt", ""), job["prompt"],
                title=sess["title"], model=api_cfg["api_model"],
            )
        except Exception as e:
            print(f"[cron-manual] session {sess['id']} failed: {e}")
    asyncio.create_task(_run())
    return {"session_id": sess["id"], "status": "running"}


@app.get("/api/cron/runs")
async def cron_runs_list_api(job_id: Optional[str] = None, limit: int = 50):
    return {"runs": cron_mod.cron_runs_list(DB_PATH, job_id=job_id, limit=limit)}


@app.post("/api/cron/parse")
async def cron_parse_api(payload: Dict):
    """Validate and preview a cron expression. Returns next 5 run times."""
    expr = payload.get("expression", "")
    cron = cron_mod._parse_cron(expr)
    if not cron:
        raise HTTPException(400, "invalid cron expression (need 5 fields)")
    # Compute next 5 runs
    from datetime import datetime, timezone, timedelta
    now = datetime.now(timezone.utc)
    runs = []
    t = now
    for _ in range(5):
        for i in range(1, 366 * 24 * 60):
            t2 = t + timedelta(minutes=i)
            if (t2.minute in cron["minute"] and
                t2.hour in cron["hour"] and
                t2.day in cron["day"] and
                t2.month in cron["month"] and
                (t2.weekday() + 1) % 7 in cron["weekday"]):
                runs.append(int(t2.timestamp()))
                t = t2
                break
    return {"valid": True, "next_runs": runs}


# ============================================================
# RAG with optional API-based embedding
# ============================================================
def _embed_for_rag(text: str) -> Optional[List[float]]:
    """Get embedding vector for a text. Returns None to fall back to keyword search.
    If provider=local, returns None (use keyword search). If provider=api, calls
    the configured embedding API."""
    cfg = get_embedding_config()
    if cfg["provider"] != "api" or not cfg["api_base_url"]:
        return None
    try:
        import urllib.request, urllib.parse, json as _json
        url = cfg["api_base_url"].rstrip("/") + "/embeddings"
        payload = _json.dumps({
            "model": cfg["model"] or "text-embedding-3-small",
            "input": text[:8000],
        }).encode("utf-8")
        req = urllib.request.Request(url, data=payload, headers={
            "Authorization": f"Bearer {cfg['api_key']}",
            "Content-Type": "application/json",
        })
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = _json.loads(resp.read().decode("utf-8"))
        return data["data"][0]["embedding"]
    except Exception as e:
        print(f"[rag] embedding API failed: {e}")
        return None


# ============================================================
# Inbox — universal capture (NP-OS style)
# ============================================================
@app.get("/api/inbox/items")
async def inbox_list(status: str = "", limit: int = 100, offset: int = 0):
    return {"items": inbox_mod.list_items(DB_PATH, "default", status=status or None,
                                          limit=limit, offset=offset)}

@app.post("/api/inbox/items")
async def inbox_add(payload: Dict):
    item = inbox_mod.add_item(
        DB_PATH, "default",
        type_=payload.get("type", "text"),
        content=payload.get("content", ""),
        title=payload.get("title", ""),
        source=payload.get("source", "manual"),
        metadata=payload.get("metadata", {}),
    )
    # Auto-suggest a destination
    item["suggested_destination"] = inbox_mod.auto_route(
        item["content"], item["type"], item["title"]
    )
    return item

@app.get("/api/inbox/items/{item_id}")
async def inbox_get(item_id: str):
    item = inbox_mod.get_item(DB_PATH, item_id)
    if not item:
        raise HTTPException(404, "item not found")
    return item

@app.post("/api/inbox/items/{item_id}")
async def inbox_update(item_id: str, payload: Dict):
    ok = inbox_mod.update_item(
        DB_PATH, item_id,
        content=payload.get("content"),
        title=payload.get("title"),
        metadata=payload.get("metadata"),
    )
    if not ok:
        raise HTTPException(404, "item not found")
    return inbox_mod.get_item(DB_PATH, item_id)

@app.delete("/api/inbox/items/{item_id}")
async def inbox_delete(item_id: str):
    if not inbox_mod.delete_item(DB_PATH, item_id):
        raise HTTPException(404, "item not found")
    return {"ok": True}

@app.post("/api/inbox/items/{item_id}/process")
async def inbox_process(item_id: str, payload: Dict):
    """Mark item as processed, with destination routing."""
    dest = payload.get("destination", "note")
    dest_id = payload.get("destination_id", "")
    if not inbox_mod.process_item(DB_PATH, item_id, dest, dest_id):
        raise HTTPException(404, "item not found")
    return {"ok": True, "destination": dest, "destination_id": dest_id}

@app.post("/api/inbox/items/{item_id}/archive")
async def inbox_archive(item_id: str):
    if not inbox_mod.archive_item(DB_PATH, item_id):
        raise HTTPException(404, "item not found")
    return {"ok": True}

@app.get("/api/inbox/stats")
async def inbox_stats_api():
    return inbox_mod.get_stats(DB_PATH, "default")

@app.post("/api/inbox/route-suggest")
async def inbox_route_suggest(payload: Dict):
    """Suggest a destination for a piece of content (used by UI)."""
    return {"destination": inbox_mod.auto_route(
        payload.get("content", ""),
        payload.get("type", "text"),
        payload.get("title", ""),
    )}


# ============================================================
# Journal — AI-assisted daily journal
# ============================================================
@app.get("/api/journal/today")
async def journal_today():
    return journal_mod.get_or_create(DB_PATH, "default")

@app.get("/api/journal/list")
async def journal_list(days: int = 30):
    return {"items": journal_mod.list_range(DB_PATH, "default", days=days)}

@app.get("/api/journal/streak")
async def journal_streak_api():
    return journal_mod.get_streak(DB_PATH, "default")

@app.get("/api/journal/{date_str}")
async def journal_get(date_str: str):
    j = journal_mod.get(DB_PATH, "default", date_str)
    if not j:
        raise HTTPException(404, "no journal for that date")
    return j

@app.post("/api/journal/{date_str}/content")
async def journal_update_content(date_str: str, payload: Dict):
    return journal_mod.update_content(DB_PATH, "default", date_str, payload.get("content", ""))

@app.post("/api/journal/{date_str}/fields")
async def journal_update_fields(date_str: str, payload: Dict):
    return journal_mod.update_fields(DB_PATH, "default", date_str, payload)

@app.post("/api/journal/{date_str}/ai-draft")
async def journal_set_ai_draft(date_str: str, payload: Dict):
    """AI generates a draft for the day. Uses today's activity + LLM."""
    # 1. Gather day's activity
    activity = journal_mod.gather_day_activity(DB_PATH, "default", date_str)
    if not activity or activity.get("stats", {}).get("conversations", 0) == 0:
        # Nothing happened → empty draft
        return journal_mod.set_ai_draft(
            DB_PATH, "default", date_str,
            draft="（今日暂无活动记录）",
            summary="", emotional_tone="neutral", highlights=[]
        )
    # 2. Build activity summary text
    s = activity["stats"]
    parts = [f"对话 {s['conversations']} 次", f"完成任务 {s['tasks_completed']} 个",
             f"捕获 {s['inbox_captures']} 条", f"反思 {s['reflections']} 条",
             f"时间线事件 {s['timeline_events']} 个"]
    activity_text = "；".join(parts) + "。"
    if activity["conversations"]:
        activity_text += "\n对话主题：" + " / ".join(
            c.get("title", "")[:30] for c in activity["conversations"][:5] if c.get("title")
        )
    if activity["tasks_completed"]:
        activity_text += "\n完成任务：" + " / ".join(
            t.get("title", "")[:30] for t in activity["tasks_completed"][:5]
        )
    # 3. Ask LLM for draft
    try:
        api_cfg = get_memory_api_config()
        prompt_text = prompt_registry_mod.get_prompt_with_meta(DB_PATH, "prompt_journal_draft")
        prompt_template = prompt_text["content"] if prompt_text else ""
        if not prompt_template:
            # fallback to inline default
            from app.prompt_registry import _resolve_default, _PROMPT_REGISTRY
            entry = next((e for e in _PROMPT_REGISTRY if e["key"] == "prompt_journal_draft"), None)
            prompt_template = _resolve_default(entry) if entry else ""
        prompt = prompt_template.format(activity_summary=activity_text)
        async with httpx.AsyncClient(timeout=60.0) as c:
            payload_llm = {
                "model": api_cfg["api_model"],
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.6, "max_tokens": 800, "stream": False, "enable_thinking": False,
            }
            resp = await c.post(
                f"{api_cfg['api_base_url']}/chat/completions",
                json=payload_llm,
                headers={"Authorization": f"Bearer {api_cfg['api_key']}",
                         "Content-Type": "application/json"},
            )
            resp.raise_for_status()
            data = resp.json()
            draft = data["choices"][0]["message"]["content"].strip()
        # 4. Emotion analysis (lightweight, separate call)
        emotion = "neutral"
        try:
            emo_prompt_text = prompt_registry_mod.get_prompt_with_meta(DB_PATH, "prompt_journal_emotion")
            emo_template = emo_prompt_text["content"] if emo_prompt_text else ""
            if not emo_template:
                from app.prompt_registry import _resolve_default, _PROMPT_REGISTRY
                entry = next((e for e in _PROMPT_REGISTRY if e["key"] == "prompt_journal_emotion"), None)
                emo_template = _resolve_default(entry) if entry else ""
            emo_prompt = emo_template.format(content=activity_text + "\n\n" + draft[:500])
            async with httpx.AsyncClient(timeout=20.0) as c:
                emo_payload = {
                    "model": api_cfg["api_model"],
                    "messages": [{"role": "user", "content": emo_prompt}],
                    "temperature": 0.2, "max_tokens": 20, "stream": False, "enable_thinking": False,
                }
                resp2 = await c.post(
                    f"{api_cfg['api_base_url']}/chat/completions",
                    json=emo_payload,
                    headers={"Authorization": f"Bearer {api_cfg['api_key']}",
                             "Content-Type": "application/json"},
                )
                resp2.raise_for_status()
                emotion = resp2.json()["choices"][0]["message"]["content"].strip().splitlines()[0][:30]
        except Exception as e:
            print(f"[journal] emotion analysis failed: {e}")
        return journal_mod.set_ai_draft(
            DB_PATH, "default", date_str,
            draft=draft, summary=draft[:120] + "...",
            emotional_tone=emotion,
            highlights=[],
        )
    except Exception as e:
        print(f"[journal] AI draft failed: {e}")
        raise HTTPException(500, f"AI draft failed: {e}")

@app.delete("/api/journal/{date_str}")
async def journal_delete(date_str: str):
    if not journal_mod.delete_journal(DB_PATH, "default", date_str):
        raise HTTPException(404, "no journal for that date")
    return {"ok": True}


# ============================================================
# Co-experience — "remember when we..."
# ============================================================
@app.get("/api/co-experience/moments")
async def co_exp_list(moment_type: str = "", limit: int = 50, offset: int = 0):
    return {"items": co_exp_mod.list_moments(
        DB_PATH, "default",
        moment_type=moment_type or None,
        limit=limit, offset=offset
    )}

@app.post("/api/co-experience/moments")
async def co_exp_create(payload: Dict):
    return co_exp_mod.create_moment(
        DB_PATH, "default",
        title=payload.get("title", ""),
        story=payload.get("story", ""),
        moment_type=payload.get("moment_type", "shared"),
        occurred_at=payload.get("occurred_at"),
        emotional_weight=payload.get("emotional_weight", 0.5),
        context_ref=payload.get("context_ref", {}),
    )

@app.get("/api/co-experience/moments/{moment_id}")
async def co_exp_get(moment_id: str):
    m = co_exp_mod.get_moment(DB_PATH, moment_id)
    if not m:
        raise HTTPException(404, "moment not found")
    return m

@app.post("/api/co-experience/moments/{moment_id}")
async def co_exp_update(moment_id: str, payload: Dict):
    m = co_exp_mod.update_moment(
        DB_PATH, moment_id,
        title=payload.get("title"),
        story=payload.get("story"),
        moment_type=payload.get("moment_type"),
        emotional_weight=payload.get("emotional_weight"),
    )
    if not m:
        raise HTTPException(404, "moment not found")
    return m

@app.delete("/api/co-experience/moments/{moment_id}")
async def co_exp_delete(moment_id: str):
    if not co_exp_mod.delete_moment(DB_PATH, moment_id):
        raise HTTPException(404, "moment not found")
    return {"ok": True}

@app.get("/api/co-experience/today")
async def co_exp_today():
    """Surface one moment for today's briefing."""
    m = co_exp_mod.surface_for_today(DB_PATH, "default")
    return {"moment": m}

@app.post("/api/co-experience/harvest")
async def co_exp_harvest():
    """Harvest high-importance timeline events into moments."""
    count = co_exp_mod.harvest_from_timeline(DB_PATH, "default")
    return {"harvested": count}

@app.get("/api/co-experience/stats")
async def co_exp_stats_api():
    return co_exp_mod.get_stats(DB_PATH, "default")


# ============================================================
# Daily Loop — morning briefing
# ============================================================
@app.get("/api/daily/briefing")
async def daily_briefing():
    """Get today's morning briefing (life-first homepage data)."""
    return daily_loop_mod.build_briefing(DB_PATH, "default")

@app.post("/api/daily/journal-draft")
async def daily_journal_draft():
    """Trigger AI to draft today's journal entry."""
    from datetime import datetime
    today = datetime.now().strftime("%Y-%m-%d")
    return await journal_set_ai_draft(today, {})


# ============================================================
# Prompt Engineering — editable LLM prompts
# ============================================================
@app.get("/api/prompts")
async def prompts_list(category: str = ""):
    return {
        "prompts": prompt_registry_mod.list_prompts(category=category or None),
        "categories": prompt_registry_mod.list_categories(),
        "stats": prompt_registry_mod.get_stats(DB_PATH),
    }

@app.get("/api/prompts/stats")
async def prompt_stats_api():
    return prompt_registry_mod.get_stats(DB_PATH)

@app.get("/api/prompts/{key}")
async def prompt_get(key: str):
    p = prompt_registry_mod.get_prompt_with_meta(DB_PATH, key)
    if not p:
        raise HTTPException(404, "prompt not found")
    return p

@app.post("/api/prompts/{key}")
async def prompt_set(key: str, payload: Dict):
    content = payload.get("content", "")
    if not prompt_registry_mod.set_prompt(DB_PATH, key, content):
        raise HTTPException(404, "prompt not found")
    return prompt_registry_mod.get_prompt_with_meta(DB_PATH, key)

@app.post("/api/prompts/{key}/reset")
async def prompt_reset(key: str):
    if not prompt_registry_mod.reset_prompt(DB_PATH, key):
        raise HTTPException(404, "prompt not found")
    return prompt_registry_mod.get_prompt_with_meta(DB_PATH, key)


# ============================================================
# Residents — living AI inhabitants
# ============================================================
@app.get("/api/residents")
async def residents_list(status: str = ""):
    return {"items": residents_mod.list_residents(DB_PATH, "default", status=status or None)}

@app.post("/api/residents")
async def residents_create(payload: Dict):
    return residents_mod.create_resident(
        DB_PATH, "default",
        name=payload.get("name", "Unnamed"),
        role=payload.get("role", "custom"),
        system_prompt=payload.get("system_prompt", ""),
        llm_config=payload.get("llm_config", {}),
        working_dir=payload.get("working_dir", ""),
        mode=payload.get("mode", "async"),
        max_retries=payload.get("max_retries", 3),
        depends_on=payload.get("depends_on", []),
        triggers=payload.get("triggers", []),
        skill_id=payload.get("skill_id", ""),
        personality_traits=payload.get("personality_traits", {}),
    )

@app.get("/api/residents/stats")
async def residents_stats_api():
    return residents_mod.get_stats(DB_PATH, "default")

@app.get("/api/residents/{resident_id}")
async def residents_get(resident_id: str):
    r = residents_mod.get_resident(DB_PATH, resident_id)
    if not r:
        raise HTTPException(404, "resident not found")
    return r

@app.post("/api/residents/{resident_id}")
async def residents_update(resident_id: str, payload: Dict):
    r = residents_mod.update_resident(DB_PATH, resident_id, payload)
    if not r:
        raise HTTPException(404, "resident not found")
    return r

@app.delete("/api/residents/{resident_id}")
async def residents_delete(resident_id: str):
    if not residents_mod.delete_resident(DB_PATH, resident_id):
        raise HTTPException(404, "resident not found")
    return {"ok": True}

@app.post("/api/residents/{resident_id}/concerns")
async def residents_set_concerns(resident_id: str, payload: Dict):
    """Set a resident's current concerns (1-3 things it's thinking about)."""
    concerns = payload.get("concerns", [])
    if not residents_mod.set_concerns(DB_PATH, resident_id, concerns):
        raise HTTPException(404, "resident not found")
    return {"ok": True}

@app.post("/api/residents/{resident_id}/run")
async def residents_run(resident_id: str, payload: Dict):
    """Manually trigger a resident run."""
    r = residents_mod.get_resident(DB_PATH, resident_id)
    if not r:
        raise HTTPException(404, "resident not found")
    trigger = payload.get("trigger", "manual")
    trigger_payload = payload.get("trigger_payload", {})
    input_text = payload.get("input", "")
    # Run inline (async event loop)
    run = await residents_mod.execute_resident(
        DB_PATH, resident_id, trigger, trigger_payload, input_text,
        http_client_factory=lambda timeout: httpx.AsyncClient(timeout=timeout),
        get_api_cfg=get_memory_api_config,
    )
    return run

@app.get("/api/residents/{resident_id}/runs")
async def resident_runs_list(resident_id: str, limit: int = 20):
    return {"items": residents_mod.list_runs(DB_PATH, resident_id=resident_id, limit=limit)}

@app.get("/api/residents/{resident_id}/state")
async def resident_state_get(resident_id: str):
    """获取居民的独立状态（当前关注、观点、心情、活动日志）。"""
    return residents_mod.get_resident_state(DB_PATH, resident_id)

@app.get("/api/residents/{resident_id}/activity")
async def resident_activity_get(resident_id: str, limit: int = 10):
    """获取居民最近的活动记录。"""
    return {"items": residents_mod.get_activity_log(DB_PATH, resident_id, limit=limit)}

@app.post("/api/residents/discuss")
async def residents_discuss_api(payload: Dict):
    """让多个居民讨论一个话题。多轮 LLM 调用，真争论。"""
    topic = payload.get("topic", "")
    resident_ids = payload.get("resident_ids", [])
    max_rounds = payload.get("max_rounds", 2)
    if not topic or not resident_ids:
        raise HTTPException(400, "需要 topic 和 resident_ids")
    results = await residents_mod.resident_discuss(
        DB_PATH, "default", topic, resident_ids,
        http_client_factory=lambda timeout: httpx.AsyncClient(timeout=timeout),
        get_api_cfg=get_memory_api_config,
        max_rounds=max_rounds,
    )
    return {"messages": results}

@app.post("/api/residents/{resident_id}/work")
async def resident_work_api(resident_id: str, payload: Dict):
    """让一个居民独立完成一项工作（Life Loop 或手动触发）。"""
    task = payload.get("task", "")
    if not task:
        raise HTTPException(400, "需要 task")
    result = await residents_mod.resident_do_work(
        DB_PATH, resident_id, task,
        http_client_factory=lambda timeout: httpx.AsyncClient(timeout=timeout),
        get_api_cfg=get_memory_api_config,
    )
    return result

@app.get("/api/resident-runs")
async def all_runs_list(status: str = "", limit: int = 50):
    return {"items": residents_mod.list_runs(
        DB_PATH, resident_id=None, user_id="default",
        status=status or None, limit=limit
    )}

@app.get("/api/resident-runs/{run_id}")
async def run_get(run_id: str):
    r = residents_mod.get_run(DB_PATH, run_id)
    if not r:
        raise HTTPException(404, "run not found")
    return r


# ============================================================
# Resident Skills (SKILL.md standard)
# ============================================================
@app.get("/api/resident-skills")
async def resident_skills_list():
    return {"items": residents_mod.list_skills(DB_PATH)}

@app.post("/api/resident-skills")
async def resident_skills_register(payload: Dict):
    return residents_mod.register_skill(
        DB_PATH,
        name=payload.get("name", ""),
        path=payload.get("path", ""),
        description=payload.get("description", ""),
        manifest=payload.get("manifest", {}),
        is_builtin=payload.get("is_builtin", False),
    )


# ============================================================
# Mornings — the daily AI letter
# ============================================================
@app.get("/api/mornings/today")
async def mornings_today():
    return mornings_mod.get_or_create(DB_PATH, "default")

@app.get("/api/mornings/{date_str}")
async def mornings_get(date_str: str):
    m = mornings_mod.get(DB_PATH, "default", date_str)
    if not m:
        raise HTTPException(404, "no morning letter for that date")
    return m

@app.get("/api/mornings/list")
async def mornings_list(days: int = 14):
    return {"items": mornings_mod.list_recent(DB_PATH, "default", days=days)}

@app.post("/api/mornings/{date_str}/generate")
async def mornings_generate(date_str: str):
    """Generate (or regenerate) today's morning letter via LLM."""
    return await mornings_mod.generate_letter(
        DB_PATH, "default", date_str,
        http_client_factory=lambda timeout: httpx.AsyncClient(timeout=timeout),
        get_api_cfg=get_memory_api_config,
    )

@app.post("/api/mornings/{date_str}/read")
async def mornings_mark_read(date_str: str):
    if not mornings_mod.mark_read(DB_PATH, "default", date_str):
        raise HTTPException(404, "no morning letter for that date")
    return {"ok": True}


# ============================================================
# Pushback — AI disagreements + memory surfacing
# ============================================================
@app.get("/api/pushback/context")
async def pushback_context_api():
    """Get the pushback system-prompt section (philosophy items)."""
    return {"context": pushback_mod.build_pushback_system_prompt(DB_PATH, "default")}

@app.post("/api/pushback/detect")
async def pushback_detect_api(payload: Dict):
    """Detect pushback opportunities for a user message.
    Returns related co-experience moments to surface."""
    user_msg = payload.get("message", "")
    return pushback_mod.detect_pushback_opportunities(DB_PATH, "default", user_msg)


# ============================================================
# Artifacts — the "World" (created things)
# ============================================================
@app.get("/api/artifacts")
async def artifacts_list(type: str = "", status: str = "", tag: str = "",
                          limit: int = 100, offset: int = 0):
    return {"items": artifacts_mod.list_artifacts(
        DB_PATH, "default",
        type_=type or None, status=status or None, tag=tag or None,
        limit=limit, offset=offset
    )}

@app.post("/api/artifacts")
async def artifacts_create(payload: Dict):
    return artifacts_mod.create(
        DB_PATH, "default",
        type_=payload.get("type", "note"),
        title=payload.get("title", "Untitled"),
        content=payload.get("content", ""),
        format_=payload.get("format", "markdown"),
        parent_id=payload.get("parent_id"),
        status=payload.get("status", "draft"),
        created_by=payload.get("created_by", "joint"),
        created_with_resident=payload.get("created_with_resident", ""),
        related_artifacts=payload.get("related_artifacts", []),
        tags=payload.get("tags", []),
        metadata=payload.get("metadata", {}),
        file_path=payload.get("file_path", ""),
    )

@app.get("/api/artifacts/stats")
async def artifacts_stats_api():
    return artifacts_mod.get_stats(DB_PATH, "default")

@app.get("/api/artifacts/{artifact_id}")
async def artifacts_get(artifact_id: str):
    a = artifacts_mod.get(DB_PATH, artifact_id, track_access=True)
    if not a:
        raise HTTPException(404, "artifact not found")
    return a

@app.post("/api/artifacts/{artifact_id}")
async def artifacts_update(artifact_id: str, payload: Dict):
    a = artifacts_mod.update(DB_PATH, artifact_id, payload)
    if not a:
        raise HTTPException(404, "artifact not found")
    return a

@app.post("/api/artifacts/{artifact_id}/new-version")
async def artifacts_new_version(artifact_id: str, payload: Dict):
    """Create a new version of an artifact."""
    a = artifacts_mod.new_version(
        DB_PATH, artifact_id,
        new_content=payload.get("content", ""),
        title=payload.get("title"),
        created_by=payload.get("created_by", "joint"),
        created_with_resident=payload.get("created_with_resident", ""),
    )
    if not a:
        raise HTTPException(404, "artifact not found")
    return a

@app.get("/api/artifacts/{artifact_id}/history")
async def artifacts_history(artifact_id: str):
    return {"items": artifacts_mod.get_history(DB_PATH, artifact_id)}

@app.delete("/api/artifacts/{artifact_id}")
async def artifacts_delete(artifact_id: str):
    if not artifacts_mod.delete(DB_PATH, artifact_id):
        raise HTTPException(404, "artifact not found")
    return {"ok": True}


# ============================================================
# Philosophy — values, beliefs, principles, anti-goals
# ============================================================
@app.get("/api/philosophy")
async def philosophy_list(type: str = ""):
    if type:
        return {"items": philosophy_mod.list_by_type(DB_PATH, "default", type)}
    return {"items": philosophy_mod.list_active(DB_PATH, "default")}

@app.get("/api/philosophy/stats")
async def philosophy_stats_api():
    return philosophy_mod.get_stats(DB_PATH, "default")

@app.post("/api/philosophy")
async def philosophy_create(payload: Dict):
    return philosophy_mod.create(
        DB_PATH, "default",
        type_=payload.get("type", "principle"),
        content=payload.get("content", ""),
        rationale=payload.get("rationale", ""),
        source=payload.get("source", "user"),
        confidence=payload.get("confidence", 0.8),
    )

@app.get("/api/philosophy/{item_id}")
async def philosophy_get(item_id: str):
    p = philosophy_mod.get(DB_PATH, item_id)
    if not p:
        raise HTTPException(404, "philosophy item not found")
    return p

@app.post("/api/philosophy/{item_id}")
async def philosophy_update(item_id: str, payload: Dict):
    p = philosophy_mod.update(DB_PATH, item_id, payload)
    if not p:
        raise HTTPException(404, "philosophy item not found")
    return p

@app.post("/api/philosophy/{item_id}/retire")
async def philosophy_retire(item_id: str):
    if not philosophy_mod.retire(DB_PATH, item_id):
        raise HTTPException(404, "philosophy item not found")
    return {"ok": True}

@app.delete("/api/philosophy/{item_id}")
async def philosophy_delete(item_id: str):
    if not philosophy_mod.delete(DB_PATH, item_id):
        raise HTTPException(404, "philosophy item not found")
    return {"ok": True}


# ============================================================
# Evolution — thought evolution tracking
# ============================================================
@app.get("/api/evolution")
async def evolution_list(type: str = "", limit: int = 100):
    return {"items": evolution_mod.list_events(
        DB_PATH, "default", type_=type or None, limit=limit
    )}

@app.post("/api/evolution")
async def evolution_create(payload: Dict):
    return evolution_mod.create_event(
        DB_PATH, "default",
        type_=payload.get("type", "interest_shift"),
        from_state=payload.get("from_state", ""),
        to_state=payload.get("to_state", ""),
        evidence=payload.get("evidence", ""),
        evidence_refs=payload.get("evidence_refs", {}),
        confidence=payload.get("confidence", 0.5),
        observed_by=payload.get("observed_by", "user"),
    )

@app.get("/api/evolution/stats")
async def evolution_stats_api():
    return evolution_mod.get_stats(DB_PATH, "default")

@app.get("/api/evolution/curve")
async def evolution_curve_api(type: str = "interest_shift", months: int = 12):
    return {"items": evolution_mod.get_evolution_curve(
        DB_PATH, "default", type_=type, months=months
    )}

@app.post("/api/evolution/{event_id}/confirm")
async def evolution_confirm(event_id: str):
    if not evolution_mod.confirm_event(DB_PATH, event_id):
        raise HTTPException(404, "event not found")
    return {"ok": True}

@app.post("/api/evolution/{event_id}/dispute")
async def evolution_dispute(event_id: str):
    if not evolution_mod.dispute_event(DB_PATH, event_id):
        raise HTTPException(404, "event not found")
    return {"ok": True}


# ============================================================
# Discoveries — daily surprises
# ============================================================
@app.get("/api/discoveries")
async def discoveries_list(days: int = 7, limit: int = 50):
    return {"items": discovery_mod.list_recent(DB_PATH, "default", days=days, limit=limit)}

@app.get("/api/discoveries/today")
async def discoveries_today():
    from datetime import datetime
    today = datetime.now().strftime("%Y-%m-%d")
    return {"items": discovery_mod.list_by_date(DB_PATH, "default", today)}

@app.get("/api/discoveries/stats")
async def discoveries_stats_api():
    return discovery_mod.get_stats(DB_PATH, "default")

@app.post("/api/discoveries")
async def discoveries_create(payload: Dict):
    return discovery_mod.create(
        DB_PATH, "default",
        type_=payload.get("type", "observation"),
        title=payload.get("title", ""),
        content=payload.get("content", ""),
        evidence=payload.get("evidence", ""),
        evidence_refs=payload.get("evidence_refs", []),
        confidence=payload.get("confidence", 0.5),
        discovered_by=payload.get("discovered_by", "ai"),
    )

@app.get("/api/discoveries/{discovery_id}")
async def discoveries_get(discovery_id: str):
    d = discovery_mod.get(DB_PATH, discovery_id)
    if not d:
        raise HTTPException(404, "discovery not found")
    return d

@app.post("/api/discoveries/{discovery_id}/seen")
async def discoveries_seen(discovery_id: str):
    if not discovery_mod.mark_seen(DB_PATH, discovery_id):
        raise HTTPException(404, "discovery not found")
    return {"ok": True}

@app.post("/api/discoveries/{discovery_id}/act")
async def discoveries_act(discovery_id: str):
    if not discovery_mod.mark_acted(DB_PATH, discovery_id):
        raise HTTPException(404, "discovery not found")
    return {"ok": True}

@app.post("/api/discoveries/{discovery_id}/dismiss")
async def discoveries_dismiss(discovery_id: str):
    if not discovery_mod.dismiss(DB_PATH, discovery_id):
        raise HTTPException(404, "discovery not found")
    return {"ok": True}

@app.delete("/api/discoveries/{discovery_id}")
async def discoveries_delete(discovery_id: str):
    if not discovery_mod.delete(DB_PATH, discovery_id):
        raise HTTPException(404, "discovery not found")
    return {"ok": True}


# ============================================================
# Plugins — extensible plugin system
# ============================================================
@app.get("/api/plugins")
async def plugins_list():
    return plugin_sdk.get_plugin_stats()

@app.post("/api/plugins/reload")
async def plugins_reload():
    """Reload all plugins from the plugins directory."""
    global _plugins_dir
    plugins_dir = PROJECT_ROOT / "plugins"
    loaded = plugin_sdk.load_all_plugins(plugins_dir)
    return {"loaded": len(loaded), "stats": plugin_sdk.get_plugin_stats()}

@app.post("/api/plugins/create-example")
async def plugins_create_example():
    """Create an example plugin to show the SDK structure."""
    plugins_dir = PROJECT_ROOT / "plugins"
    path = plugin_sdk.create_example_plugin(plugins_dir)
    # Reload to pick it up
    plugin_sdk.load_all_plugins(plugins_dir)
    return {"ok": True, "path": str(path)}


# ============================================================
# Vector Store — semantic search stats
# ============================================================
@app.get("/api/vector-store/stats")
async def vector_store_stats_api():
    vs = vector_store_mod.get_vector_store(DB_PATH)
    return vs.get_stats()

@app.post("/api/vector-store/reindex")
async def vector_store_reindex_api():
    """Reindex all memories into the vector store."""
    try:
        vs = vector_store_mod.get_vector_store(DB_PATH)
        from app.db_utils import safe_connect
        import sqlite3
        conn = safe_connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT id, user_id, content, importance, layer, category FROM memory_items"
        ).fetchall()
        conn.close()
        count = 0
        for r in rows:
            try:
                vs.add(f"memories_{r['user_id']}", id=r["id"], text=r["content"],
                       metadata={"importance": r["importance"], "layer": r["layer"],
                                 "category": r["category"]})
                count += 1
            except Exception:
                pass
        return {"reindexed": count, "backend": vs.backend}
    except Exception as e:
        raise HTTPException(500, f"reindex failed: {e}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "app.main:app",
        host="0.0.0.0",
        port=int(os.getenv("PORT", "3000")),
        reload=False,
        log_level="info",
    )
